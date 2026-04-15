# -*- coding: utf-8 -*-
#############################################################################
#
#    Partenaire Succes Pvt. Ltd.
#
#    Copyright (C) 2025-TODAY Partenaire Succes(<https://www.partenairesucces.com/>)
#    Author: Adama KONE
#
#############################################################################
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class ImportLoyaltyPointsWizard(models.TransientModel):
    _name = 'import.loyalty.points.wizard'
    _description = 'Import Excel - Mise à jour Points Fidélité'

    file = fields.Binary(string="Fichier Excel", required=False)
    file_name = fields.Char()
    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('loaded', 'Chargé'),
        ('done', 'Terminé'),
    ], default='draft')
    line_ids = fields.One2many(
        'import.loyalty.points.line',
        'wizard_id',
        string='Lignes'
    )
    count_ok = fields.Integer(string='Lignes valides', compute='_compute_counts')
    count_errors = fields.Integer(string='Lignes en erreur', compute='_compute_counts')

    @api.depends('line_ids.status')
    def _compute_counts(self):
        for wizard in self:
            wizard.count_ok = len(wizard.line_ids.filtered(lambda l: l.status == 'ok'))
            wizard.count_errors = len(wizard.line_ids.filtered(lambda l: l.status != 'ok'))

    # -------------------------------------------------------------------------
    # TÉLÉCHARGER LE TEMPLATE EXCEL
    # -------------------------------------------------------------------------
    def action_download_template(self):
        """Génère et télécharge un fichier Excel modèle pour l'import des points."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl est requise pour générer le template."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Import Points Fidélité"

        # En-têtes
        headers = ['code_carte', 'points']
        ws.append(headers)

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Ligne exemple
        ws.append(['044XXXXXXXXXXXXXXXXX', 150.0])

        # Largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'template_import_points_fidelite.xlsx',
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'self',
        }

    # -------------------------------------------------------------------------
    # CHARGER LE FICHIER
    # -------------------------------------------------------------------------
    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez charger un fichier Excel."))

        # Supprimer les anciennes lignes
        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        file_data = io.BytesIO(decoded_file)

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_data)
        except Exception:
            raise UserError(_(
                "Impossible de lire le fichier. "
                "Vérifiez qu'il s'agit bien d'un fichier Excel (.xlsx)."
            ))

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        # Lecture des en-têtes (première ligne)
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]

        required_columns = ['code_carte', 'points']
        for col in required_columns:
            if col not in headers:
                raise UserError(
                    _("Colonne manquante : '%s'\nColonnes trouvées : %s") % (
                        col, ', '.join(headers)
                    )
                )

        idx_code = headers.index('code_carte')
        idx_points = headers.index('points')

        lines_vals = []

        for row_num, row in enumerate(rows[1:], start=2):
            # Ignorer les lignes vides
            if not row or row[idx_code] is None:
                continue

            # Normalisation : Excel lit les codes numériques comme float (ex: 1012 → 1012.0)
            raw_code = row[idx_code]
            if isinstance(raw_code, float) and raw_code == int(raw_code):
                code_carte = str(int(raw_code))
            else:
                code_carte = str(raw_code).strip()
            if not code_carte:
                continue

            try:
                points_new = float(row[idx_points] or 0.0)
            except (ValueError, TypeError):
                points_new = 0.0

            # Recherche de la carte de fidélité par son code
            card = self.env['loyalty.card'].search(
                [('code', '=', code_carte)],
                limit=1
            )

            if not card:
                lines_vals.append((0, 0, {
                    'code_carte': code_carte,
                    'card_id': False,
                    'partner_id': False,
                    'program_id': False,
                    'points_old': 0.0,
                    'points_new': points_new,
                    'status': 'card_not_found',
                }))
                continue

            lines_vals.append((0, 0, {
                'code_carte': code_carte,
                'card_id': card.id,
                'partner_id': card.partner_id.id if card.partner_id else False,
                'program_id': card.program_id.id if card.program_id else False,
                'points_old': card.points,
                'points_new': points_new,
                'status': 'ok',
            }))

        if not lines_vals:
            raise UserError(_(
                "Aucune ligne de données trouvée dans le fichier (hors en-têtes)."
            ))

        self.write({
            'line_ids': lines_vals,
            'state': 'loaded',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # -------------------------------------------------------------------------
    # CONFIRMER LA MISE À JOUR
    # -------------------------------------------------------------------------
    def action_confirm(self):
        self.ensure_one()

        valid_lines = self.line_ids.filtered(lambda l: l.status == 'ok')

        if not valid_lines:
            raise UserError(_(
                "Aucune ligne valide à traiter. "
                "Vérifiez les erreurs dans le tableau."
            ))

        for line in valid_lines:
            line.card_id.write({'points': line.points_new})

        self.write({'state': 'done'})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Import terminé"),
                'message': _(
                    "%d carte(s) de fidélité mise(s) à jour avec succès."
                ) % len(valid_lines),
                'type': 'success',
                'sticky': False,
            },
        }


class ImportLoyaltyPointsLine(models.TransientModel):
    _name = 'import.loyalty.points.line'
    _description = 'Ligne import points fidélité'

    wizard_id = fields.Many2one('import.loyalty.points.wizard', ondelete='cascade')
    code_carte = fields.Char(string='Code Carte (Excel)', readonly=True)
    card_id = fields.Many2one('loyalty.card', string='Carte trouvée', readonly=True)
    partner_id = fields.Many2one('res.partner', string='Partenaire', readonly=True)
    program_id = fields.Many2one('loyalty.program', string='Programme', readonly=True)
    points_old = fields.Float(string='Solde actuel', readonly=True)
    points_new = fields.Float(string='Nouveau solde', readonly=True)
    status = fields.Selection([
        ('ok', 'Prêt'),
        ('card_not_found', 'Carte non trouvée'),
    ], string='Statut', readonly=True)
