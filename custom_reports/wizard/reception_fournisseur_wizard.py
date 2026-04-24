# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class ReceptionFournisseurWizard(models.TransientModel):
    _name = 'reception.fournisseur.wizard'
    _description = 'Wizard Rapport Réceptions Fournisseurs'

    date_debut = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today
    )
    date_fin = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company
    )
    fournisseur_ids = fields.Many2many(
        'res.partner',
        string='Fournisseurs',
        domain=[('supplier_rank', '>', 0)],
        help='Laissez vide pour inclure tous les fournisseurs'
    )
    currency_id = fields.Many2one(
        'res.currency',
        string='Devise',
        default=lambda self: self.env.company.currency_id,
    )

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for record in self:
            if record.date_debut > record.date_fin:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    def _get_pickings_data(self):
        """
        Récupère TOUS les pickings entrants validés (réceptions par commande ET
        réceptions directes) et calcule les montants réels ligne par ligne :
        montant = qty_done × price_unit du move.
        """
        self.ensure_one()

        domain = [
            ('picking_type_code', '=', 'incoming'),
            ('state', '=', 'done'),
            ('date_done', '>=', self.date_debut),
            ('date_done', '<=', self.date_fin),
            ('company_id', '=', self.company_id.id),
        ]
        if self.fournisseur_ids:
            domain.append(('partner_id', 'in', self.fournisseur_ids.ids))

        pickings = self.env['stock.picking'].search(domain, order='partner_id, date_done')

        if not pickings:
            raise UserError("Aucune réception trouvée pour la période sélectionnée.")

        # Grouper par fournisseur (clé = partner.id ou 0 si pas de fournisseur)
        fournisseurs_data = {}

        for picking in pickings:
            partner = picking.partner_id
            key = partner.id if partner else 0

            if key not in fournisseurs_data:
                fournisseurs_data[key] = {
                    'partner': partner,
                    'receptions': [],
                    'total': 0.0,
                }

            # Calcul du montant réel de ce picking : somme(qty_done × price_unit)
            total_picking = 0.0
            for move in picking.move_ids.filtered(lambda m: m.state == 'done'):
                qty_done = (
                    sum(move.move_line_ids.mapped('quantity'))
                    or move.product_uom_qty
                )
                total_picking += qty_done * (move.price_unit or 0.0)

            fournisseurs_data[key]['receptions'].append({
                'type': picking.picking_type_id.name[:3].upper() if picking.picking_type_id else 'REC',
                'date': picking.date_done,
                'numero': picking.name,
                'ref': picking.origin or '',
                'total': total_picking,
            })
            fournisseurs_data[key]['total'] += total_picking

        return list(fournisseurs_data.values())

    def action_print_report(self):
        self.ensure_one()
        self._get_pickings_data()   # lève UserError si vide
        return self.env.ref('custom_reports.action_report_reception_fournisseur').report_action(self)

    # ── Export Excel ──────────────────────────────────────────────────────────
    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        fournisseurs_data = self._get_pickings_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Réceptions Fournisseurs"

        # ── Styles ────────────────────────────────────────────────────────────
        BLUE   = "1A5276"
        LBLUE  = "D6EAF8"
        WHITE  = "FFFFFF"
        GRAY   = "F0F0F0"
        thin   = Side(style='thin', color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_font(bold=True):
            return Font(name="Arial", bold=bold, color=WHITE, size=10)

        def cell_font(bold=False):
            return Font(name="Arial", bold=bold, size=9)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        # ── Ligne 1 : société ─────────────────────────────────────────────────
        ws.merge_cells("A1:F1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        # ── Ligne 2 : titre ───────────────────────────────────────────────────
        ws.merge_cells("A2:F2")
        ws["A2"] = (
            f"RÉCAPITULATIF RÉCEPTIONS FOURNISSEURS — "
            f"Du {self.date_debut.strftime('%d/%m/%Y')} au {self.date_fin.strftime('%d/%m/%Y')}"
        )
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Ligne 3 : vide ────────────────────────────────────────────────────
        ws.append([])

        # ── En-têtes colonnes ─────────────────────────────────────────────────
        headers = ["Type Mvt", "Date", "N° Réception", "Fournisseur", "N° Fact / BL", "Montant Total"]
        ws.append(headers)
        hdr_row = ws.max_row
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=hdr_row, column=col)
            cell.font = hdr_font()
            cell.fill = fill(BLUE)
            cell.alignment = aln("center")
            cell.border = border

        # ── Données ───────────────────────────────────────────────────────────
        grand_total = 0.0
        for fdata in fournisseurs_data:
            partner_name = fdata['partner'].name if fdata['partner'] else '—'
            for reception in fdata['receptions']:
                row = [
                    reception['type'],
                    reception['date'].strftime('%d/%m/%Y') if reception['date'] else '',
                    reception['numero'],
                    partner_name,
                    reception['ref'],
                    reception['total'],
                ]
                ws.append(row)
                r = ws.max_row
                for col in range(1, 7):
                    c = ws.cell(row=r, column=col)
                    c.font = cell_font()
                    c.border = border
                    c.alignment = aln("right" if col == 6 else "center" if col in (1, 2) else "left")
                ws.cell(row=r, column=6).number_format = '#,##0'

            # Sous-total fournisseur
            ws.append(["", "", "", "", "Sous-total " + partner_name, fdata['total']])
            r = ws.max_row
            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE)
                c.border = border
                c.alignment = aln("right" if col >= 5 else "left")
            ws.cell(row=r, column=6).number_format = '#,##0'
            grand_total += fdata['total']

        # ── Total général ─────────────────────────────────────────────────────
        ws.append(["", "", "", "", "TOTAL GÉNÉRAL", grand_total])
        r = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = border
            c.alignment = aln("right" if col >= 5 else "left")
        ws.cell(row=r, column=6).number_format = '#,##0'

        # ── Largeurs colonnes ─────────────────────────────────────────────────
        for col, width in enumerate([10, 13, 18, 28, 22, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # ── Sauvegarde ────────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        filename = f"Receptions_Fournisseurs_{self.date_debut.strftime('%d%m%Y')}_{self.date_fin.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })

        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
