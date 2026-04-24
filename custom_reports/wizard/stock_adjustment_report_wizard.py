import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class StockAdjustmentReportWizard(models.TransientModel):
    _name = 'stock.adjustment.report.wizard'
    _description = 'Ajustement de stock'

    date_from = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today
    )
    date_to = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company
    )

    scrap_lines_ids = fields.Many2many(
        'stock.scrap',
        string='Rébuts',
        compute='_compute_scrap_lines_ids'
    )

    @api.depends('date_from', 'date_to', 'company_id')
    def _compute_scrap_lines_ids(self):
        for record in self:
            record.scrap_lines_ids = self.env['stock.scrap'].search([
                ('date_done', '>=', record.date_from),
                ('date_done', '<=', record.date_to),
                ('company_id', '=', record.company_id.id),
                ('state', '=', 'done')
            ], order='date_done, name')

    def action_print_report(self):
        self.ensure_one()
        if not self.scrap_lines_ids:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")

        return self.env.ref('custom_reports.action_report_stock_adjustment').report_action(self)

    def get_grouped_data(self):
        """Grouper les rebuts par type d'ajustement et raison"""
        self.ensure_one()

        grouped_data = {}

        for scrap in self.scrap_lines_ids:
            # Déterminer le type d'ajustement
            adjustment_type = self._get_adjustment_type(scrap)

            if adjustment_type not in grouped_data:
                grouped_data[adjustment_type] = {}

            # Grouper par raison
            reasons = scrap.scrap_reason_tag_ids
            if not reasons:
                reason_key = 'Sans raison'
                reason_name = 'Sans raison'
            else:
                # Utiliser toutes les raisons associées
                reason_key = ','.join(reasons.mapped('name'))
                reason_name = ', '.join(reasons.mapped('name'))

            if reason_key not in grouped_data[adjustment_type]:
                grouped_data[adjustment_type][reason_key] = {
                    'reason_name': reason_name,
                    'lines': []
                }

            grouped_data[adjustment_type][reason_key]['lines'].append({
                'document': scrap.name,
                'date': scrap.date_done,
                'cashier': scrap.create_uid.name,
                'article': scrap.product_id.code_article or '',
                'designation': scrap.product_id.name,
                'quantity': scrap.scrap_qty,
                'total_pa': scrap.scrap_qty * scrap.product_id.standard_price,
            })

        return grouped_data

    def _get_adjustment_type(self, scrap):
        """Déterminer le type d'ajustement basé sur la localisation ou d'autres critères"""
        # Logique pour déterminer le type (AUTOCONSOMMATION, CASSE, etc.)
        # À adapter selon vos besoins
        if scrap.location_id.usage == 'internal':
            return 'AUTOCONSOMMATION'
        else:
            return 'CASSE'

    def get_totals_by_type(self, grouped_data):
        """Calculer les totaux par type d'ajustement"""
        totals = {}
        for adj_type, reasons in grouped_data.items():
            total_qty = 0
            total_pa = 0
            for reason_data in reasons.values():
                for line in reason_data['lines']:
                    total_qty += line['quantity']
                    total_pa += line['total_pa']
            totals[adj_type] = {
                'quantity': total_qty,
                'total_pa': total_pa
            }
        return totals

    def get_grand_totals(self, totals):
        grand_total_qty = sum(t['quantity'] for t in totals.values())
        grand_total_pa = sum(t['total_pa'] for t in totals.values())
        return {'quantity': grand_total_qty, 'total_pa': grand_total_pa}

    def action_export_excel(self):
        self.ensure_one()
        if not self.scrap_lines_ids:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        wb = Workbook(); ws = wb.active; ws.title = "Ajustements Stock"
        BLUE = "1A5276"; LBLUE = "D6EAF8"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        ws.merge_cells("A1:I1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)
        ws.merge_cells("A2:I2")
        ws["A2"] = f"AJUSTEMENTS STOCK — Du {self.date_from.strftime('%d/%m/%Y')} au {self.date_to.strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE); ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18; ws.append([])

        headers = ["Type", "Raison", "N° Document", "Date", "Opérateur", "Code Article", "Désignation", "Quantité", "Total PA"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 10):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        grouped = self.get_grouped_data()
        grand_qty = grand_pa = 0.0
        for adj_type, reasons in grouped.items():
            for reason_data in reasons.values():
                for line in reason_data['lines']:
                    ws.append([adj_type, reason_data['reason_name'], line['document'],
                                line['date'].strftime('%d/%m/%Y') if line['date'] else '',
                                line['cashier'], line['article'], line['designation'],
                                line['quantity'], line['total_pa']])
                    r = ws.max_row
                    for col in range(1, 10):
                        c = ws.cell(row=r, column=col)
                        c.font = Font(name="Arial", size=9); c.border = brd
                        c.alignment = aln("right" if col >= 8 else "center" if col == 4 else "left")
                    ws.cell(row=r, column=9).number_format = '#,##0.00'
                    grand_qty += line['quantity']; grand_pa += line['total_pa']

        ws.append(["", "", "", "", "", "", "", "TOTAL GÉNÉRAL", grand_pa])
        r = ws.max_row
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.border = brd
            c.alignment = aln("right" if col >= 8 else "left")
        ws.cell(row=r, column=9).number_format = '#,##0.00'

        for col, width in enumerate([18, 20, 16, 13, 18, 14, 28, 10, 14], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = f"Ajustements_Stock_{self.date_from.strftime('%d%m%Y')}_{self.date_to.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename, 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}



