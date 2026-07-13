# -*- coding: utf-8 -*-
import io
import base64
from collections import defaultdict
from datetime import datetime as dt, time as t
from odoo import models, fields, api
from odoo.exceptions import UserError
from odoo.tools import float_round


class StockValoriseReport(models.TransientModel):
    _name = 'stock.valorise.report'
    _description = 'Rapport de Stock Valorisé'

    date_report = fields.Datetime(
        string='Date de valorisation',
        required=True,
        default=lambda self: dt.combine(fields.Date.context_today(self), t(23, 59, 59))
    )

    location_id = fields.Many2one(
        'stock.location',
        string='Emplacement',
        required=True,
        domain=[('usage', '=', 'internal')]
    )

    category_ids = fields.Many2many(
        'product.category',
        string='Catégories'
    )

    cat_gestion_ids = fields.Many2many(
        'product.category.x3',
        string='Niveau 5'
    )

    product_ids = fields.Many2many(
        'product.product',
        string='Produits',
        compute='_compute_product_ids',
    )

    company_id = fields.Many2one(
        'res.company',
        string='Société',
        default=lambda self: self.env.company,
    )


    @api.depends('company_id', 'cat_gestion_ids')
    def _compute_product_ids(self):
        """Détermine les produits concernés par la société et les catégories choisies."""
        for record in self:
            domain = [
                ('product_tmpl_id.allowed_company_ids', 'in', [record.company_id.id]),
                ('product_tmpl_id.type', '=', 'consu'),
                ('product_tmpl_id.prod_type_x3_id.name', '=', 'TS'),
            ]
            if record.cat_gestion_ids:
                domain.append(('cat_gestion_id', 'in', record.cat_gestion_ids.ids))

            record.product_ids = self.env['product.product'].search(domain)

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)

        if 'location_id' in fields_list and not res.get('location_id'):
            location = self.env['stock.location'].search([
                ('usage', '=', 'internal'),
                '|',
                ('name', 'ilike', 'Stock'),
                ('complete_name', 'ilike', '/Stock')
            ], limit=1)

            if not location:
                warehouse = self.env['stock.warehouse'].search([
                    ('company_id', '=', self.env.company.id)
                ], limit=1)
                if warehouse:
                    location = warehouse.lot_stock_id

            if not location:
                location = self.env['stock.location'].search([
                    ('usage', '=', 'internal')
                ], limit=1)

            if location:
                res['location_id'] = location.id

        return res

    # -------------------------------------------------------------------------
    # LOGIQUE DU RAPPORT
    # -------------------------------------------------------------------------
    def _compute_avco_at_date(self):
        """Coût moyen pondéré (AVCO) réel par produit à la date du rapport.

        On n'utilise PAS product.avg_cost avec le contexte to_date : ce champ
        (stock_account/models/product.py::_compute_value) divise la valeur
        historique rejouée par product.qty_available, qui est recalculée sur
        un périmètre quants/emplacements différent de celui du rejeu de
        mouvements — les deux quantités "à la date" peuvent diverger et donc
        fausser le ratio.
        On appelle directement _run_average_batch / _run_fifo_batch (le
        moteur interne utilisé par avg_cost), qui rejoue les mouvements
        is_in/is_out avec exactement le même algorithme que
        stock_account.stock_avco_report, pour garantir un résultat cohérent
        avec ce rapport de justification.
        """
        self.ensure_one()
        avco_by_product_id = {}
        if not self.product_ids:
            return avco_by_product_id

        products = self.product_ids.with_company(self.company_id).with_context(
            allowed_company_ids=self.company_id.ids
        )
        products_by_cost_method = defaultdict(lambda: self.env['product.product'])
        for product in products:
            products_by_cost_method[product.cost_method] |= product

        for cost_method, method_products in products_by_cost_method.items():
            if cost_method == 'average':
                std_prices, __ = method_products._run_average_batch(at_date=self.date_report)
            elif cost_method == 'fifo':
                std_prices, __ = method_products._run_fifo_batch(at_date=self.date_report)
            else:
                std_prices = {p.id: p.standard_price for p in method_products}
            avco_by_product_id.update(std_prices)

        return avco_by_product_id

    def _get_stock_by_category(self):
        """Retourne les données de valorisation à la date du rapport."""
        self.ensure_one()
        categories = {}
        total_articles = 0
        total_qty = 0.0
        total_valorisation = 0.0

        # Quantités à l'emplacement en batch
        # ANCIEN CODE : to_date ignoré par Odoo 19 (qty_available utilise stock.quant, pas stock.move.line)
        # loc_qtys = self.product_ids.with_context(
        #     location=self.location_id.id, to_date=self.date_report
        # ).mapped('qty_available')
        # qty_loc = {p.id: q for p, q in zip(self.product_ids, loc_qtys)}
        self.env.cr.execute("""
            SELECT product_id,
                   SUM(CASE WHEN location_dest_id = %s THEN quantity ELSE -quantity END) AS qty
            FROM stock_move_line
            WHERE product_id = ANY(%s)
              AND state = 'done'
              AND date <= %s
              AND (location_id = %s OR location_dest_id = %s)
            GROUP BY product_id
        """, [self.location_id.id, list(self.product_ids.ids), self.date_report, self.location_id.id, self.location_id.id])
        qty_loc = {row[0]: row[1] for row in self.env.cr.fetchall()}

        avco_by_product_id = self._compute_avco_at_date()

        for product in self.product_ids:
            categ = product.cat_gestion_id
            cat_gestion_id = categ.id
            code_article = product.code_article or product.product_tmpl_id.code_article or ''

            qty = qty_loc.get(product.id, 0.0)
            if not qty or qty <= 0:
                continue

            pamp = float_round(avco_by_product_id.get(product.id, 0.0) or 0.0, 2)

            valorisation = float_round(qty * pamp, 2)

            if cat_gestion_id not in categories:
                categories[cat_gestion_id] = {
                    'category': categ.name,
                    'category_code': categ.description or categ.name,
                    'products': [],
                    'total': 0.0,
                }

            categories[cat_gestion_id]['products'].append({
                'code_article': code_article,
                'default_code': product.default_code or '',
                'name': product.name,
                'category_code': categ.description,
                'qty': float_round(qty, 2),
                'pamp': float_round(pamp, 2),
                'valorisation': valorisation,
            })

            categories[cat_gestion_id]['total'] += valorisation
            total_articles += 1
            total_qty += qty
            total_valorisation += valorisation

        pamp_moyen = float_round(total_valorisation / total_qty, 2) if total_qty else 0.0

        return {
            'categories': sorted(categories.values(), key=lambda c: c['category']),
            'total_articles': total_articles,
            'total_qty': float_round(total_qty, 2),
            'pamp_moyen': pamp_moyen,
            'total_valorisation': float_round(total_valorisation, 2),
        }

    # -------------------------------------------------------------------------
    # ACTION DE RAPPORT
    # -------------------------------------------------------------------------


    def action_print_report(self):
        """Générer le rapport PDF et fermer le wizard"""
        self.ensure_one()
        report_action = self.env.ref('custom_reports.action_report_stock_valorise').report_action(self)
        report_action['close_on_report_download'] = True
        return report_action

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self._get_stock_by_category()
        if not data['categories']:
            raise UserError("Aucune donnée de stock trouvée.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Valorisé"

        BLUE = "1A5276"; LBLUE = "D6EAF8"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        ws.merge_cells("A1:G1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        ws.merge_cells("A2:G2")
        ws["A2"] = f"STOCK VALORISÉ — {self.date_report.strftime('%d/%m/%Y')} — {self.location_id.complete_name}"
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        headers = ["Code Article", "Réf. Interne", "Désignation", "Catégorie", "Qté", "PAMP", "Valorisation"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        for cat in data['categories']:
            for p in cat['products']:
                ws.append([p['code_article'], p['default_code'], p['name'],
                            cat['category'], p['qty'], p['pamp'], p['valorisation']])
                r = ws.max_row
                for col in range(1, 8):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9); c.border = brd
                    c.alignment = aln("right" if col >= 5 else "left")
                for col in (5, 6, 7):
                    ws.cell(row=r, column=col).number_format = '#,##0.00'

            ws.append(["", "", "", "Sous-total " + cat['category'], "", "", cat['total']])
            r = ws.max_row
            for col in range(1, 8):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE); c.border = brd
                c.alignment = aln("right" if col >= 6 else "left")
            ws.cell(row=r, column=7).number_format = '#,##0.00'

        ws.append(["", "", "", "", "", "TOTAL GÉNÉRAL", data['total_valorisation']])
        r = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.border = brd
            c.alignment = aln("right" if col >= 6 else "left")
        ws.cell(row=r, column=7).number_format = '#,##0.00'

        for col, width in enumerate([14, 14, 30, 22, 10, 14, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = f"Stock_Valorise_{self.date_report.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename, 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}





