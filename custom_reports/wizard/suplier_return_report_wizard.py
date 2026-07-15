# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class SupplierReturnReportWizard(models.TransientModel):
    _name = 'supplier.return.report.wizard'
    _description = 'Assistant Rapport Retours Fournisseurs'

    date_from = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today
    )
    date_to = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today
    )
    company_ids = fields.Many2many(
        'res.company',
        string='Sociétés',
        required=True,
        default=lambda self: self.env.company,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        compute='_compute_company_id',
        help="Première société sélectionnée, utilisée pour l'en-tête du document.",
    )
    partner_ids = fields.Many2many(
        'res.partner',
        string='Fournisseurs',
        domain=[('supplier_rank', '>', 0)],
        help='Laissez vide pour tous les fournisseurs'
    )

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    @api.depends('company_ids')
    def _compute_company_id(self):
        for record in self:
            record.company_id = record.company_ids[:1]

    def action_print_report(self):
        self.ensure_one()
        self._get_report_data()  # lève UserError si vide
        return self.env.ref('custom_reports.action_report_supplier_returns').report_action(self)

    def _get_domain(self):
        domain = [
            ('location_dest_id.usage', '=', 'supplier'),
            ('state', '=', 'done'),
            ('date_done', '>=', self.date_from),
            ('date_done', '<=', self.date_to),
            ('company_id', 'in', self.company_ids.ids),
        ]
        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))
        return domain

    def _get_report_data(self):
        pickings = self.env['stock.picking'].search(
            self._get_domain(), order='partner_id, date_done'
        )
        if not pickings:
            raise UserError("Aucun retour fournisseur trouvé pour la période sélectionnée.")

        data_by_supplier = {}
        for picking in pickings:
            supplier = picking.partner_id
            key = supplier.id if supplier else 0
            if key not in data_by_supplier:
                data_by_supplier[key] = {
                    'supplier_name': supplier.name if supplier else '—',
                    'returns': [],
                    'total_amount': 0.0,
                }

            amount = 0.0
            for move in picking.move_ids.filtered(lambda m: m.state == 'done'):
                qty_done = (
                    sum(move.move_line_ids.mapped('quantity'))
                    or move.product_uom_qty
                )
                prix = (
                    move.price_unit
                    or move.origin_returned_move_id.price_unit
                    or move.product_id.standard_price
                )
                amount += qty_done * prix

            data_by_supplier[key]['returns'].append({
                'date': picking.date_done,
                'reference': picking.name,
                'origin': picking.origin or '',
                'amount': amount,
            })
            data_by_supplier[key]['total_amount'] += amount

        return {
            'date_from': self.date_from,
            'date_to': self.date_to,
            'company': self.company_id,
            'suppliers': list(data_by_supplier.values()),
            'grand_total': sum(s['total_amount'] for s in data_by_supplier.values()),
        }

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self._get_report_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Retours Fournisseurs"

        BLUE  = "1A5276"
        LBLUE = "D6EAF8"
        WHITE = "FFFFFF"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(h):
            return PatternFill("solid", fgColor=h)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        ws.merge_cells("A1:F1")
        ws["A1"] = ', '.join(self.company_ids.mapped('name'))
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        ws.merge_cells("A2:F2")
        ws["A2"] = (
            f"RÉCAPITULATIF RETOURS FOURNISSEURS — "
            f"Du {self.date_from.strftime('%d/%m/%Y')} au {self.date_to.strftime('%d/%m/%Y')}"
        )
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        headers = ["Type Mvt", "Date", "N° Retour", "Fournisseur", "Référence", "Montant Total"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.alignment = aln("center")
            c.border = brd

        for supplier in data['suppliers']:
            for ret in supplier['returns']:
                ws.append([
                    'BRF',
                    ret['date'].strftime('%d/%m/%Y') if ret['date'] else '',
                    ret['reference'],
                    supplier['supplier_name'],
                    ret['origin'],
                    ret['amount'],
                ])
                r = ws.max_row
                for col in range(1, 7):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9)
                    c.border = brd
                    c.alignment = aln("right" if col == 6 else "center" if col in (1, 2) else "left")
                ws.cell(row=r, column=6).number_format = '#,##0'

            ws.append(["", "", "", "", "Sous-total " + supplier['supplier_name'], supplier['total_amount']])
            r = ws.max_row
            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE)
                c.border = brd
                c.alignment = aln("right" if col >= 5 else "left")
            ws.cell(row=r, column=6).number_format = '#,##0'

        ws.append(["", "", "", "", "TOTAL GÉNÉRAL", data['grand_total']])
        r = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = brd
            c.alignment = aln("right" if col >= 5 else "left")
        ws.cell(row=r, column=6).number_format = '#,##0'

        for col, width in enumerate([10, 13, 18, 28, 22, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        filename = f"Retours_Fournisseurs_{self.date_from.strftime('%d%m%Y')}_{self.date_to.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
