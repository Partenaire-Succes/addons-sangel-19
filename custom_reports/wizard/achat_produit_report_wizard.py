# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class AchatProduitReportWizard(models.TransientModel):
    _name = 'achat.produit.report.wizard'
    _description = 'Rapport Commandes & Réceptions par Produit'

    date_debut = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today,
    )
    date_fin = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )
    fournisseur_ids = fields.Many2many(
        'res.partner',
        string='Fournisseurs',
        domain=[('supplier_rank', '>', 0)],
        help='Laissez vide pour inclure tous les fournisseurs',
    )

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for rec in self:
            if rec.date_debut > rec.date_fin:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    def _get_report_data(self):
        """
        Retourne pour chaque produit :
          - les totaux (header)
          - des sous-lignes (une par couple commande/réception liés)
        """
        self.ensure_one()
        from datetime import datetime as _dt

        # ── Lignes de commandes d'achat ───────────────────────────────────────
        po_domain = [
            ('order_id.state', 'in', ['purchase', 'done']),
            ('order_id.date_order', '>=', str(self.date_debut) + ' 00:00:00'),
            ('order_id.date_order', '<=', str(self.date_fin) + ' 23:59:59'),
            ('order_id.company_id', '=', self.company_id.id),
        ]
        if self.fournisseur_ids:
            po_domain.append(('order_id.partner_id', 'in', self.fournisseur_ids.ids))
        po_lines = self.env['purchase.order.line'].search(po_domain)

        # ── Mouvements réceptions liés aux lignes PO (par date de réception) ──
        moves_by_po_line = {}
        if po_lines:
            linked_moves = self.env['stock.move'].search([
                ('purchase_line_id', 'in', po_lines.ids),
                ('state', '=', 'done'),
                ('picking_id.picking_type_code', '=', 'incoming'),
                ('picking_id.date_done', '>=', self.date_debut),
                ('picking_id.date_done', '<=', self.date_fin),
                ('location_id.usage', 'in', ['supplier', 'transit']),
            ])
            for m in linked_moves:
                moves_by_po_line.setdefault(m.purchase_line_id.id, []).append(m)

        # ── Réceptions directes (sans lien commande) ──────────────────────────
        standalone_domain = [
            ('purchase_line_id', '=', False),
            ('picking_id.picking_type_code', '=', 'incoming'),
            ('state', '=', 'done'),
            ('picking_id.date_done', '>=', self.date_debut),
            ('picking_id.date_done', '<=', self.date_fin),
            ('picking_id.company_id', '=', self.company_id.id),
            ('location_id.usage', 'in', ['supplier', 'transit']),
        ]
        if self.fournisseur_ids:
            standalone_domain.append(('picking_id.partner_id', 'in', self.fournisseur_ids.ids))
        standalone_moves = self.env['stock.move'].search(standalone_domain)

        if not po_lines and not standalone_moves:
            raise UserError("Aucune donnée trouvée pour la période et les filtres sélectionnés.")

        products_data = {}

        # ── Traitement lignes PO ──────────────────────────────────────────────
        for line in po_lines:
            product = line.product_id
            if not product:
                continue
            key = product.id
            if key not in products_data:
                products_data[key] = {
                    'product': product,
                    'po_ids': set(),
                    'reception_ids': set(),
                    'qty_commandee': 0.0,
                    'montant_commande': 0.0,
                    'qty_receptionnee': 0.0,
                    'montant_reception': 0.0,
                    'sub_lines': [],
                }
            products_data[key]['qty_commandee'] += line.product_qty
            products_data[key]['montant_commande'] += line.product_qty * line.price_unit
            products_data[key]['po_ids'].add(line.order_id.name)

            linked = moves_by_po_line.get(line.id, [])
            if linked:
                for move in linked:
                    qty_done = (
                        sum(move.move_line_ids.mapped('quantity'))
                        or move.product_uom_qty
                    )
                    prix_rec = move.price_unit or line.price_unit
                    products_data[key]['qty_receptionnee'] += qty_done
                    products_data[key]['montant_reception'] += qty_done * prix_rec
                    products_data[key]['reception_ids'].add(move.picking_id.name)
                    products_data[key]['sub_lines'].append({
                        'po_ref':           line.order_id.name,
                        'date_commande':    line.order_id.date_order,
                        'qty_commandee':    line.product_qty,
                        'prix_commande':    line.price_unit,
                        'montant_commande': line.product_qty * line.price_unit,
                        'reception_ref':    move.picking_id.name,
                        'date_reception':   move.picking_id.date_done,
                        'qty_receptionnee': qty_done,
                        'prix_reception':   prix_rec,
                        'montant_reception': qty_done * prix_rec,
                    })
            else:
                # Commande sans réception dans la période
                products_data[key]['sub_lines'].append({
                    'po_ref':           line.order_id.name,
                    'date_commande':    line.order_id.date_order,
                    'qty_commandee':    line.product_qty,
                    'prix_commande':    line.price_unit,
                    'montant_commande': line.product_qty * line.price_unit,
                    'reception_ref':    '',
                    'date_reception':   None,
                    'qty_receptionnee': 0.0,
                    'prix_reception':   0.0,
                    'montant_reception': 0.0,
                })

        # ── Réceptions directes ───────────────────────────────────────────────
        for move in standalone_moves:
            product = move.product_id
            if not product:
                continue
            key = product.id
            if key not in products_data:
                products_data[key] = {
                    'product': product,
                    'po_ids': set(),
                    'reception_ids': set(),
                    'qty_commandee': 0.0,
                    'montant_commande': 0.0,
                    'qty_receptionnee': 0.0,
                    'montant_reception': 0.0,
                    'sub_lines': [],
                }
            qty_done = (
                sum(move.move_line_ids.mapped('quantity'))
                or move.product_uom_qty
            )
            prix_rec = move.price_unit
            products_data[key]['qty_receptionnee'] += qty_done
            products_data[key]['montant_reception'] += qty_done * prix_rec
            if move.picking_id:
                products_data[key]['reception_ids'].add(move.picking_id.name)
            products_data[key]['sub_lines'].append({
                'po_ref':           '',
                'date_commande':    None,
                'qty_commandee':    0.0,
                'prix_commande':    0.0,
                'montant_commande': 0.0,
                'reception_ref':    move.picking_id.name if move.picking_id else '',
                'date_reception':   move.picking_id.date_done if move.picking_id else None,
                'qty_receptionnee': qty_done,
                'prix_reception':   prix_rec,
                'montant_reception': qty_done * prix_rec,
            })

        result = []
        for data in products_data.values():
            sub = sorted(
                data['sub_lines'],
                key=lambda x: x['date_commande'] or x['date_reception'] or _dt.min,
            )
            result.append({
                'ref':               data['product'].default_code or '',
                'name':              data['product'].name,
                'nb_commandes':      len(data['po_ids']),
                'qty_commandee':     data['qty_commandee'],
                'montant_commande':  data['montant_commande'],
                'nb_receptions':     len(data['reception_ids']),
                'qty_receptionnee':  data['qty_receptionnee'],
                'montant_reception': data['montant_reception'],
                'sub_lines':         sub,
            })

        result.sort(key=lambda x: (x['ref'].lower() if x['ref'] else x['name'].lower()))
        return result

    def action_print_report(self):
        self.ensure_one()
        self._get_report_data()
        return self.env.ref('custom_reports.action_report_achat_produit').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self._get_report_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Cmdes & Réceptions Produits"

        # ── Palette ───────────────────────────────────────────────────────────
        BLUE    = "1A5276"
        GRAY    = "D5D8DC"   # sous-ligne sans réception
        LBLUE   = "EBF5FB"   # alternance lignes paires
        WHITE   = "FFFFFF"
        ORANGE  = "FAD7A0"   # réception directe (sans commande)
        thin    = Side(style='thin', color="BBBBBB")
        brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

        def font(bold=False, color="000000", size=9):
            return Font(name="Arial", bold=bold, color=color, size=size)

        def fill(c):
            return PatternFill("solid", fgColor=c)

        def aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def fmt_date(d):
            return d.strftime('%d/%m/%Y') if d else ''

        # ── En-tête société / titre ───────────────────────────────────────────
        NCOLS = 11
        ws.merge_cells("A1:K1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = font(bold=True, size=11)

        ws.merge_cells("A2:K2")
        ws["A2"] = (
            f"COMMANDES & RÉCEPTIONS PAR PRODUIT — "
            f"Du {self.date_debut.strftime('%d/%m/%Y')} au {self.date_fin.strftime('%d/%m/%Y')}"
        )
        ws["A2"].font = font(bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        # ── En-têtes colonnes ─────────────────────────────────────────────────
        #  A          B               C          D          E             F
        #  Réf.  | Désignation  | Date Cmd | Qté Cmd | Prix Unit | Montant Cmd
        #  G               H         I         J             K
        #  Réf. Réception | Date Réc | Qté Réc | Prix Unit | Montant Réc
        hdrs = [
            "Réf. Produit", "Désignation / Réf. Commande",
            "Date Commande", "Qté Commandée", "Prix Unit. Cmd", "Montant Cmdé",
            "Réf. Réception", "Date Réception", "Qté Réceptionnée",
            "Prix Unit. Réc.", "Montant Réc.",
        ]
        ws.append(hdrs)
        hr = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=hr, column=col)
            c.font = font(bold=True, color=WHITE, size=9)
            c.fill = fill(BLUE)
            c.alignment = aln("center")
            c.border = brd
        ws.row_dimensions[hr].height = 22

        # ── Données ───────────────────────────────────────────────────────────
        tot_qty_cmd = 0.0
        tot_mnt_cmd = 0.0
        tot_qty_rec = 0.0
        tot_mnt_rec = 0.0

        for prod in data:
            # ── Ligne produit (header) ────────────────────────────────────────
            ws.append([
                prod['ref'],
                prod['name'],
                f"{prod['nb_commandes']} cmde(s)",
                prod['qty_commandee'],
                '',
                prod['montant_commande'],
                f"{prod['nb_receptions']} réc.",
                '',
                prod['qty_receptionnee'],
                '',
                prod['montant_reception'],
            ])
            r = ws.max_row
            for col in range(1, NCOLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = font(bold=True, color=WHITE, size=9)
                c.fill = fill(BLUE)
                c.border = brd
                c.alignment = aln("right" if col in (4, 6, 9, 11) else "center" if col == 3 else "left")
            ws.cell(row=r, column=4).number_format = '#,##0.##'
            ws.cell(row=r, column=6).number_format = '#,##0'
            ws.cell(row=r, column=9).number_format = '#,##0.##'
            ws.cell(row=r, column=11).number_format = '#,##0'
            ws.row_dimensions[r].height = 16

            tot_qty_cmd += prod['qty_commandee']
            tot_mnt_cmd += prod['montant_commande']
            tot_qty_rec += prod['qty_receptionnee']
            tot_mnt_rec += prod['montant_reception']

            # ── Sous-lignes ───────────────────────────────────────────────────
            for i, sl in enumerate(prod['sub_lines']):
                has_po  = bool(sl['po_ref'])
                has_rec = bool(sl['reception_ref'])
                if has_po and has_rec:
                    bg = LBLUE if i % 2 == 0 else WHITE
                elif has_po:
                    bg = GRAY   # commandé mais pas encore reçu
                else:
                    bg = ORANGE  # réception directe

                ws.append([
                    '',
                    sl['po_ref'],
                    fmt_date(sl['date_commande']),
                    sl['qty_commandee'] or '',
                    sl['prix_commande'] or '',
                    sl['montant_commande'] or '',
                    sl['reception_ref'],
                    fmt_date(sl['date_reception']),
                    sl['qty_receptionnee'] or '',
                    sl['prix_reception'] or '',
                    sl['montant_reception'] or '',
                ])
                r = ws.max_row
                for col in range(1, NCOLS + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = font(size=8)
                    c.fill = fill(bg)
                    c.border = brd
                    c.alignment = aln("right" if col in (4, 5, 6, 9, 10, 11) else "center" if col == 3 else "left")
                for col in (4, 9):
                    ws.cell(row=r, column=col).number_format = '#,##0.##'
                for col in (5, 10):
                    ws.cell(row=r, column=col).number_format = '#,##0.00'
                for col in (6, 11):
                    ws.cell(row=r, column=col).number_format = '#,##0'
                ws.row_dimensions[r].height = 14

        # ── Ligne grand total ─────────────────────────────────────────────────
        ws.append(["", "TOTAL GÉNÉRAL", "", tot_qty_cmd, "", tot_mnt_cmd,
                   "", "", tot_qty_rec, "", tot_mnt_rec])
        r = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = font(bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = brd
            c.alignment = aln("right" if col in (4, 6, 9, 11) else "left")
        ws.cell(row=r, column=4).number_format = '#,##0.##'
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=9).number_format = '#,##0.##'
        ws.cell(row=r, column=11).number_format = '#,##0'
        ws.row_dimensions[r].height = 18

        # ── Légende ───────────────────────────────────────────────────────────
        ws.append([])
        ws.append(["Légende :"])
        ws.cell(ws.max_row, 1).font = font(bold=True, size=8)
        for txt, bg in [
            ("Commandé + Réceptionné", LBLUE),
            ("Commandé sans réception", GRAY),
            ("Réception directe (sans commande)", ORANGE),
        ]:
            ws.append(["", txt])
            r = ws.max_row
            ws.cell(r, 2).font = font(size=8)
            ws.cell(r, 2).fill = fill(bg)
            ws.cell(r, 2).border = brd

        # ── Largeurs colonnes ─────────────────────────────────────────────────
        for col, w in enumerate([12, 32, 14, 12, 13, 14, 18, 14, 14, 13, 14], 1):
            ws.column_dimensions[chr(64 + col)].width = w

        # ── Export ────────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read()).decode()

        fname = (
            f"Commandes_Receptions_Produits_"
            f"{self.date_debut.strftime('%d%m%Y')}_{self.date_fin.strftime('%d%m%Y')}.xlsx"
        )
        att = self.env['ir.attachment'].create({
            'name':      fname,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'new',
        }
