# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class SaleStatReportWizard(models.TransientModel):
    _name = 'sale.stat.report.wizard'
    _description = 'Assistant Statistiques de Ventes'

    date_start_period1 = fields.Date(string='Début Période 1')
    date_end_period1 = fields.Date(string='Fin Période 1')
    date_start_period2 = fields.Date(string='Début Période 2')
    date_end_period2 = fields.Date(string='Fin Période 2')

    report_mode = fields.Selection([
        ('single', 'Une période'),
        ('comparison', 'Comparaison deux périodes'),
    ], string='Mode', default='comparison', required=True)

    order_source = fields.Selection([
        ('pos', 'Point de Vente (POS)'),
        ('sale', 'Ventes'),
        ('both', 'Les deux'),
    ], string='Source', default='both', required=True)

    group_by = fields.Selection([
        ('customer_category', 'Catégorie Client'),
        ('customer', 'Client'),
        ('product_category', 'Catégorie Produit'),
        ('product', 'Produit'),
    ], string='Grouper par', default='customer_category', required=True)

    # --- Filtres ---
    partner_ids = fields.Many2many(
        'res.partner',
        string='Clients',
        domain=[('customer_rank', '>', 0)],
    )
    category_ids = fields.Many2many(
        'res.partner.category',
        string='Catégories clients',
    )
    product_category_ids = fields.Many2many(
        'product.category',
        string='Catégories produits',
    )
    product_ids = fields.Many2many(
        'product.product',
        string='Produits',
        domain=[('sale_ok', '=', True)],
    )

    company_ids = fields.Many2many(
        'res.company',
        string='Sociétés',
        required=True,
        default=lambda self: self.env.company,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        compute='_compute_company_id',
        help="Première société sélectionnée, utilisée pour l'en-tête du document.",
    )

    @api.depends('company_ids')
    def _compute_company_id(self):
        for record in self:
            record.company_id = record.company_ids[:1]

    @api.constrains('date_start_period1', 'date_end_period1', 'date_start_period2', 'date_end_period2')
    def _check_dates(self):
        for record in self:
            if record.date_start_period1 and record.date_end_period1:
                if record.date_start_period1 > record.date_end_period1:
                    raise UserError("La date de début de la période 1 doit être avant la date de fin.")
            if record.report_mode == 'comparison':
                if record.date_start_period2 and record.date_end_period2:
                    if record.date_start_period2 > record.date_end_period2:
                        raise UserError("La date de début de la période 2 doit être avant la date de fin.")

    # ------------------------------------------------------------------
    # Helpers internes
    # ------------------------------------------------------------------

    def _build_order_domain(self, period):
        """Domaine pour POS — filtre par date_order."""
        date_start = self.date_start_period1 if period == 1 else self.date_start_period2
        date_end   = self.date_end_period1   if period == 1 else self.date_end_period2
        domain = [('company_id', 'in', self.company_ids.ids)]
        if date_start:
            domain.append(('date_order', '>=', fields.Datetime.to_datetime(date_start)))
        if date_end:
            dt_end = fields.Datetime.to_datetime(date_end).replace(hour=23, minute=59, second=59)
            domain.append(('date_order', '<=', dt_end))
        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))
        if self.category_ids:
            domain.append(('partner_id.category_id', 'in', self.category_ids.ids))
        return domain

    def _build_invoice_domain(self, period):
        """Domaine pour factures/avoirs vente — filtre par invoice_date."""
        date_start = self.date_start_period1 if period == 1 else self.date_start_period2
        date_end   = self.date_end_period1   if period == 1 else self.date_end_period2
        domain = [('company_id', 'in', self.company_ids.ids), ('state', '=', 'posted')]
        if date_start:
            domain.append(('invoice_date', '>=', date_start))
        if date_end:
            domain.append(('invoice_date', '<=', date_end))
        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))
        if self.category_ids:
            domain.append(('partner_id.category_id', 'in', self.category_ids.ids))
        return domain

    def _fetch_orders(self, period):
        result = {
            'pos': self.env['pos.order'],
            'sale': self.env['sale.order'],
            'refund': self.env['account.move'],
        }
        inv_base = self._build_invoice_domain(period)
        if self.order_source in ('pos', 'both'):
            base_pos = self._build_order_domain(period)
            result['pos'] = self.env['pos.order'].search(
                [('state', 'in', ['paid', 'done', 'invoiced'])] + base_pos
            )
        if self.order_source in ('sale', 'both'):
            # Commandes de vente facturées et validées (hors POS)
            invoices = self.env['account.move'].search(
                inv_base + [('move_type', '=', 'out_invoice'), ('pos_order_ids', '=', False)]
            )
            result['sale'] = self.env['sale.order'].search([('invoice_ids', 'in', invoices.ids)])
        # Avoirs de vente (hors POS) toujours inclus quelle que soit la source :
        # le rapport CA les soustrait du CA dans tous les modes (POS seul inclus)
        result['refund'] = self.env['account.move'].search(
            inv_base + [('move_type', '=', 'out_refund'), ('pos_order_ids', '=', False)]
        )
        return result

    def _filter_orders_by_product(self, orders_dict):
        """Garde uniquement les commandes/avoirs contenant au moins un produit/catégorie filtrés."""
        result = {}
        for src, orders in orders_dict.items():
            if not orders:
                result[src] = orders
                continue
            if src == 'refund':
                ld = [('move_id', 'in', orders.ids), ('product_id', '!=', False)]
                if self.product_ids:
                    ld.append(('product_id', 'in', self.product_ids.ids))
                if self.product_category_ids:
                    ld.append(('product_id.categ_id', 'child_of', self.product_category_ids.ids))
                valid_ids = set(self.env['account.move.line'].search(ld).mapped('move_id').ids)
                result[src] = orders.filtered(lambda o: o.id in valid_ids)
            else:
                line_model = 'pos.order.line' if src == 'pos' else 'sale.order.line'
                ld = [('order_id', 'in', orders.ids)]
                if self.product_ids:
                    ld.append(('product_id', 'in', self.product_ids.ids))
                if self.product_category_ids:
                    ld.append(('product_id.categ_id', 'child_of', self.product_category_ids.ids))
                valid_ids = set(self.env[line_model].search(ld).mapped('order_id').ids)
                result[src] = orders.filtered(lambda o: o.id in valid_ids)
        return result

    def _fetch_lines(self, orders_dict):
        """Retourne les lignes filtrées par produit/catégorie. Les lignes d'avoirs sont incluses."""
        result = []
        for src, orders in orders_dict.items():
            if not orders:
                continue
            if src == 'refund':
                ld = [('move_id', 'in', orders.ids), ('product_id', '!=', False)]
                if self.product_ids:
                    ld.append(('product_id', 'in', self.product_ids.ids))
                if self.product_category_ids:
                    ld.append(('product_id.categ_id', 'child_of', self.product_category_ids.ids))
                result.append(('refund', self.env['account.move.line'].search(ld)))
            else:
                line_model = 'pos.order.line' if src == 'pos' else 'sale.order.line'
                ld = [('order_id', 'in', orders.ids)]
                if self.product_ids:
                    ld.append(('product_id', 'in', self.product_ids.ids))
                if self.product_category_ids:
                    ld.append(('product_id.categ_id', 'child_of', self.product_category_ids.ids))
                result.append((src, self.env[line_model].search(ld)))
        return result

    # ------------------------------------------------------------------
    # Méthode principale de calcul
    # ------------------------------------------------------------------

    def get_sale_data(self):
        """Retourne les données de vente groupées selon self.group_by.

        Les avoirs (remboursements) de vente sont intégrés avec des valeurs négatives.
        POS : filtré par date_order. Ventes : filtré par invoice_date (factures validées).
        """
        self.ensure_one()
        groups = {}

        def make_group(name, gid):
            return {
                'group_name': name, 'group_id': gid, 'items': {},
                'total_p1_qty': 0.0, 'total_p1_ca': 0.0, 'total_p1_ca_ttc': 0.0, 'total_p1_margin': 0.0,
                'total_p2_qty': 0.0, 'total_p2_ca': 0.0, 'total_p2_ca_ttc': 0.0, 'total_p2_margin': 0.0,
            }

        def make_item(name, ref=''):
            return {
                'name': name, 'ref': ref,
                'qty_p1': 0.0, 'ca_p1': 0.0, 'ca_ttc_p1': 0.0, 'margin_p1': 0.0, 'margin_pct_p1': 0.0,
                'qty_p2': 0.0, 'ca_p2': 0.0, 'ca_ttc_p2': 0.0, 'margin_p2': 0.0, 'margin_pct_p2': 0.0,
                'prog_qty': 0.0, 'prog_ca': 0.0, 'prog_ca_pct': 0.0,
                'prog_margin': 0.0, 'prog_margin_pct': 0.0,
            }

        def acc(gkey, gname, gid, ikey, iname, iref, qty, ca, ca_ttc, margin, period):
            if gkey not in groups:
                groups[gkey] = make_group(gname, gid)
            g = groups[gkey]
            if ikey not in g['items']:
                g['items'][ikey] = make_item(iname, iref)
            it = g['items'][ikey]
            if period == 1:
                it['qty_p1'] += qty; it['ca_p1'] += ca; it['ca_ttc_p1'] += ca_ttc; it['margin_p1'] += margin
                g['total_p1_qty'] += qty; g['total_p1_ca'] += ca; g['total_p1_ca_ttc'] += ca_ttc; g['total_p1_margin'] += margin
            else:
                it['qty_p2'] += qty; it['ca_p2'] += ca; it['ca_ttc_p2'] += ca_ttc; it['margin_p2'] += margin
                g['total_p2_qty'] += qty; g['total_p2_ca'] += ca; g['total_p2_ca_ttc'] += ca_ttc; g['total_p2_margin'] += margin

        # --- Stratégies de groupement ---

        def by_customer_category(orders_dict, period):
            for src, orders in orders_dict.items():
                for order in orders:
                    p = order.partner_id
                    if src == 'refund':
                        ca     = -(order.amount_untaxed or 0.0)
                        ca_ttc = -(order.amount_total or 0.0)
                        mg     = 0.0
                        qty    = -1
                    else:
                        ca     = (order.amount_total or 0.0) - (order.amount_tax or 0.0)
                        ca_ttc = order.amount_total or 0.0
                        mg     = sum(l.margin for l in order.lines) if src == 'pos' else (getattr(order, 'margin', 0.0) or 0.0)
                        qty    = 1
                    cats = p.category_id
                    if self.category_ids:
                        cats = cats.filtered(lambda c: c.id in self.category_ids.ids)
                    if not cats:
                        acc('_no_cat', 'Sans Catégorie', 0,
                            p.id, p.name or '', p.ref or '', qty, ca, ca_ttc, mg, period)
                    else:
                        for cat in cats:
                            acc(cat.name, cat.name, cat.id,
                                p.id, p.name or '', p.ref or '', qty, ca, ca_ttc, mg, period)

        def by_customer(orders_dict, period):
            for src, orders in orders_dict.items():
                for order in orders:
                    p = order.partner_id
                    if src == 'refund':
                        ca     = -(order.amount_untaxed or 0.0)
                        ca_ttc = -(order.amount_total or 0.0)
                        mg     = 0.0
                        qty    = -1
                    else:
                        ca     = (order.amount_total or 0.0) - (order.amount_tax or 0.0)
                        ca_ttc = order.amount_total or 0.0
                        mg     = sum(l.margin for l in order.lines) if src == 'pos' else (getattr(order, 'margin', 0.0) or 0.0)
                        qty    = 1
                    gkey = p.id or 0
                    acc(gkey, p.name or 'Inconnu', gkey,
                        gkey, p.name or 'Inconnu', p.ref or '', qty, ca, ca_ttc, mg, period)

        def by_product_category(lines_list, period):
            for src, lines in lines_list:
                for line in lines:
                    prod = line.product_id
                    cat  = prod.categ_id
                    if src == 'refund':
                        ca     = -(line.price_subtotal or 0.0)
                        ca_ttc = -(line.price_total or 0.0)
                        mg     = 0.0
                        qty    = -(line.quantity or 0.0)
                    else:
                        ca     = line.price_subtotal or 0.0
                        ca_ttc = (line.price_subtotal_incl if src == 'pos' else line.price_total) or 0.0
                        mg     = getattr(line, 'margin', 0.0) or 0.0
                        qty    = line.qty if src == 'pos' else line.product_uom_qty
                    gkey  = cat.name if cat else '_no_cat'
                    gname = cat.name if cat else 'Sans Catégorie'
                    gid   = cat.id   if cat else 0
                    acc(gkey, gname, gid,
                        prod.id, prod.name or '', prod.default_code or '', qty, ca, ca_ttc, mg, period)

        def by_product(lines_list, period):
            for src, lines in lines_list:
                for line in lines:
                    prod = line.product_id
                    if src == 'refund':
                        ca     = -(line.price_subtotal or 0.0)
                        ca_ttc = -(line.price_total or 0.0)
                        mg     = 0.0
                        qty    = -(line.quantity or 0.0)
                    else:
                        ca     = line.price_subtotal or 0.0
                        ca_ttc = (line.price_subtotal_incl if src == 'pos' else line.price_total) or 0.0
                        mg     = getattr(line, 'margin', 0.0) or 0.0
                        qty    = line.qty if src == 'pos' else line.product_uom_qty
                    gkey = prod.id or 0
                    acc(gkey, prod.name or 'Inconnu', gkey,
                        gkey, prod.name or 'Inconnu', prod.default_code or '', qty, ca, ca_ttc, mg, period)

        # --- Exécution par période ---

        def run_period(period):
            orders = self._fetch_orders(period)
            gb = self.group_by
            if gb in ('customer_category', 'customer') and (self.product_ids or self.product_category_ids):
                orders = self._filter_orders_by_product(orders)
            if gb == 'customer_category':
                by_customer_category(orders, period)
            elif gb == 'customer':
                by_customer(orders, period)
            elif gb == 'product_category':
                by_product_category(self._fetch_lines(orders), period)
            elif gb == 'product':
                by_product(self._fetch_lines(orders), period)

        run_period(1)
        if self.report_mode == 'comparison':
            run_period(2)

        # --- Calcul des % et progressions ---
        for gdata in groups.values():
            for vals in gdata['items'].values():
                if vals['ca_p1']:
                    vals['margin_pct_p1'] = vals['margin_p1'] / vals['ca_p1'] * 100
                if self.report_mode == 'comparison':
                    if vals['ca_p2']:
                        vals['margin_pct_p2'] = vals['margin_p2'] / vals['ca_p2'] * 100
                    vals['prog_qty']    = vals['qty_p2']    - vals['qty_p1']
                    vals['prog_ca']     = vals['ca_p2']     - vals['ca_p1']
                    vals['prog_margin'] = vals['margin_p2'] - vals['margin_p1']
                    vals['prog_ca_pct'] = (
                        vals['prog_ca'] / vals['ca_p1'] * 100 if vals['ca_p1']
                        else (100.0 if vals['ca_p2'] else 0.0)
                    )
                    vals['prog_margin_pct'] = (
                        vals['prog_margin'] / vals['margin_p1'] * 100 if vals['margin_p1']
                        else (100.0 if vals['margin_p2'] else 0.0)
                    )

            gdata['total_p1_margin_pct'] = (
                gdata['total_p1_margin'] / gdata['total_p1_ca'] * 100
                if gdata['total_p1_ca'] else 0.0
            )
            if self.report_mode == 'comparison':
                gdata['total_p2_margin_pct'] = (
                    gdata['total_p2_margin'] / gdata['total_p2_ca'] * 100
                    if gdata['total_p2_ca'] else 0.0
                )
                gdata['total_prog_qty']    = gdata['total_p2_qty']    - gdata['total_p1_qty']
                gdata['total_prog_ca']     = gdata['total_p2_ca']     - gdata['total_p1_ca']
                gdata['total_prog_margin'] = gdata['total_p2_margin'] - gdata['total_p1_margin']
                gdata['total_prog_ca_pct'] = (
                    gdata['total_prog_ca'] / gdata['total_p1_ca'] * 100
                    if gdata['total_p1_ca'] else (100.0 if gdata['total_p2_ca'] else 0.0)
                )

        return dict(sorted(groups.items(), key=lambda x: str(x[0])))

    def get_sale_data_by_category(self):
        """Alias maintenu pour compatibilité avec le template PDF."""
        return self.get_sale_data()

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def action_print_report(self):
        self.ensure_one()
        return self.env.ref('custom_reports.action_report_sale_statistics').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self.get_sale_data()
        if not data:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")

        col_labels = {
            'customer_category': ('Catég. Client', 'Client'),
            'customer':          ('Client',         'Réf. Client'),
            'product_category':  ('Catég. Produit', 'Produit'),
            'product':           ('Produit',        'Code Article'),
        }
        col1_label, col2_label = col_labels.get(self.group_by, ('Groupe', 'Détail'))
        use_ref_col2 = self.group_by in ('product', 'customer')

        wb = Workbook(); ws = wb.active; ws.title = "Stats Ventes"
        BLUE = "1A5276"; LBLUE = "D6EAF8"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        is_comparison = self.report_mode == 'comparison'
        p1_label = (
            f"{self.date_start_period1.strftime('%d/%m/%Y')} → {self.date_end_period1.strftime('%d/%m/%Y')}"
            if self.date_start_period1 else "Période 1"
        )

        # Single  : Groupe|Code|Qté|CA HT|CA TTC|Marge|% Marge  (7)
        # Compar. : id. + P2(Qté|CA HT|CA TTC|Marge|% Marge) + Prog CA HT|% Prog (14)
        if is_comparison:
            last_col = 14
            headers = [
                col1_label, col2_label,
                "Qté P1", "CA HT P1", "CA TTC P1", "Marge P1", "% Marge P1",
                "Qté P2", "CA HT P2", "CA TTC P2", "Marge P2", "% Marge P2",
                "Prog CA HT", "% Prog CA",
            ]
            money_cols = (4, 5, 6, 9, 10, 11, 13)
        else:
            last_col = 7
            headers = [col1_label, col2_label, "Qté", "CA HT", "CA TTC", "Marge", "% Marge"]
            money_cols = (4, 5, 6)

        last_col_letter = chr(64 + last_col)

        # ── En-tête titre ──────────────────────────────────────────────────────
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = ', '.join(self.company_ids.mapped('name'))
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        ws.merge_cells(f"A2:{last_col_letter}2")
        if is_comparison:
            p2_label = (
                f"{self.date_start_period2.strftime('%d/%m/%Y')} → {self.date_end_period2.strftime('%d/%m/%Y')}"
                if self.date_start_period2 else "Période 2"
            )
            ws["A2"] = f"STATISTIQUES VENTES — P1: {p1_label} | P2: {p2_label}"
        else:
            ws["A2"] = f"STATISTIQUES VENTES — {p1_label}"
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE); ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18; ws.append([])

        # ── En-têtes colonnes ──────────────────────────────────────────────────
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, last_col + 1):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        # ── Données ────────────────────────────────────────────────────────────
        for gdata in data.values():
            for item in gdata['items'].values():
                col2_val = item['ref'] if use_ref_col2 else item['name']
                if is_comparison:
                    row_data = [
                        gdata['group_name'], col2_val,
                        round(item['qty_p1'], 3), item['ca_p1'], item['ca_ttc_p1'],
                        item['margin_p1'], round(item['margin_pct_p1'], 2),
                        round(item['qty_p2'], 3), item['ca_p2'], item['ca_ttc_p2'],
                        item['margin_p2'], round(item['margin_pct_p2'], 2),
                        item['prog_ca'], round(item['prog_ca_pct'], 2),
                    ]
                else:
                    row_data = [
                        gdata['group_name'], col2_val,
                        round(item['qty_p1'], 3), item['ca_p1'], item['ca_ttc_p1'],
                        item['margin_p1'], round(item['margin_pct_p1'], 2),
                    ]
                ws.append(row_data)
                r = ws.max_row
                for col in range(1, last_col + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9); c.border = brd
                    c.alignment = aln("right" if col >= 3 else "left")
                for col in money_cols:
                    ws.cell(row=r, column=col).number_format = '#,##0.00'

            # ── Sous-total groupe ──────────────────────────────────────────────
            if is_comparison:
                sub_row = [
                    "Sous-total " + gdata['group_name'], "",
                    round(gdata['total_p1_qty'], 3), gdata['total_p1_ca'], gdata['total_p1_ca_ttc'],
                    gdata['total_p1_margin'], round(gdata.get('total_p1_margin_pct', 0.0), 2),
                    round(gdata['total_p2_qty'], 3), gdata['total_p2_ca'], gdata['total_p2_ca_ttc'],
                    gdata['total_p2_margin'], round(gdata.get('total_p2_margin_pct', 0.0), 2),
                    gdata.get('total_prog_ca', 0.0), round(gdata.get('total_prog_ca_pct', 0.0), 2),
                ]
            else:
                sub_row = [
                    "Sous-total " + gdata['group_name'], "",
                    round(gdata['total_p1_qty'], 3), gdata['total_p1_ca'], gdata['total_p1_ca_ttc'],
                    gdata['total_p1_margin'], round(gdata.get('total_p1_margin_pct', 0.0), 2),
                ]
            ws.append(sub_row)
            r = ws.max_row
            for col in range(1, last_col + 1):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE); c.border = brd
                c.alignment = aln("right" if col >= 3 else "left")
            for col in money_cols:
                ws.cell(row=r, column=col).number_format = '#,##0.00'

        # ── Largeurs colonnes ──────────────────────────────────────────────────
        if is_comparison:
            col_widths = [22, 14, 8, 13, 13, 13, 9, 8, 13, 13, 13, 9, 13, 9]
        else:
            col_widths = [22, 14, 8, 13, 13, 13, 9]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        attachment = self.env['ir.attachment'].create({
            'name': 'Stats_Ventes.xlsx', 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}
