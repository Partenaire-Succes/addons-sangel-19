# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class RetoursConsolidesReportWizard(models.TransientModel):
    """
    BLOC 6 — Rapport consolidé des retours et réceptions internes.

    Couvre :
      - Réceptions directes (BLOC 2) avec prix de réception et prix mis à jour.
      - Retours fournisseurs (BLOC 5) : pickings incoming inversés (stock → fournisseur).
      - Retours inventaires (BLOC 4) : avoirs fournisseurs générés depuis l'inventaire.
    """
    _name = 'retours.consolides.report.wizard'
    _description = 'Rapport Consolidé Retours et Réceptions'

    type_rapport = fields.Selection([
        ('receptions', 'Réceptions directes uniquement'),
        ('retours',    'Retours uniquement (Fournisseurs + Inventaires)'),
        ('tous',       'Tout (Réceptions + Retours)'),
    ], string='Contenu du rapport', default='tous', required=True)

    date_from = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today,
    )
    date_to = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for rec in self:
            if rec.date_from > rec.date_to:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    # ────────────────────────────────────────────────────────────────────────
    # DONNÉES RAPPORT
    # ────────────────────────────────────────────────────────────────────────

    # ────────────────────────────────────────────────────────────────────────
    # HELPERS FORMAT
    # ────────────────────────────────────────────────────────────────────────
    def _fmt_date(self, d):
        if not d:
            return '—'
        if hasattr(d, 'strftime'):
            return d.strftime('%d/%m/%Y')
        return str(d)

    def _fmt_amount(self, amount):
        symbol = self.company_id.currency_id.symbol or ''
        return '{:,.0f} {}'.format(amount or 0, symbol).replace(',', ' ')

    # ────────────────────────────────────────────────────────────────────────
    # DONNÉES RAPPORT
    # ────────────────────────────────────────────────────────────────────────

    def _get_receptions_directes(self):
        """Réceptions sans BdC (BLOC 2) — prix gravé sur le stock.move (historique réel)."""
        pickings = self.env['stock.picking'].search([
            ('state', '=', 'done'),
            ('origin', '=', 'Réception Directe'),
            ('date_done', '>=', self.date_from),
            ('date_done', '<=', self.date_to),
            ('company_id', '=', self.company_id.id),
        ], order='date_done')

        rows = []
        for picking in pickings:
            for move in picking.move_ids.filtered(lambda m: m.state == 'done'):
                tmpl = move.product_id.product_tmpl_id
                prix_standard = tmpl.standard_price
                # move.price_unit = coût gravé lors de la réception (nouveau prix si modifié)
                prix_reception = move.price_unit
                a_nouveau_prix = (
                    prix_reception > 0
                    and abs(prix_reception - prix_standard) > 0.001
                )
                qty_done = sum(move.move_line_ids.mapped('quantity')) or move.product_uom_qty
                rows.append({
                    'date': self._fmt_date(picking.date_done),
                    'reference': picking.name,
                    'fournisseur': picking.partner_id.name or '—',
                    'notes': picking.note or '—',
                    'produit': move.product_id.display_name,
                    'qty': qty_done,
                    'uom': move.product_uom.name,
                    'prix_standard': prix_standard,
                    'prix_reception': prix_reception,
                    'a_nouveau_prix': a_nouveau_prix,
                    'montant': qty_done * prix_reception,
                })
        return rows

    def _get_retours_fournisseur(self):
        """Retours fournisseurs (BLOC 5) : stock → fournisseur."""
        pickings = self.env['stock.picking'].search([
            ('state', '=', 'done'),
            ('picking_type_code', '=', 'incoming'),
            ('location_id.usage', '=', 'internal'),
            ('location_dest_id.usage', '=', 'supplier'),
            ('date_done', '>=', self.date_from),
            ('date_done', '<=', self.date_to),
            ('company_id', '=', self.company_id.id),
        ], order='date_done')

        rows = []
        for picking in pickings:
            for move in picking.move_ids.filtered(lambda m: m.state == 'done'):
                qty_done = sum(move.move_line_ids.mapped('quantity')) or move.product_uom_qty
                montant = qty_done * move.price_unit
                rows.append({
                    'date': self._fmt_date(picking.date_done),
                    'reference': picking.name,
                    'origine': picking.origin or '—',
                    'fournisseur': picking.partner_id.name or '—',
                    'produit': move.product_id.display_name,
                    'qty': qty_done,
                    'uom': move.product_uom.name,
                    'prix_unitaire': move.price_unit,
                    'montant': montant,
                })
        return rows

    def _get_retours_inventaire(self):
        """Retours inventaires (BLOC 4) : avoirs fournisseurs issus d'inventaires physiques."""
        avoirs = self.env['account.move'].search([
            ('move_type', '=', 'in_refund'),
            ('state', '!=', 'cancel'),
            ('ref', 'like', 'Retour inventaire%'),
            ('invoice_date', '>=', self.date_from),
            ('invoice_date', '<=', self.date_to),
            ('company_id', '=', self.company_id.id),
        ], order='invoice_date')

        etat_labels = {
            'draft': 'Brouillon',
            'posted': 'Validé',
            'cancel': 'Annulé',
        }
        rows = []
        for avoir in avoirs:
            for line in avoir.invoice_line_ids.filtered(lambda l: not l.display_type):
                rows.append({
                    'date': self._fmt_date(avoir.invoice_date),
                    'reference': avoir.name or '—',
                    'origine': avoir.ref or '—',
                    'fournisseur': avoir.partner_id.name or '—',
                    'produit': line.product_id.display_name if line.product_id else line.name,
                    'qty': line.quantity,
                    'prix_unitaire': line.price_unit,
                    'montant': line.price_subtotal,
                    'etat': etat_labels.get(avoir.state, avoir.state),
                    'etat_code': avoir.state,
                })
        return rows

    def _get_type_label(self):
        labels = {
            'receptions': 'Réceptions directes',
            'retours':    'Retours (Fournisseurs + Inventaires)',
            'tous':       'Tout (Réceptions + Retours)',
        }
        return labels.get(self.type_rapport, '')

    def _get_report_data(self):
        """Structure complète des données pour le template QWeb."""
        receptions   = self._get_receptions_directes()   if self.type_rapport in ('receptions', 'tous') else []
        retours_four = self._get_retours_fournisseur()   if self.type_rapport in ('retours', 'tous')    else []
        retours_inv  = self._get_retours_inventaire()    if self.type_rapport in ('retours', 'tous')    else []

        total_receptions = sum(r['montant'] for r in receptions)
        total_retours_four = sum(r['montant'] for r in retours_four)
        total_retours_inv = sum(r['montant'] for r in retours_inv)
        grand_total = total_receptions + total_retours_four + total_retours_inv

        currency = self.company_id.currency_id

        return {
            'date_from': self._fmt_date(self.date_from),
            'date_to': self._fmt_date(self.date_to),
            'company': self.company_id,
            'currency': currency,
            'type_rapport': self.type_rapport,
            'type_label': self._get_type_label(),
            'receptions_directes': receptions,
            'retours_fournisseur': retours_four,
            'retours_inventaire': retours_inv,
            'total_receptions': total_receptions,
            'total_retours_four': total_retours_four,
            'total_retours_inv': total_retours_inv,
            'grand_total': grand_total,
            'fmt': self._fmt_amount,
        }

    # ────────────────────────────────────────────────────────────────────────
    # ACTION
    # ────────────────────────────────────────────────────────────────────────

    def action_print_report(self):
        self.ensure_one()
        data = self._get_report_data()
        total_lines = (
            len(data['receptions_directes'])
            + len(data['retours_fournisseur'])
            + len(data['retours_inventaire'])
        )
        if not total_lines:
            raise UserError(
                "Aucune opération trouvée pour la période et la société sélectionnées."
            )
        return self.env.ref(
            'custom_reports.action_report_retours_consolides'
        ).report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self._get_report_data()
        total_lines = (
            len(data['receptions_directes'])
            + len(data['retours_fournisseur'])
            + len(data['retours_inventaire'])
        )
        if not total_lines:
            raise UserError(
                "Aucune opération trouvée pour la période et la société sélectionnées."
            )

        wb = Workbook()

        BLUE  = "1A5276"
        LBLUE = "D6EAF8"
        WHITE = "FFFFFF"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(h):
            return PatternFill("solid", fgColor=h)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        def write_section(ws, title, headers, rows, col_widths):
            ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
            ws["A1"] = self.company_id.name
            ws["A1"].font = Font(name="Arial", bold=True, size=11)

            ws.merge_cells(f"A2:{chr(64 + len(headers))}2")
            ws["A2"] = (
                f"{title} — "
                f"Du {self.date_from.strftime('%d/%m/%Y')} au {self.date_to.strftime('%d/%m/%Y')}"
            )
            ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
            ws["A2"].fill = fill(BLUE)
            ws["A2"].alignment = aln("center")
            ws.row_dimensions[2].height = 18
            ws.append([])

            ws.append(headers)
            hrow = ws.max_row
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=hrow, column=col)
                c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
                c.fill = fill(BLUE)
                c.alignment = aln("center")
                c.border = brd

            for row_data in rows:
                ws.append(row_data)
                r = ws.max_row
                for col in range(1, len(headers) + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9)
                    c.border = brd
                    c.alignment = aln("right" if col == len(headers) else "left")
                ws.cell(row=r, column=len(headers)).number_format = '#,##0.00'

            total = sum(row[-1] for row in rows if isinstance(row[-1], (int, float)))
            ws.append([""] * (len(headers) - 2) + ["TOTAL", total])
            r = ws.max_row
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
                c.fill = fill(BLUE)
                c.border = brd
                c.alignment = aln("right" if col >= len(headers) - 1 else "left")
            ws.cell(row=r, column=len(headers)).number_format = '#,##0.00'

            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # ── Onglet Réceptions directes ────────────────────────────────────────
        if data['receptions_directes']:
            ws1 = wb.active
            ws1.title = "Réceptions"
            rows1 = [
                [r['date'], r['reference'], r['fournisseur'], r['produit'],
                 r['qty'], r['uom'], r['prix_reception'], r['montant']]
                for r in data['receptions_directes']
            ]
            write_section(ws1, "RÉCEPTIONS DIRECTES",
                          ["Date", "Référence", "Fournisseur", "Produit", "Qté", "UdM", "Prix Unit.", "Montant"],
                          rows1, [13, 18, 25, 30, 8, 8, 12, 14])
        else:
            wb.active.title = "Réceptions"

        # ── Onglet Retours fournisseurs ───────────────────────────────────────
        if data['retours_fournisseur']:
            ws2 = wb.create_sheet("Retours Fournisseurs")
            rows2 = [
                [r['date'], r['reference'], r['fournisseur'], r['produit'],
                 r['qty'], r['uom'], r['prix_unitaire'], r['montant']]
                for r in data['retours_fournisseur']
            ]
            write_section(ws2, "RETOURS FOURNISSEURS",
                          ["Date", "Référence", "Fournisseur", "Produit", "Qté", "UdM", "Prix Unit.", "Montant"],
                          rows2, [13, 18, 25, 30, 8, 8, 12, 14])

        # ── Onglet Retours inventaires ────────────────────────────────────────
        if data['retours_inventaire']:
            ws3 = wb.create_sheet("Retours Inventaires")
            rows3 = [
                [r['date'], r['reference'], r['fournisseur'], r['produit'],
                 r['qty'], r['prix_unitaire'], r['montant'], r['etat']]
                for r in data['retours_inventaire']
            ]
            write_section(ws3, "RETOURS INVENTAIRES",
                          ["Date", "Référence", "Fournisseur", "Produit", "Qté", "Prix Unit.", "Montant", "État"],
                          rows3, [13, 18, 25, 30, 8, 12, 14, 12])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        filename = (
            f"Retours_Consolides_{self.date_from.strftime('%d%m%Y')}"
            f"_{self.date_to.strftime('%d%m%Y')}.xlsx"
        )
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
