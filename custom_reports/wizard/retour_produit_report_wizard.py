# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class RetourProduitReportWizard(models.TransientModel):
    _name = 'retour.produit.report.wizard'
    _description = 'Rapport Retours Fournisseurs par Produit'

    date_debut = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today,
    )
    date_fin = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )
    fournisseur_ids = fields.Many2many(
        'res.partner',
        string='Fournisseurs',
        domain=[('supplier_rank', '>', 0)],
        help='Laissez vide pour inclure tous les fournisseurs',
    )

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for rec in self:
            if rec.date_debut > rec.date_fin:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    def _get_report_data(self):
        """
        Retourne pour chaque produit :
          - les totaux (header)
          - des sous-lignes (une par mouvement de retour fournisseur)

        Les retours fournisseurs (retour.fournisseur.wizard) sont des
        opérations autonomes (stock → fournisseur), sans lien vers une
        commande ou une réception d'origine — contrairement aux réceptions,
        il n'y a donc qu'un seul "côté" à afficher par sous-ligne.
        """
        self.ensure_one()

        domain = [
            ('state', '=', 'done'),
            ('location_dest_id.usage', '=', 'supplier'),
            ('picking_id.date_done', '>=', self.date_debut),
            ('picking_id.date_done', '<=', self.date_fin),
            ('picking_id.company_id', '=', self.company_id.id),
        ]
        if self.fournisseur_ids:
            domain.append(('picking_id.partner_id', 'in', self.fournisseur_ids.ids))
        moves = self.env['stock.move'].search(domain)

        if not moves:
            raise UserError("Aucun retour fournisseur trouvé pour la période et les filtres sélectionnés.")

        products_data = {}
        for move in moves:
            product = move.product_id
            if not product:
                continue
            key = product.id
            if key not in products_data:
                products_data[key] = {
                    'product': product,
                    'retour_ids': set(),
                    'qty_retournee': 0.0,
                    'montant_retour': 0.0,
                    'sub_lines': [],
                }

            qty_done = (
                sum(move.move_line_ids.mapped('quantity'))
                or move.product_uom_qty
            )
            prix = (
                move.price_unit
                or move.origin_returned_move_id.price_unit
                or product.standard_price
            )
            montant = qty_done * prix

            products_data[key]['qty_retournee'] += qty_done
            products_data[key]['montant_retour'] += montant
            if move.picking_id:
                products_data[key]['retour_ids'].add(move.picking_id.name)

            products_data[key]['sub_lines'].append({
                'retour_ref':     move.picking_id.name if move.picking_id else '',
                'date_retour':    move.picking_id.date_done if move.picking_id else None,
                'fournisseur':    move.picking_id.partner_id.name if move.picking_id.partner_id else '',
                'qty_retournee':  qty_done,
                'prix_retour':    prix,
                'montant_retour': montant,
            })

        result = []
        for data in products_data.values():
            sub = sorted(
                data['sub_lines'],
                key=lambda x: x['date_retour'] or fields.Datetime.now(),
            )
            result.append({
                'ref':             data['product'].default_code or '',
                'name':            data['product'].name,
                'nb_retours':      len(data['retour_ids']),
                'qty_retournee':   data['qty_retournee'],
                'montant_retour':  data['montant_retour'],
                'sub_lines':       sub,
            })

        result.sort(key=lambda x: (x['ref'].lower() if x['ref'] else x['name'].lower()))
        return result

    def action_print_report(self):
        self.ensure_one()
        self._get_report_data()
        return self.env.ref('custom_reports.action_report_retour_produit').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        data = self._get_report_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Retours Produits"

        BLUE  = "1A5276"
        WHITE = "FFFFFF"
        SL_A  = "D6EAF8"
        SL_B  = "EBF5FB"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def font(bold=False, color="000000", size=9):
            return Font(name="Arial", bold=bold, color=color, size=size)

        def fill(c):
            return PatternFill("solid", fgColor=c)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        def fmt_date(d):
            return d.strftime('%d/%m/%Y') if d else ''

        NCOLS = 6
        ws.merge_cells("A1:F1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = font(bold=True, size=11)

        ws.merge_cells("A2:F2")
        ws["A2"] = (
            f"RETOURS FOURNISSEURS PAR PRODUIT — "
            f"Du {self.date_debut.strftime('%d/%m/%Y')} au {self.date_fin.strftime('%d/%m/%Y')}"
        )
        ws["A2"].font = font(bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        hdrs = [
            "Réf. Produit", "Désignation / N° Retour", "Date Retour",
            "Fournisseur", "Qté Retournée", "Montant Retourné",
        ]
        ws.append(hdrs)
        hr = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=hr, column=col)
            c.font = font(bold=True, color=WHITE, size=9)
            c.fill = fill(BLUE)
            c.alignment = aln("center")
            c.border = brd
        ws.row_dimensions[hr].height = 22

        tot_qty = 0.0
        tot_mnt = 0.0

        for prod in data:
            ws.append([
                prod['ref'], prod['name'],
                f"{prod['nb_retours']} retour(s)", '',
                prod['qty_retournee'], prod['montant_retour'],
            ])
            r = ws.max_row
            for col in range(1, NCOLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = font(bold=True, color=WHITE, size=9)
                c.fill = fill(BLUE)
                c.border = brd
                c.alignment = aln("right" if col in (5, 6) else "left")
            ws.cell(row=r, column=5).number_format = '#,##0.##'
            ws.cell(row=r, column=6).number_format = '#,##0'
            ws.row_dimensions[r].height = 16

            tot_qty += prod['qty_retournee']
            tot_mnt += prod['montant_retour']

            for i, sl in enumerate(prod['sub_lines']):
                bg = SL_A if i % 2 == 0 else SL_B
                ws.append([
                    '', sl['retour_ref'], fmt_date(sl['date_retour']),
                    sl['fournisseur'], sl['qty_retournee'], sl['montant_retour'],
                ])
                r = ws.max_row
                for col in range(1, NCOLS + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = font(size=8)
                    c.fill = fill(bg)
                    c.border = brd
                    c.alignment = aln("right" if col in (5, 6) else "center" if col == 3 else "left")
                ws.cell(row=r, column=5).number_format = '#,##0.##'
                ws.cell(row=r, column=6).number_format = '#,##0'
                ws.row_dimensions[r].height = 14

        ws.append(["", "TOTAL GÉNÉRAL", "", "", tot_qty, tot_mnt])
        r = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = font(bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = brd
            c.alignment = aln("right" if col in (5, 6) else "left")
        ws.cell(row=r, column=5).number_format = '#,##0.##'
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.row_dimensions[r].height = 18

        for col, w in enumerate([14, 32, 14, 28, 14, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read()).decode()

        fname = (
            f"Retours_Produits_"
            f"{self.date_debut.strftime('%d%m%Y')}_{self.date_fin.strftime('%d%m%Y')}.xlsx"
        )
        att = self.env['ir.attachment'].create({
            'name':      fname,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'new',
        }
