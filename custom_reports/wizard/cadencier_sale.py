# -*- coding: utf-8 -*-
import base64
import io
from odoo import fields, models, api
from datetime import date
from collections import defaultdict


class CadencierWizard(models.TransientModel):
    _name = 'cadencier.ventes.wizard'
    _description = 'Wizard Cadencier Stat Ventes Articles'

    year = fields.Selection(
        selection='_get_years',
        string="Année",
        required=True,
        default=lambda self: str(date.today().year),
    )
    company_id = fields.Many2one(
        'res.company',
        string="Société",
        required=True,
        default=lambda self: self.env.company,
        readonly=True,
    )
    famille_ids = fields.Many2many(
        'product.category',
        string="Familles (optionnel)",
        help="Laisser vide pour toutes les familles",
    )

    nbre_product = fields.Integer(
        string='Nombres articles',
        compute='_compute_nbre_product',
        store=False
    )
    

    @api.model
    def _get_years(self):
        current_year = date.today().year
        return [(str(y), str(y)) for y in range(current_year - 5, current_year + 2)]

    def action_print_report(self):
        self.ensure_one()
        return self.env.ref('custom_reports.action_report_cadencier').report_action(self)

    # ─────────────────────────────────────────────────────────────
    # EXPORT EXCEL
    # ─────────────────────────────────────────────────────────────
    def action_export_excel(self):
        """Génère le cadencier au format Excel et le retourne en téléchargement."""
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. Installez-la via : pip install openpyxl")

        lines = self._get_report_data(
            self.year,
            self.company_id.id,
            self.famille_ids.ids,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f"Cadencier {self.year}"

        # ── Couleurs ──────────────────────────────────────────────
        COLOR_DARK   = "1A1A2E"
        COLOR_CYAN   = "00C8E0"
        COLOR_ORANGE = "FF6B35"
        COLOR_GREY   = "C1B7B7"
        COLOR_LIGHT  = "D9F5F8"
        COLOR_TOTAL  = "FFF3EE"
        COLOR_SUB_BG = "0A3D4D"
        COLOR_WHITE  = "FFFFFF"
        COLOR_GREEN  = "2ECC71"
        COLOR_RED    = "E74C3C"

        thin = Side(style='thin', color=COLOR_DARK)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def font(bold=False, color=COLOR_DARK, size=9):
            return Font(name="Arial", bold=bold, color=color, size=size)

        def align(h="center", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        # ── Ligne 1 : titre société ────────────────────────────────
        ws.merge_cells("A1:V1")
        company = self.company_id
        ws["A1"] = (
            f"{company.name}"
            + (f"  –  {company.street}" if company.street else "")
            + (f", {company.city}" if company.city else "")
        )
        ws["A1"].font = font(bold=True, size=10)
        ws["A1"].alignment = align(h="left")

        # ── Ligne 2 : titre rapport ────────────────────────────────
        ws.merge_cells("A2:V2")
        ws["A2"] = (
            f"CADENCIER STAT VENTES ARTICLES  –  ANNEE {self.year}"
            f"  –  SOURCE : POS + MODULE VENTE"
        )
        ws["A2"].font = font(bold=True, color=COLOR_WHITE, size=12)
        ws["A2"].fill = fill(COLOR_DARK)
        ws["A2"].alignment = align()
        ws.row_dimensions[2].height = 20

        # ── Ligne 3 vide ───────────────────────────────────────────
        ws.append([])

        # ── En-têtes colonnes (ligne 4) ────────────────────────────
        MONTHS = ['JAN', 'FEV', 'MAR', 'AVR', 'MAI', 'JUIN',
                'JLT', 'AOT', 'SPT', 'OCT', 'NOV', 'DEC']
        headers = [
            'CODE', 'DESIGNATION', 'STA.', 'FAMILLE',
            'ST.DISP.', 'MAXI', 'CMD', 'MARG%', 'PVTC'
        ] + MONTHS + ['TOTAL']

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.alignment = align()
            cell.border = border_all
            if header in MONTHS:
                cell.fill = fill(COLOR_CYAN)
                cell.font = font(bold=True, color=COLOR_DARK, size=9)
            elif header == 'TOTAL':
                cell.fill = fill(COLOR_ORANGE)
                cell.font = font(bold=True, color=COLOR_WHITE, size=9)
            else:
                cell.fill = fill(COLOR_DARK)
                cell.font = font(bold=True, color=COLOR_WHITE, size=9)

        ws.row_dimensions[header_row].height = 18

        # ── Données ────────────────────────────────────────────────
        data_start_row = 5
        for row_offset, line in enumerate(lines):
            r = data_start_row + row_offset

            if line.get('is_subtotal'):
                # ── Sous-total famille ─────────────────────────────

                # Colonnes A:D — label fusionné
                label = f"▶ TOTAL  {line['code_famille']} — {line['famille']}"
                ws.merge_cells(f"A{r}:D{r}")
                ws[f"A{r}"] = label
                ws[f"A{r}"].font      = font(bold=True, color=COLOR_DARK, size=9)
                ws[f"A{r}"].fill      = fill(COLOR_GREY)
                ws[f"A{r}"].alignment = align(h="left")
                ws[f"A{r}"].border    = border_all

                # Colonnes 5 à 9 : ST.DISP, MAXI, CMD, MARG%, PVTC
                subtotal_cols = {
                    5: ('st_disp', '#,##0.00', False),  # ST.DISP
                    6: (None,      None,       False),  # MAXI  — vide
                    7: ('cmd',     '#,##0.00', False),  # CMD
                    8: ('marg',    '0.00"%"',  True),   # MARG% — formule agrégée
                    9: (None,      None,       False),  # PVTC  — vide
                }

                for c, (key, fmt, is_marg) in subtotal_cols.items():
                    cell = ws.cell(row=r, column=c)
                    cell.border    = border_all
                    cell.alignment = align()
                    cell.fill      = fill(COLOR_GREY)
                    if key and key in line and line[key] is not None:
                        val = line[key]
                        cell.value         = val
                        cell.number_format = fmt
                        if is_marg:
                            cell.font = font(
                                bold=True,
                                color=COLOR_GREEN if val >= 0 else COLOR_RED,
                                size=9
                            )
                        else:
                            cell.font = font(bold=True, color=COLOR_DARK, size=9)

                # Ventes mensuelles — colonnes 10 à 21
                for m_idx, qty in enumerate(line['ventes']):
                    c    = 10 + m_idx
                    cell = ws.cell(row=r, column=c)
                    cell.value         = qty if qty > 0 else None
                    cell.fill          = fill(COLOR_SUB_BG if qty > 0 else COLOR_GREY)
                    cell.font          = font(
                        bold=True,
                        color=COLOR_CYAN if qty > 0 else COLOR_DARK,
                        size=9
                    )
                    cell.alignment     = align()
                    cell.border        = border_all
                    cell.number_format = '#,##0'

                # TOTAL — colonne 22
                cell_total                = ws.cell(row=r, column=22)
                cell_total.value          = line['total']
                cell_total.font           = font(bold=True, color=COLOR_WHITE, size=9)
                cell_total.fill           = fill(COLOR_ORANGE)
                cell_total.alignment      = align()
                cell_total.border         = border_all
                cell_total.number_format  = '#,##0'

                ws.row_dimensions[r].height = 16

            else:
                # ── Ligne produit ──────────────────────────────────
                row_values = [
                    line['code'],
                    line['designation'],
                    line['sta'],
                    line['famille'],
                    line['st_disp'],
                    line['maxi'],
                    line['cmd'],
                    line['marg'],
                    line['pvtc'],
                ] + [v if v > 0 else 0 for v in line['ventes']] + [line['total']]

                for col_idx, value in enumerate(row_values, start=1):
                    cell           = ws.cell(row=r, column=col_idx, value=value)
                    cell.border    = border_all
                    cell.alignment = align(
                        h="left" if col_idx == 2 else "center",
                        wrap=(col_idx == 2)
                    )
                    cell.font = font(size=9)

                    if col_idx == 9:                        # PVTC
                        cell.number_format = '#,##0'

                    elif col_idx == 8:                      # MARG%
                        cell.number_format = '0.00"%"'
                        cell.font = font(
                            bold=False,
                            color=COLOR_GREEN if (value or 0) >= 0 else COLOR_RED,
                            size=9
                        )

                    elif col_idx in range(10, 22):          # Mois
                        cell.number_format = '#,##0'
                        if value and value > 0:
                            cell.fill = fill(COLOR_LIGHT)
                            cell.font = font(bold=True, size=9)

                    elif col_idx == 22:                     # TOTAL
                        cell.number_format = '#,##0'
                        if line['total'] > 0:
                            cell.fill = fill(COLOR_TOTAL)
                            cell.font = font(bold=True, color=COLOR_ORANGE, size=9)

                ws.row_dimensions[r].height = 15

        # ── Largeurs de colonnes ───────────────────────────────────
        col_widths = [
            9,   # CODE
            28,  # DESIGNATION
            5,   # STA
            14,  # FAMILLE
            8,   # ST.DISP
            7,   # MAXI
            7,   # CMD
            7,   # MARG%
            12,  # PVTC
        ] + [7] * 12 + [9]  # 12 mois + TOTAL

        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Figer les 4 premières colonnes et la ligne d'en-tête
        ws.freeze_panes = "E5"

        # ── Sauvegarde en mémoire ──────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        filename   = f"Cadencier_Ventes_{self.year}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })

        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    # ─────────────────────────────────────────────────────────────
    # DONNÉES RAPPORT
    # ─────────────────────────────────────────────────────────────
    def _compute_nbre_product(self):
        for record in self:
            data = record._get_report_data(
                year=record.year,
                company_id=record.company_id.id
            )
            record.nbre_product = len([
                x for x in data if not x.get('is_subtotal')
            ])


    # def _get_report_data(self, year, company_id, famille_ids=None):
    #     company = self.env['res.company'].browse(company_id)
    #     date_from = date(int(year), 1, 1)
    #     date_to   = date(int(year), 12, 31)
    #     product_data = defaultdict(lambda: defaultdict(float))

    #     products = self.env['product.template'].search([
    #         ('allowed_company_ids', 'in', company_id),
    #         ('type', '=', 'consu'),
    #         ('active', '=', True),
    #         ('cat_gestion_id.name', 'in', ['01', '02', '04', '05', '06', 'DI'])
    #     ])

    #     # ✅ Règle métier :
    #     #    - Statut C → toujours inclus
    #     #    - Statut D → seulement si stock > 0
    #     final_products = self.env['product.template']
    #     for p in products:
    #         code = p.current_company_status_id.code if p.current_company_status_id else False
    #         if code == 'C':
    #             final_products |= p
    #         elif code == 'D':
    #             if any(v.qty_available > 0 for v in p.product_variant_ids):
    #                 final_products |= p
    #         # Autres statuts → exclus

    #     for p in final_products:
    #         for variant in p.product_variant_ids:
    #             product_data[variant.id]

    #     # ✅ IDs valides pré-calculés — utilisés pour filtrer les lignes sans boucle
    #     valid_product_ids = set(product_data.keys())

    #     if not valid_product_ids:
    #         return []

    #     # ── Domaines ─────────────────────────────────────────────
    #     date_from_str = str(date_from)
    #     date_to_str   = str(date_to)
    #     cat_filter    = ['01', '02', '04', '05', '06', 'DI']

    #     common_filters = [
    #         ('product_id', 'in', list(valid_product_ids)),  # ✅ filtre direct sur IDs
    #         ('product_id.type', '=', 'consu'),
    #         ('product_id.cat_gestion_id.name', 'in', cat_filter),
    #         ('product_id.active', '=', True),
    #     ]

    #     sale_domain = [
    #         ('order_id.state', 'in', ['sale', 'done']),
    #         ('order_id.company_id', '=', company_id),
    #         ('order_id.date_order', '>=', date_from_str),
    #         ('order_id.date_order', '<=', date_to_str),
    #     ] + common_filters

    #     pos_domain = [
    #         ('order_id.state', 'in', ['done', 'paid', 'invoiced']),
    #         ('order_id.company_id', '=', company_id),
    #         ('order_id.date_order', '>=', date_from_str),
    #         ('order_id.date_order', '<=', date_to_str),
    #     ] + common_filters

    #     sale_lines = self.env['sale.order.line'].with_context(company_id=company_id).search(sale_domain)
    #     pos_lines  = self.env['pos.order.line'].with_context(company_id=company_id).search(pos_domain)

    #     # ✅ Plus besoin de _filter_lines — le domaine SQL fait déjà le travail

    #     # ── Pré-agrégation ────────────────────────────────────────
    #     ca_by_product = defaultdict(float)

    #     for line in sale_lines:
    #         month_idx = line.order_id.date_order.month - 1
    #         product_data[line.product_id.id][month_idx] += line.product_uom_qty
    #         ca_by_product[line.product_id.id] += line.price_subtotal

    #     for line in pos_lines:
    #         month_idx = line.order_id.date_order.month - 1
    #         product_data[line.product_id.id][month_idx] += line.qty
    #         ca_by_product[line.product_id.id] += line.price_subtotal

    #     product_ids = list(product_data.keys())
    #     products    = self.env['product.product'].browse(product_ids)

    #     if famille_ids:
    #         products = products.filtered(lambda p: p.categ_id.id in famille_ids)

    #     # ── Pré-calculs ───────────────────────────────────────────
    #     stock_by_product = {
    #         p.id: p.with_company(company).qty_available
    #         for p in products
    #     }

    #     ttc_by_product = {}
    #     for product in products:
    #         taxes = product.taxes_id.compute_all(
    #             product.list_price,
    #             currency=product.currency_id,
    #             quantity=1,
    #             product=product,
    #             partner=None,
    #         )
    #         ttc_by_product[product.id] = taxes['total_included']

    #     # ── Construction résultat ─────────────────────────────────
    #     result = []
    #     for product in products.sorted(
    #         key=lambda p: ((p.categ_id.code or '').lower(), (p.default_code or '').lower())
    #     ):
    #         monthly_qtys = product_data[product.id]
    #         ventes = [round(monthly_qtys.get(i, 0), 2) for i in range(12)]
    #         total  = sum(ventes)

    #         stock      = stock_by_product.get(product.id, 0.0)
    #         pmp        = product.avg_cost or product.standard_price
    #         total_ca   = ca_by_product.get(product.id, 0.0)
    #         total_cost = total * pmp
    #         taux_marge = (
    #             round((total_ca - total_cost) / total_ca * 100, 2)
    #             if total_ca > 0 else 0.0
    #         )

    #         result.append({
    #             'code':         product.default_code or '',
    #             'designation':  product.name,
    #             'sta':          product.current_company_status_id.code if product.current_company_status_id else '',
    #             'maxi':         product.max_qty_orderpoint,
    #             'cmd':          product.pending_reception_qty,
    #             'marg':         taux_marge,
    #             'famille':      product.categ_id.name,
    #             'code_famille': product.categ_id.code,
    #             'st_disp':      round(stock, 2),
    #             'pvtc':         ttc_by_product.get(product.id, 0.0),
    #             'ventes':       ventes,
    #             'total':        total,
    #             # ✅ Stocker ca et cost pour le calcul agrégé du sous-total
    #             '_ca':          total_ca,
    #             '_cost':        total_cost,
    #         })

    #     # ── Sous-totaux ───────────────────────────────────────────
    #     final_result = []
    #     current_famille_code = None
    #     famille_ventes = [0.0] * 12
    #     famille_total  = 0.0
    #     famille_label  = ''
    #     famille_st_disp = 0.0   # ✅
    #     famille_cmd     = 0.0   # ✅
    #     famille_ca      = 0.0   # ✅ pour la formule de marge agrégée
    #     famille_cost    = 0.0   # ✅

    #     def _compute_subtotal(label, code, ventes, total, st_disp, cmd, ca, cost):
    #         marge = round((ca - cost) / ca * 100, 2) if ca > 0 else 0.0
    #         return {
    #             'is_subtotal':  True,
    #             'famille':      label,
    #             'code_famille': code,
    #             'marg':         marge,       # ✅ formule agrégée
    #             'st_disp':      round(st_disp, 2),  # ✅ total st_disp
    #             'cmd':          round(cmd, 2),      # ✅ total cmd
    #             'ventes':       [round(v, 2) for v in ventes],
    #             'total':        round(total, 2),
    #         }

    #     for item in result:
    #         item['is_subtotal'] = False
    #         if current_famille_code is not None and item['code_famille'] != current_famille_code:
    #             final_result.append(_compute_subtotal(
    #                 famille_label, current_famille_code,
    #                 famille_ventes, famille_total,
    #                 famille_st_disp, famille_cmd,
    #                 famille_ca, famille_cost,
    #             ))
    #             famille_ventes  = [0.0] * 12
    #             famille_total   = 0.0
    #             famille_st_disp = 0.0
    #             famille_cmd     = 0.0
    #             famille_ca      = 0.0
    #             famille_cost    = 0.0

    #         current_famille_code = item['code_famille']
    #         famille_label        = item['famille']
    #         final_result.append(item)
    #         for i in range(12):
    #             famille_ventes[i] += item['ventes'][i]
    #         famille_total   += item['total']
    #         famille_st_disp += item['st_disp']   # ✅
    #         famille_cmd     += item['cmd']       # ✅
    #         famille_ca      += item['_ca']       # ✅
    #         famille_cost    += item['_cost']     # ✅

    #     if current_famille_code is not None:
    #         final_result.append(_compute_subtotal(
    #             famille_label, current_famille_code,
    #             famille_ventes, famille_total,
    #             famille_st_disp, famille_cmd,
    #             famille_ca, famille_cost,
    #         ))

    #     return final_result


    def _get_report_data(self, year, company_id, famille_ids=None):
        company = self.env['res.company'].browse(company_id)
        date_from = date(int(year), 1, 1)
        date_to   = date(int(year), 12, 31)
        product_data = defaultdict(lambda: defaultdict(float))

        products = self.env['product.template'].search([
            ('allowed_company_ids', 'in', company_id),
            ('type', '=', 'consu'),
            ('active', '=', True),
            ('cat_gestion_id.name', 'in', ['01', '02', '04', '05', '06', 'DI'])
        ])

        # ✅ Règle métier :
        #    - Statut C → toujours inclus
        #    - Statut D → seulement si stock > 0
        final_products = self.env['product.template']
        for p in products:
            code = p.current_company_status_id.code if p.current_company_status_id else False
            if code == 'C':
                final_products |= p
            elif code == 'D':
                if any(v.qty_available > 0 for v in p.product_variant_ids):
                    final_products |= p
            # Autres statuts → exclus

        for p in final_products:
            for variant in p.product_variant_ids:
                product_data[variant.id]

        # ✅ IDs valides pré-calculés — utilisés pour filtrer les lignes sans boucle
        valid_product_ids = set(product_data.keys())

        if not valid_product_ids:
            return []

        # ── Domaines ─────────────────────────────────────────────
        date_from_str = str(date_from)
        date_to_str   = str(date_to)
        cat_filter    = ['01', '02', '04', '05', '06', 'DI']

        common_filters = [
            ('product_id', 'in', list(valid_product_ids)),
            ('product_id.type', '=', 'consu'),
            ('product_id.cat_gestion_id.name', 'in', cat_filter),
            ('product_id.active', '=', True),
        ]

        sale_domain = [
            ('order_id.state', 'in', ['sale', 'done']),
            ('order_id.company_id', '=', company_id),
            ('order_id.date_order', '>=', date_from_str),
            ('order_id.date_order', '<=', date_to_str),
        ] + common_filters

        pos_domain = [
            ('order_id.state', 'in', ['done', 'paid', 'invoiced']),
            ('order_id.company_id', '=', company_id),
            ('order_id.date_order', '>=', date_from_str),
            ('order_id.date_order', '<=', date_to_str),
        ] + common_filters

        sale_lines = self.env['sale.order.line'].with_context(company_id=company_id).search(sale_domain)
        pos_lines  = self.env['pos.order.line'].with_context(company_id=company_id).search(pos_domain)

        # ── Pré-agrégation ────────────────────────────────────────
        ca_by_product     = defaultdict(float)
        margin_by_product = defaultdict(float)  # ✅ marge brute directe depuis les lignes

        for line in sale_lines:
            month_idx = line.order_id.date_order.month - 1
            product_data[line.product_id.id][month_idx] += line.product_uom_qty
            ca_by_product[line.product_id.id]     += line.price_subtotal
            margin_by_product[line.product_id.id] += line.margin  # ✅ champ natif SO

        for line in pos_lines:
            month_idx = line.order_id.date_order.month - 1
            product_data[line.product_id.id][month_idx] += line.qty
            ca_by_product[line.product_id.id]     += line.price_subtotal
            margin_by_product[line.product_id.id] += line.margin  # ✅ champ natif POS

        product_ids = list(product_data.keys())
        products    = self.env['product.product'].browse(product_ids)

        if famille_ids:
            products = products.filtered(lambda p: p.categ_id.id in famille_ids)

        # ── Pré-calculs ───────────────────────────────────────────
        stock_by_product = {
            p.id: p.with_company(company).qty_available
            for p in products
        }

        ttc_by_product = {}
        for product in products:
            taxes = product.taxes_id.compute_all(
                product.list_price,
                currency=product.currency_id,
                quantity=1,
                product=product,
                partner=None,
            )
            ttc_by_product[product.id] = taxes['total_included']

        # ── Construction résultat ─────────────────────────────────
        result = []
        for product in products.sorted(
            key=lambda p: ((p.categ_id.code or '').lower(), (p.default_code or '').lower())
        ):
            monthly_qtys = product_data[product.id]
            ventes = [round(monthly_qtys.get(i, 0), 2) for i in range(12)]
            total  = sum(ventes)

            stock        = stock_by_product.get(product.id, 0.0)
            total_ca     = ca_by_product.get(product.id, 0.0)
            total_margin = margin_by_product.get(product.id, 0.0)  # ✅
            taux_marge   = (
                round(total_margin / total_ca * 100, 2)
                if total_ca > 0 else 0.0
            )

            result.append({
                'code':         product.default_code or '',
                'designation':  product.name,
                'sta':          product.current_company_status_id.code if product.current_company_status_id else '',
                'maxi':         product.max_qty_orderpoint,
                'cmd':          product.pending_reception_qty,
                'marg':         taux_marge,
                'famille':      product.categ_id.name,
                'code_famille': product.categ_id.code,
                'st_disp':      round(stock, 2),
                'pvtc':         ttc_by_product.get(product.id, 0.0),
                'ventes':       ventes,
                'total':        total,
                '_ca':          total_ca,
                '_margin':      total_margin,  # ✅ remplace _cost
            })

        # ── Sous-totaux ───────────────────────────────────────────
        final_result = []
        current_famille_code = None
        famille_ventes  = [0.0] * 12
        famille_total   = 0.0
        famille_label   = ''
        famille_st_disp = 0.0
        famille_cmd     = 0.0
        famille_ca      = 0.0
        famille_margin  = 0.0  # ✅ remplace famille_cost

        def _compute_subtotal(label, code, ventes, total, st_disp, cmd, ca, margin):
            marge = round(margin / ca * 100, 2) if ca > 0 else 0.0  # ✅
            return {
                'is_subtotal':  True,
                'famille':      label,
                'code_famille': code,
                'marg':         marge,
                'st_disp':      round(st_disp, 2),
                'cmd':          round(cmd, 2),
                'ventes':       [round(v, 2) for v in ventes],
                'total':        round(total, 2),
            }

        for item in result:
            item['is_subtotal'] = False
            if current_famille_code is not None and item['code_famille'] != current_famille_code:
                final_result.append(_compute_subtotal(
                    famille_label, current_famille_code,
                    famille_ventes, famille_total,
                    famille_st_disp, famille_cmd,
                    famille_ca, famille_margin,  # ✅
                ))
                famille_ventes  = [0.0] * 12
                famille_total   = 0.0
                famille_st_disp = 0.0
                famille_cmd     = 0.0
                famille_ca      = 0.0
                famille_margin  = 0.0  # ✅

            current_famille_code = item['code_famille']
            famille_label        = item['famille']
            final_result.append(item)
            for i in range(12):
                famille_ventes[i] += item['ventes'][i]
            famille_total   += item['total']
            famille_st_disp += item['st_disp']
            famille_cmd     += item['cmd']
            famille_ca      += item['_ca']
            famille_margin  += item['_margin']  # ✅

        if current_famille_code is not None:
            final_result.append(_compute_subtotal(
                famille_label, current_famille_code,
                famille_ventes, famille_total,
                famille_st_disp, famille_cmd,
                famille_ca, famille_margin,  # ✅
            ))

        return final_result


class ReportCadencierVentes(models.AbstractModel):
    _name = 'report.custom_reports.report_cadencier_template'
    _description = 'Rapport Cadencier Ventes'

    @api.model
    def _get_report_values(self, docids, data=None):
        wizard = self.env['cadencier.ventes.wizard'].browse(docids)
        lines = wizard._get_report_data(
            wizard.year,
            wizard.company_id.id,
            wizard.famille_ids.ids,
        )
        return {
            'doc': wizard,
            'company': wizard.company_id,
            'year': wizard.year,
            'lines': lines,
        }