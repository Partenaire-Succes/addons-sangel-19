# -*- coding: utf-8 -*-
import io
import base64

from odoo import models, fields, api
from odoo.exceptions import UserError
from odoo.tools import float_round


class StockMovementReportWizard(models.TransientModel):
    _name = 'stock.movement.report.wizard'
    _description = 'Rapport Stock et Mouvements entre deux dates'

    date_from = fields.Datetime(
        string='Date A',
        required=True,
        default=fields.Datetime.now,
        help="Stock reconstitué à cet instant précis.",
    )
    date_to = fields.Datetime(
        string='Date B',
        required=True,
        default=fields.Datetime.now,
        help="Stock reconstitué à cet instant précis.",
    )
    location_id = fields.Many2one(
        'stock.location',
        string='Emplacement',
        required=True,
        domain=[('usage', '=', 'internal')],
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )
    category_ids = fields.Many2many(
        'product.category.x3',
        string='Niveau 5',
    )

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise UserError("La Date A doit être antérieure ou égale à la Date B.")

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'location_id' in fields_list and not res.get('location_id'):
            location = self.env['stock.location'].search([
                ('usage', '=', 'internal'),
                '|',
                ('name', 'ilike', 'Stock'),
                ('complete_name', 'ilike', '/Stock'),
            ], limit=1)
            if not location:
                warehouse = self.env['stock.warehouse'].search([
                    ('company_id', '=', self.env.company.id),
                ], limit=1)
                if warehouse:
                    location = warehouse.lot_stock_id
            if location:
                res['location_id'] = location.id
        return res

    # -------------------------------------------------------------------------
    # LOGIQUE DU RAPPORT
    # -------------------------------------------------------------------------
    def _get_products(self):
        self.ensure_one()
        domain = [
            ('product_tmpl_id.allowed_company_ids', 'in', [self.company_id.id]),
            ('product_tmpl_id.type', '=', 'consu'),
            ('product_tmpl_id.prod_type_x3_id.name', '=', 'TS'),
        ]
        if self.category_ids:
            domain.append(('cat_gestion_id', 'in', self.category_ids.ids))
        return self.env['product.product'].search(domain)

    def _get_stock_at(self, product_ids, at):
        """Reconstitue le stock à l'instant `at` pour `product_ids`, à
        l'emplacement sélectionné, à partir des stock.move.line validées."""
        if not product_ids:
            return {}
        self.env.cr.execute("""
            SELECT product_id,
                   SUM(CASE WHEN location_dest_id = %s THEN quantity ELSE -quantity END) AS qty
            FROM stock_move_line
            WHERE product_id = ANY(%s)
              AND state = 'done'
              AND date <= %s
              AND (location_id = %s OR location_dest_id = %s)
            GROUP BY product_id
        """, [self.location_id.id, list(product_ids), at, self.location_id.id, self.location_id.id])
        return {row[0]: row[1] for row in self.env.cr.fetchall()}

    def _get_movements(self, product_ids):
        """Entrées / sorties entre Date A (exclue) et Date B (incluse), à
        l'emplacement sélectionné. Le net (entrée - sortie) correspond
        exactement aux "Mvts" tels que calculés auparavant.

        Exclut les ajustements d'inventaire (stock.move.is_inventory=True,
        posés par stock.quant._apply_inventory : inventaires physiques
        journaliers, éclatement carton/sachet...) — on ne veut ici que les
        mouvements "métier" (ventes, réceptions, rebut, transferts...)."""
        if not product_ids:
            return {}
        self.env.cr.execute("""
            SELECT sml.product_id,
                   SUM(CASE WHEN sml.location_dest_id = %s
                       THEN sml.quantity ELSE 0 END) AS entree,
                   SUM(CASE WHEN sml.location_id = %s AND sml.location_dest_id != %s
                       THEN sml.quantity ELSE 0 END) AS sortie
            FROM stock_move_line sml
            LEFT JOIN stock_move sm ON sm.id = sml.move_id
            WHERE sml.product_id = ANY(%s)
              AND sml.state = 'done'
              AND sml.date > %s
              AND sml.date <= %s
              AND (sml.location_id = %s OR sml.location_dest_id = %s)
              AND COALESCE(sm.is_inventory, false) = false
            GROUP BY sml.product_id
        """, [
            self.location_id.id, self.location_id.id, self.location_id.id,
            list(product_ids), self.date_from, self.date_to,
            self.location_id.id, self.location_id.id,
        ])
        return {row[0]: (row[1], row[2]) for row in self.env.cr.fetchall()}

    def _get_last_inventory_dates(self, tmpl_ids):
        """Date du dernier inventaire physique journalier validé (date_done)
        par produit (product_tmpl_id), toutes dates confondues.

        On regroupe sur product_tmpl_id et non product_id : ce dernier est
        recalculé sur product_tmpl_id.product_variant_id (voir
        PhysicalInventoryLine._compute_product_id) et peut être vide ou
        obsolète (variante archivée/changée) alors que product_tmpl_id reste
        toujours renseigné."""
        if not tmpl_ids:
            return {}
        self.env.cr.execute("""
            SELECT pil.product_tmpl_id, MAX(pi.date_done) AS last_date
            FROM physical_inventory_line pil
            JOIN physical_inventory pi ON pi.id = pil.inventory_physical_id
            WHERE pil.product_tmpl_id = ANY(%s)
              AND pil.active = true
              AND pi.company_id = %s
              AND pi.state = 'done'
            GROUP BY pil.product_tmpl_id
        """, [list(tmpl_ids), self.company_id.id])
        return {row[0]: row[1] for row in self.env.cr.fetchall()}

    def _get_cumulative_gap(self, tmpl_ids):
        """Écart cumulé (qty_diff) et écart cumulé valorisé (valorisation) :
        somme des lignes d'inventaire physique journalier validées (toutes
        dates confondues) par produit (product_tmpl_id, cf. note ci-dessus
        sur la fiabilité de product_id)."""
        if not tmpl_ids:
            return {}, {}
        groups = self.env['physical.inventory.line'].read_group(
            domain=[
                ('product_tmpl_id', 'in', tmpl_ids),
                ('company_id', '=', self.company_id.id),
                ('active', '=', True),
                ('state', '=', 'done'),
            ],
            fields=['qty_diff:sum', 'valorisation:sum'],
            groupby=['product_tmpl_id'],
        )
        cumulative_gap = {g['product_tmpl_id'][0]: g['qty_diff'] for g in groups}
        cumulative_gap_valued = {g['product_tmpl_id'][0]: g['valorisation'] for g in groups}
        return cumulative_gap, cumulative_gap_valued

    def _get_report_lines(self):
        self.ensure_one()
        products = self._get_products()
        product_ids = products.ids

        tmpl_ids = products.mapped('product_tmpl_id').ids

        stock_a = self._get_stock_at(product_ids, self.date_from)
        stock_b = self._get_stock_at(product_ids, self.date_to)
        movements = self._get_movements(product_ids)
        cumulative_gap, cumulative_gap_valued = self._get_cumulative_gap(tmpl_ids)
        last_inventory_dates = self._get_last_inventory_dates(tmpl_ids)

        lines = []
        for product in products:
            qty_a = stock_a.get(product.id, 0.0)
            qty_b = stock_b.get(product.id, 0.0)
            entree, sortie = movements.get(product.id, (0.0, 0.0))
            mvt = entree - sortie
            gap = cumulative_gap.get(product.product_tmpl_id.id, 0.0)
            gap_valued = cumulative_gap_valued.get(product.product_tmpl_id.id, 0.0)
            last_inventory_date = last_inventory_dates.get(product.product_tmpl_id.id)

            if not qty_a and not qty_b and not mvt and not gap:
                continue

            lines.append({
                'code_article': product.code_article or product.product_tmpl_id.code_article or '',
                'name': product.name,
                'last_inventory_date': last_inventory_date,
                'cost': float_round(product.with_company(self.company_id).standard_price or 0.0, 2),
                'stock_a': float_round(qty_a, 2),
                'stock_b': float_round(qty_b, 2),
                'entree': float_round(entree, 2),
                'sortie': float_round(sortie, 2),
                'movement': float_round(mvt, 2),
                'ecart_cumule': float_round(gap, 2),
                'ecart_cumule_valorise': float_round(gap_valued, 2),
            })

        return sorted(lines, key=lambda l: l['code_article'])

    # -------------------------------------------------------------------------
    # ACTIONS
    # -------------------------------------------------------------------------
    def action_print_report(self):
        self.ensure_one()
        report_action = self.env.ref('custom_reports.action_report_stock_movement').report_action(self)
        report_action['close_on_report_download'] = True
        return report_action

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        lines = self._get_report_lines()
        if not lines:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock et Mouvements"

        BLUE = "1A5276"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        ws.merge_cells("A1:K1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        ws.merge_cells("A2:K2")
        ws["A2"] = (
            f"STOCK ET MOUVEMENTS — {self.location_id.complete_name} — "
            f"{self.date_from.strftime('%d/%m/%Y %H:%M')} au {self.date_to.strftime('%d/%m/%Y %H:%M')}"
        )
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        headers = [
            "Code Article", "Nom produit", "Date dernier inventaire",
            "Coût",
            f"Stock {self.date_from.strftime('%d/%m/%Y %H:%M')}",
            f"Stock {self.date_to.strftime('%d/%m/%Y %H:%M')}",
            "Entrée", "Sortie",
            "Mvts (Entrée-Sortie)",
            "Ecart cumulé",
            "Ecart cumulé valorisé",
        ]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 12):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        for line in lines:
            last_inv = (
                line['last_inventory_date'].strftime('%d/%m/%Y %H:%M')
                if line['last_inventory_date'] else ''
            )
            ws.append([
                line['code_article'], line['name'], last_inv, line['cost'],
                line['stock_a'], line['stock_b'],
                line['entree'], line['sortie'], line['movement'], line['ecart_cumule'],
                line['ecart_cumule_valorise'],
            ])
            r = ws.max_row
            for col in range(1, 12):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", size=9); c.border = brd
                c.alignment = aln("right" if col >= 4 else "left")
            for col in (4, 5, 6, 7, 8, 9, 10, 11):
                ws.cell(row=r, column=col).number_format = '#,##0.00'

        for col, width in enumerate([14, 34, 20, 14, 20, 20, 14, 14, 18, 14, 18], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = (
            f"Stock_Mouvements_{self.date_from.strftime('%Y%m%d_%Hh%M')}_"
            f"{self.date_to.strftime('%Y%m%d_%Hh%M')}.xlsx"
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename, 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}
