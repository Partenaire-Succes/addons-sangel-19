import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class CumulInventaryReportWizard(models.TransientModel):
    _name = 'cumul.inventary.report.wizard'
    _description = 'Cumul Inventaire'

    date_from = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today
    )
    date_to = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company
    )

    physical_lines_ids = fields.Many2many(
        'physical.inventory.line',
        string='Lignes physiques',
        compute='_compute_physical_lines_ids'
    )

    code_article_filter = fields.Char(
        string='Filtrer par Code Article',
        help='Laissez vide pour afficher tous les articles'
    )

    @api.depends('date_from', 'date_to', 'company_id', 'code_article_filter')
    def _compute_physical_lines_ids(self):
        for record in self:
            domain = [
                ('create_date', '>=', record.date_from),
                ('create_date', '<=', record.date_to),
                ('company_id', '=', record.company_id.id),
                ('active', '=', True),
            ]

            # Filtrer par code article si spécifié
            if record.code_article_filter:
                domain.append(('code_article', 'ilike', record.code_article_filter))

            record.physical_lines_ids = self.env['physical.inventory.line'].search(
                domain,
                order='code_article, create_date'
            )

    @api.depends('physical_lines_ids')
    def _compute_grouped_data(self):
        """Grouper les données par article et calculer les totaux"""
        for record in self:
            grouped = {}
            for line in record.physical_lines_ids:
                key = (line.code_article, line.product_tmpl_id.id)
                if key not in grouped:
                    grouped[key] = {
                        'code_article': line.code_article,
                        'designation': line.product_tmpl_id.name,
                        'lines': [],
                        'total_ecart': 0.0,
                        'total_montant': 0.0,
                    }

                grouped[key]['lines'].append(line)
                grouped[key]['total_ecart'] += line.qty_diff or 0.0
                grouped[key]['total_montant'] += (line.qty_diff or 0.0) * (line.price or 0.0)

            record.grouped_data = str(grouped)

    def _get_grouped_lines(self):
        """
        Méthode pour grouper les lignes par article (code + désignation)
        Utilisée dans le template QWeb
        """
        self.ensure_one()
        grouped = {}

        for line in self.physical_lines_ids:
            # Créer une clé unique pour chaque article
            key = (line.code_article or '', line.product_tmpl_id.id)

            if key not in grouped:
                grouped[key] = {
                    'code_article': line.code_article or '',
                    'designation': line.product_tmpl_id.name or '',
                    'lines': [],
                    'total_ecart': 0.0,
                    'total_montant': 0.0,
                }

            grouped[key]['lines'].append(line)
            grouped[key]['total_ecart'] += line.qty_diff or 0.0
            grouped[key]['total_montant'] += (line.qty_diff or 0.0) * (line.price or 0.0)

        # Convertir en liste et trier par code article
        result = sorted(grouped.values(), key=lambda x: x['code_article'])
        return result

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    def action_print_report(self):
        return self.env.ref('custom_reports.action_report_cumul_inventaire').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        if not self.physical_lines_ids:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        wb = Workbook(); ws = wb.active; ws.title = "Cumul Inventaire"
        BLUE = "1A5276"; LBLUE = "D6EAF8"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        ws.merge_cells("A1:G1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)
        ws.merge_cells("A2:G2")
        ws["A2"] = f"CUMUL INVENTAIRE — Du {self.date_from.strftime('%d/%m/%Y')} au {self.date_to.strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE); ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18; ws.append([])

        headers = ["Code Article", "Désignation", "N° Inv.", "Date", "Écart", "PAMP", "Montant"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        grouped = self._get_grouped_lines()
        total_montant = 0.0
        for group in grouped:
            for line in group['lines']:
                ws.append([
                    group['code_article'],
                    group['designation'],
                    line.id,
                    line.create_date.strftime('%d/%m/%Y') if line.create_date else '',
                    line.qty_diff or 0.0,
                    line.price or 0.0,
                    (line.qty_diff or 0.0) * (line.price or 0.0),
                ])
                r = ws.max_row
                for col in range(1, 8):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9); c.border = brd
                    c.alignment = aln("right" if col >= 5 else "center" if col == 4 else "left")
                for col in (5, 6, 7):
                    ws.cell(row=r, column=col).number_format = '#,##0.00'

            ws.append(["", group['designation'], "", "", group['total_ecart'], "", group['total_montant']])
            r = ws.max_row
            for col in range(1, 8):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE); c.border = brd
                c.alignment = aln("right" if col >= 5 else "left")
            for col in (5, 7):
                ws.cell(row=r, column=col).number_format = '#,##0.00'
            total_montant += group['total_montant']

        ws.append(["", "", "", "", "", "TOTAL GÉNÉRAL", total_montant])
        r = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.border = brd
            c.alignment = aln("right" if col >= 6 else "left")
        ws.cell(row=r, column=7).number_format = '#,##0.00'

        for col, width in enumerate([14, 30, 10, 13, 12, 14, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = f"Cumul_Inventaire_{self.date_from.strftime('%d%m%Y')}_{self.date_to.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename, 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}
