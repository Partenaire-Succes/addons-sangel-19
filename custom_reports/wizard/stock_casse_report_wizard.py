# -*- coding: utf-8 -*-
import io
import base64
from odoo import api, fields, models
from odoo.exceptions import UserError


class StockCasseReport(models.TransientModel):
    _name = 'stock.casse.report'
    _description = 'Rapport de casse'

    date_from = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today
    )
    date_to = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )

    scrap_reason_tag_ids = fields.Many2many(
        'stock.scrap.reason.tag',
        string='Raisons de casse',
    )

    scrap_lines_ids = fields.Many2many(
        'stock.scrap',
        string='Lignes de casse',
        compute='_compute_scrap_lines_ids'
    )

    @api.depends('date_from', 'date_to', 'company_id')
    def _compute_scrap_lines_ids(self):
        for record in self:
            scraps = self.env['stock.scrap'].search([
                ('date_done', '>=', record.date_from),
                ('date_done', '<=', record.date_to),
                ('company_id', '=', record.company_id.id),
                ('state', '=', 'done'),
                ('scrap_reason_tag_ids', 'in', record.scrap_reason_tag_ids.ids),
            ], order='date_done, name')
            # Consider 'CASSE' as scraps whose source location is not internal
        #    record.scrap_lines_ids = scraps.filtered(lambda s: s.location_id.usage != 'internal')
            record.scrap_lines_ids = scraps

    def action_print_report(self):
        self.ensure_one()
        if not self.scrap_lines_ids:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")
        return self.env.ref('custom_reports.action_report_casse').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        if not self.scrap_lines_ids:
            raise UserError("Aucune donnée trouvée pour la période sélectionnée.")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        wb = Workbook(); ws = wb.active; ws.title = "Casse"
        BLUE = "1A5276"; LBLUE = "D6EAF8"; WHITE = "FFFFFF"
        thin = Side(style='thin', color="AAAAAA")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)
        def aln(h="left"): return Alignment(horizontal=h, vertical="center")

        ws.merge_cells("A1:G1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)
        ws.merge_cells("A2:G2")
        ws["A2"] = f"STOCK CASSE — Du {self.date_from.strftime('%d/%m/%Y')} au {self.date_to.strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE); ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18; ws.append([])

        headers = ["N° Document", "Date", "Code Article", "Désignation", "Raisons", "Quantité", "Total PA"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.alignment = aln("center"); c.border = brd

        grand_total = 0.0
        for scrap in self.scrap_lines_ids:
            total_pa = scrap.scrap_qty * (scrap.product_id.standard_price or 0.0)
            reasons = ', '.join(scrap.scrap_reason_tag_ids.mapped('name')) if scrap.scrap_reason_tag_ids else ''
            ws.append([
                scrap.name,
                scrap.date_done.strftime('%d/%m/%Y') if scrap.date_done else '',
                scrap.product_id.default_code or '',
                scrap.product_id.name or '',
                reasons,
                scrap.scrap_qty,
                total_pa,
            ])
            r = ws.max_row
            for col in range(1, 8):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", size=9); c.border = brd
                c.alignment = aln("right" if col >= 6 else "center" if col == 2 else "left")
            ws.cell(row=r, column=7).number_format = '#,##0.00'
            grand_total += total_pa

        ws.append(["", "", "", "", "", "TOTAL GÉNÉRAL", grand_total])
        r = ws.max_row
        for col in range(1, 8):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE); c.border = brd
            c.alignment = aln("right" if col >= 6 else "left")
        ws.cell(row=r, column=7).number_format = '#,##0.00'

        for col, width in enumerate([16, 13, 14, 30, 20, 10, 14], 1):
            ws.column_dimensions[chr(64 + col)].width = width

        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = f"Casse_{self.date_from.strftime('%d%m%Y')}_{self.date_to.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename, 'type': 'binary', 'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name, 'res_id': self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new'}

