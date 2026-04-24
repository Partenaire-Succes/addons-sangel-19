# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class LivraisonReportWizard(models.TransientModel):
    _name = 'livraison.report.wizard'
    _description = 'Rapport Récapitulatif Livraisons'

    date_debut = fields.Date(
        string='Date de début',
        required=True,
        default=fields.Date.context_today,
    )
    date_fin = fields.Date(
        string='Date de fin',
        required=True,
        default=fields.Date.context_today,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )
    client_ids = fields.Many2many(
        'res.partner',
        'livraison_report_partner_rel',
        string='Clients',
        domain=[('customer_rank', '>', 0)],
        help='Laissez vide pour inclure tous les clients',
    )

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for rec in self:
            if rec.date_debut > rec.date_fin:
                raise UserError("La date de début doit être antérieure à la date de fin.")

    def _get_livraisons_data(self):
        self.ensure_one()
        domain = [
            ('picking_type_code', '=', 'outgoing'),
            ('state', '=', 'done'),
            ('date_done', '>=', self.date_debut),
            ('date_done', '<=', self.date_fin),
            ('company_id', '=', self.company_id.id),
        ]
        if self.client_ids:
            domain.append(('partner_id', 'in', self.client_ids.ids))

        pickings = self.env['stock.picking'].search(domain, order='partner_id, date_done')

        if not pickings:
            raise UserError("Aucune livraison trouvée pour la période sélectionnée.")

        clients_data = {}
        for picking in pickings:
            partner = picking.partner_id
            key = partner.id if partner else 0
            if key not in clients_data:
                clients_data[key] = {
                    'partner': partner,
                    'livraisons': [],
                    'total': 0.0,
                }

            total_picking = 0.0
            for move in picking.move_ids.filtered(lambda m: m.state == 'done'):
                qty_done = (
                    sum(move.move_line_ids.mapped('quantity'))
                    or move.product_uom_qty
                )
                total_picking += qty_done * (move.price_unit or 0.0)

            clients_data[key]['livraisons'].append({
                'date': picking.date_done,
                'numero': picking.name,
                'ref': picking.origin or '',
                'total': total_picking,
            })
            clients_data[key]['total'] += total_picking

        return list(clients_data.values())

    def action_print_report(self):
        self.ensure_one()
        self._get_livraisons_data()
        return self.env.ref('custom_reports.action_report_livraison').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        clients_data = self._get_livraisons_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Livraisons"

        BLUE  = "1A5276"
        LBLUE = "D6EAF8"
        WHITE = "FFFFFF"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(h):
            return PatternFill("solid", fgColor=h)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        ws.merge_cells("A1:F1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = Font(name="Arial", bold=True, size=11)

        ws.merge_cells("A2:F2")
        ws["A2"] = (
            f"RÉCAPITULATIF LIVRAISONS — "
            f"Du {self.date_debut.strftime('%d/%m/%Y')} au {self.date_fin.strftime('%d/%m/%Y')}"
        )
        ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        headers = ["Date", "N° Livraison", "Client", "N° Commande / Réf.", "Montant Total"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, 6):
            c = ws.cell(row=hrow, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.alignment = aln("center")
            c.border = brd

        grand_total = 0.0
        for cdata in clients_data:
            pname = cdata['partner'].name if cdata['partner'] else '—'
            for liv in cdata['livraisons']:
                ws.append([
                    liv['date'].strftime('%d/%m/%Y') if liv['date'] else '',
                    liv['numero'],
                    pname,
                    liv['ref'],
                    liv['total'],
                ])
                r = ws.max_row
                for col in range(1, 6):
                    c = ws.cell(row=r, column=col)
                    c.font = Font(name="Arial", size=9)
                    c.border = brd
                    c.alignment = aln("right" if col == 5 else "center" if col == 1 else "left")
                ws.cell(row=r, column=5).number_format = '#,##0'

            ws.append(["", "", "", "Sous-total " + pname, cdata['total']])
            r = ws.max_row
            for col in range(1, 6):
                c = ws.cell(row=r, column=col)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = fill(LBLUE)
                c.border = brd
                c.alignment = aln("right" if col >= 4 else "left")
            ws.cell(row=r, column=5).number_format = '#,##0'
            grand_total += cdata['total']

        ws.append(["", "", "", "TOTAL GÉNÉRAL", grand_total])
        r = ws.max_row
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = brd
            c.alignment = aln("right" if col >= 4 else "left")
        ws.cell(row=r, column=5).number_format = '#,##0'

        for col, width in enumerate([13, 18, 30, 24, 16], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        filename = f"Livraisons_{self.date_debut.strftime('%d%m%Y')}_{self.date_fin.strftime('%d%m%Y')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
