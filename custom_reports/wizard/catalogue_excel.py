# -*- coding: utf-8 -*-
import io
import base64
import datetime
from collections import defaultdict

from odoo import models, fields, api
from odoo.exceptions import ValidationError

import logging
_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    openpyxl = None


# ── Palette couleurs (reprises du QWeb) ───────────────────────────────────────
BLUE_DARK   = "0057A8"   # en-tête colonne / catégorie
BLUE_LIGHT  = "E8F4FD"   # fond prix de détail
WHITE       = "FFFFFF"
GRAY_BORDER = "DDE3EC"
ROW_ODD     = "F5F8FD"
ROW_EVEN    = "FFFFFF"
GREEN_POS   = "006622"
RED_NEG     = "CC0000"
TEXT_MAIN   = "1A1A2E"
GRAY_ZERO   = "AAAAAA"

# ── Définition des colonnes ───────────────────────────────────────────────────
COLUMNS = [
    ("Cde FAM.",   7,  "center"),
    ("FAMILLE",   18,  "left"),
    ("CODE",       9,  "center"),
    ("DÉSIGNATION",32, "left"),
    ("INV.",        6,  "center"),
    ("FOURN.",      9,  "center"),
    ("PMP",         9,  "right"),
    ("CONDT",       7,  "center"),
    ("MAXI",        8,  "center"),
    ("STOCK",       9,  "right"),
    ("CARTON",      9,  "right"),
    ("DÉTAIL",      9,  "right"),
    ("MARG C %",    9,  "right"),
    ("MARG D %",    9,  "right"),
    ("T.V.A",       7,  "center"),
    ("AIRSI",       7,  "center"),
    ("R",           4,  "center"),
]
NCOLS = len(COLUMNS)


def _side(color=GRAY_BORDER, style="thin"):
    return Side(border_style=style, color=color)


def _border(color=GRAY_BORDER):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill():
    return PatternFill("solid", fgColor=BLUE_DARK)


def _categ_fill():
    return PatternFill("solid", fgColor="D6E8F7")


class ProductCatalogueXlsx(models.AbstractModel):
    """
    Génère le catalogue articles en Excel (.xlsx) avec le même
    contenu et la même mise en forme que le rapport QWeb PDF.

    Appelé depuis le wizard ProductReportWizard via action_print_excel().
    """
    _name = "report.custom_reports.catalogue_xlsx"
    _description = "Catalogue Articles XLSX"

    # ── Point d'entrée public ─────────────────────────────────────────────────

    def generate(self, products, company):
        """
        Retourne les bytes du fichier .xlsx.
        products : recordset product.template
        company  : res.company
        """
        if openpyxl is None:
            raise ValidationError(
                "La bibliothèque openpyxl n'est pas installée sur le serveur.\n"
                "Exécutez : pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogue"
        ws.sheet_view.showGridLines = False

        # Figer la 1ʳᵉ ligne (en-tête colonnes)
        ws.freeze_panes = "A3"

        self._write_title_row(ws, company)
        self._write_header_row(ws)
        self._write_body(ws, products)
        self._set_column_widths(ws)
        self._set_print_setup(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Ligne titre (ligne 1) ─────────────────────────────────────────────────

    def _write_title_row(self, ws, company):
        today = datetime.date.today().strftime("%d/%m/%Y")

        # Fusion A1:P1 pour le titre
        ws.merge_cells(start_row=1, start_column=1,
                        end_row=1,   end_column=NCOLS - 1)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"CATALOGUE ARTICLES — {company.name}"
        title_cell.font = Font(name="Arial", bold=True, size=13,
                                color=WHITE, charset=1)
        title_cell.fill = _header_fill()
        title_cell.alignment = Alignment(horizontal="center",
                                          vertical="center",
                                          wrap_text=False)

        # Colonne NCOLS = date
        date_cell = ws.cell(row=1, column=NCOLS)
        date_cell.value = f"Edité le : {today}"
        date_cell.font = Font(name="Arial", bold=True, size=11,
                               color=BLUE_DARK)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    # ── Ligne en-tête colonnes (ligne 2) ──────────────────────────────────────

    def _write_header_row(self, ws):
        for col_idx, (label, _width, align) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font = Font(name="Arial", bold=True, size=11,
                              color=WHITE)
            cell.fill = _header_fill()
            cell.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        wrap_text=False)
            cell.border = Border(
                left=_side(color="3399EE"),
                right=_side(color="3399EE"),
                top=_side(color="004488"),
                bottom=_side(color="004488"),
            )
        ws.row_dimensions[2].height = 18

    # ── Corps du tableau ──────────────────────────────────────────────────────

    def _write_body(self, ws, products):
        categ_dict = defaultdict(list)
        for p in products:
            categ_dict[p.categ_id.id if p.categ_id else 0].append(p)

        row = 3
        for categ_id, prods in sorted(categ_dict.items()):
            categ = self.env["product.category"].browse(categ_id)
            categ_name = categ.name if categ else "Sans catégorie"
            categ_code = categ.code if categ else ""

            row = self._write_categ_header(ws, row, categ_name, len(prods))

            for idx, product in enumerate(prods):
                row = self._write_product_row(ws, row, idx, product,
                                               categ_code, categ_name)

            # Espaceur inter-catégorie
            ws.row_dimensions[row].height = 6
            row += 1

        # Ligne pied de page
        self._write_footer(ws, row)

    # ── En-tête catégorie ─────────────────────────────────────────────────────

    def _write_categ_header(self, ws, row, categ_name, count):
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row,   end_column=NCOLS)
        cell = ws.cell(row=row, column=1)
        cell.value = f"  {categ_name.upper()}   ({count} article{'s' if count > 1 else ''})"
        cell.font = Font(name="Arial", bold=True, size=11, color=BLUE_DARK)
        cell.fill = _categ_fill()
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
        thick = _side(color=BLUE_DARK, style="medium")
        cell.border = Border(bottom=thick)
        ws.row_dimensions[row].height = 16
        return row + 1

    # ── Ligne produit ─────────────────────────────────────────────────────────

    def _write_product_row(self, ws, row, idx, product, categ_code, categ_name):
        bg = ROW_ODD if idx % 2 == 0 else ROW_EVEN
        base_fill = PatternFill("solid", fgColor=bg)
        base_font = Font(name="Arial", size=11, color=TEXT_MAIN)
        brd       = _border()

        def _cell(col, value, align="left", bold=False,
                  color=TEXT_MAIN, fill=None, fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="Arial", size=11, color=color,
                           bold=bold)
            c.fill = fill or base_fill
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = brd
            if fmt:
                c.number_format = fmt
            return c

        # ── C FAM
        _cell(1, categ_code, "center")
        # ── FAMILLE
        _cell(2, categ_name, "left")
        # ── CODE
        _cell(3, product.default_code or "", "center")
        # ── DÉSIGNATION
        _cell(4, product.name or "", "left", bold=True)
        # ── INV.
        inv_name = product.code_inventory_id.name if product.code_inventory_id else ""
        _cell(5, inv_name, "center")
        # ── FOURN.
        fourn = ""
        if product.seller_ids:
            s = product.seller_ids[0]
            fourn = s.partner_id.ref or s.partner_id.name or ""
        _cell(6, fourn, "center")
        # ── PMP
        _cell(7, product.standard_price, "right", fmt="#,##0")
        # ── CONDT
        condt = product.uom_ids[0].relative_factor if product.uom_ids else ""
        _cell(8, condt, "center")
        # ── MAXI
        maxi = product.max_qty_orderpoint or ""
        _cell(9, maxi, "center", fmt="#,##0.00")
        # ── STOCK (coloré)
        stock_val = product.qty_available
        stock_color = GREEN_POS if stock_val > 0 else (RED_NEG if stock_val < 0 else TEXT_MAIN)
        _cell(10, stock_val, "right", color=stock_color, fmt="#,##0.00")

        # ── CARTON
        price_carton = product.price_carton or 0
        if price_carton:
            _cell(11, price_carton, "right", fmt="#,##0")
        else:
            _cell(11, 0, "right", color=GRAY_ZERO, fmt="#,##0")

        # ── DÉTAIL (prix mis en valeur)
        detail_fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        price_detail = product.price_unit_ttc if price_carton else 0
        if price_carton:
            _cell(12, price_detail, "right", bold=True,
                  color=BLUE_DARK, fill=detail_fill, fmt="#,##0")
        else:
            _cell(12, 0, "right", color=GRAY_ZERO, fmt="#,##0")

        # ── MARG C %
        if price_carton:
            tax_rate = (product.taxes_id[0].amount / 100) if product.taxes_id else 0
            price_ht  = price_carton / (1 + tax_rate) if (1 + tax_rate) else 0
            marg_c    = ((price_ht - product.standard_price) / price_ht * 100
                          if price_ht else 0)
            _cell(13, round(marg_c, 2), "right", fmt="0.00")
        else:
            _cell(13, 0, "right", color=GRAY_ZERO, fmt="0.00")

        # ── MARG D %
        if product.list_price:
            marg_d = ((product.list_price - product.standard_price)
                       / product.list_price * 100)
            _cell(14, round(marg_d, 2), "right", fmt="0.00")
        else:
            _cell(14, 0, "right", color=GRAY_ZERO, fmt="0.00")

        # ── T.V.A
        tva_name = product.taxes_id[0].name if product.taxes_id else "0"
        _cell(15, tva_name, "center")

        # ── AIRSI
        airsi = product.airsi_taxes_id[0].amount if product.airsi_taxes_id else 0
        _cell(16, airsi, "center")

        # ── R
        r_val = 1 if product.discount_ligne else 0
        _cell(17, r_val, "center")

        ws.row_dimensions[row].height = 15
        return row + 1

    # ── Pied de page ──────────────────────────────────────────────────────────

    def _write_footer(self, ws, row):
        today = datetime.date.today().strftime("%d/%m/%Y")
        ws.merge_cells(start_row=row, start_column=1,
                        end_row=row,   end_column=NCOLS)
        cell = ws.cell(row=row, column=1)
        company_name = self.env.company.name
        cell.value = (f"{company_name} — Usage interne confidentiel"
                      f"    |    Généré le {today}")
        cell.font = Font(name="Arial", size=10, color="777777", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        top = _side(color="CCD6E8")
        cell.border = Border(top=top)
        ws.row_dimensions[row].height = 14

    # ── Largeurs colonnes ─────────────────────────────────────────────────────

    def _set_column_widths(self, ws):
        for col_idx, (_label, width, _align) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Mise en page impression ───────────────────────────────────────────────

    def _set_print_setup(self, ws):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left   = 0.2
        ws.page_margins.right  = 0.2
        ws.page_margins.top    = 0.5
        ws.page_margins.bottom = 0.3
        ws.print_title_rows    = "1:2"