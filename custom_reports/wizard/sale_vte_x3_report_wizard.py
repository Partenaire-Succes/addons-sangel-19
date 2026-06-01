# -*- coding: utf-8 -*-
import io
import base64
from datetime import datetime, time
from odoo import models, fields, api
from odoo.exceptions import UserError


class SaleVteX3ReportWizard(models.TransientModel):
    _name = 'sale.vte.x3.report.wizard'
    _description = 'Rapport de Ventes X3'

    date_from = fields.Date(string='Date début', required=True, default=fields.Date.context_today)
    date_to = fields.Date(string='Date fin', required=True, default=fields.Date.context_today)

    source = fields.Selection([
        ('sale', 'Ventes (Sales)'),
        ('pos',  'Point de Vente (POS)'),
        ('both', 'Les deux'),
    ], string='Source', required=True, default='both')

    cat_gestion_ids = fields.Many2many(
        'product.category.x3',
        'sale_vte_x3_cat_gestion_rel',
        'wizard_id', 'cat_id',
        string='Catégorie de gestion',
    )
    categ_ids = fields.Many2many(
        'res.partner.category',
        'sale_vte_x3_partner_categ_rel',
        'wizard_id', 'categ_id',
        string='Catégorie client',
    )
    customer_ids = fields.Many2many(
        'res.partner',
        'sale_vte_x3_partner_rel',
        'wizard_id', 'partner_id',
        string='Clients',
        domain=[('customer_rank', '>', 0)],
    )
    default_code = fields.Char(string='Code article (CODE_ART)')

    company_id = fields.Many2one(
        'res.company',
        string='Société',
        default=lambda self: self.env.company,
    )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _dt_from(self):
        return datetime.combine(self.date_from, time.min)

    def _dt_to(self):
        return datetime.combine(self.date_to, time.max)

    def _partner_info(self, partner):
        if not partner:
            return {
                'code_catg_client': '',
                'catg_client': 'Client Comptoir',
                'code_client': '',
                'libelle_client': 'Client Comptoir',
            }
        categories = partner.category_id
        return {
            'code_catg_client': getattr(partner, 'code_family', '') or '',
            'catg_client': categories[0].name if categories else '',
            'code_client': getattr(partner, 'customer_id', '') or '',
            'libelle_client': partner.name or '',
        }

    def _product_info(self, product):
        tmpl = product.product_tmpl_id
        return {
            'cat_art': product.cat_gestion_id.name if product.cat_gestion_id else '',
            'categorie_gestion': product.cat_gestion_id.description if product.cat_gestion_id else '',
            'sousfam_art': tmpl.s_family_id.code if tmpl.s_family_id else '',
            'sousfam_art_libelle': tmpl.s_family_id.name if tmpl.s_family_id else '',
            'code_art': product.default_code or '',
            'designation_art': product.name or '',
            'prix_catalogue': tmpl.list_price,
            'rayon': tmpl.radius_id.code if tmpl.radius_id else '',
            'srayon': tmpl.s_radius_id.code if tmpl.s_radius_id else '',
        }

    # ------------------------------------------------------------------
    # Données sale.order.line
    # ------------------------------------------------------------------

    def _get_sale_lines(self):
        domain = [
            ('order_id.state', 'in', ['sale', 'done']),
            ('order_id.company_id', '=', self.company_id.id),
            ('order_id.date_order', '>=', self._dt_from()),
            ('order_id.date_order', '<=', self._dt_to()),
            ('product_id', '!=', False),
        ]
        if self.cat_gestion_ids:
            domain.append(('product_id.cat_gestion_id', 'in', self.cat_gestion_ids.ids))
        if self.categ_ids:
            domain.append(('order_id.partner_id.category_id', 'in', self.categ_ids.ids))
        if self.customer_ids:
            domain.append(('order_id.partner_id', 'in', self.customer_ids.ids))
        if self.default_code:
            domain.append(('product_id.default_code', 'ilike', self.default_code))

        lines = self.env['sale.order.line'].search(domain)

        rows = []
        for line in lines:
            order = line.order_id
            qty = line.product_uom_qty
            sign = 1 if qty >= 0 else -1
            mt_ht = line.price_subtotal
            mt_ttc = line.price_total
            pump = line.purchase_price or 0.0
            discount = line.discount or 0.0
            prix_net = line.price_unit * (1 - discount / 100.0) if line.price_unit else 0.0

            prod_info = self._product_info(line.product_id)
            part_info = self._partner_info(order.partner_id)

            rows.append({
                **prod_info,
                **part_info,
                'date': order.date_order.date() if order.date_order else False,
                'annee': order.date_order.year if order.date_order else '',
                'mois': order.date_order.month if order.date_order else '',
                'prevision_vet_mois': 0.0,
                'prix_net': round(prix_net, 5),
                'bac': '',
                'piece': order.name,
                'type_facture': 'AVOIR' if sign < 0 else 'FACTURE',
                'pump': pump,
                'qte': abs(qty),
                'qte_signe': qty,
                'mt_ht': mt_ht,
                'mt_ttc': mt_ttc,
                'mtht_signe': mt_ht * sign,
                'marge': line.margin,
                'site': order.company_id.name or '',
            })
        return rows

    # ------------------------------------------------------------------
    # Données pos.order.line
    # ------------------------------------------------------------------

    def _get_pos_lines(self):
        domain = [
            ('order_id.state', 'in', ['paid', 'done', 'invoiced']),
            ('order_id.company_id', '=', self.company_id.id),
            ('order_id.date_order', '>=', self._dt_from()),
            ('order_id.date_order', '<=', self._dt_to()),
            ('product_id', '!=', False),
        ]
        if self.cat_gestion_ids:
            domain.append(('product_id.cat_gestion_id', 'in', self.cat_gestion_ids.ids))
        if self.categ_ids:
            # Les ventes anonymes (sans partner) ne matchent aucune catégorie → exclues si filtre actif
            domain.append(('order_id.partner_id.category_id', 'in', self.categ_ids.ids))
        if self.customer_ids:
            domain.append(('order_id.partner_id', 'in', self.customer_ids.ids))
        if self.default_code:
            domain.append(('product_id.default_code', 'ilike', self.default_code))

        lines = self.env['pos.order.line'].search(domain)

        rows = []
        for line in lines:
            order = line.order_id
            qty = line.qty  # signé : négatif pour les avoirs POS
            sign = 1 if qty >= 0 else -1
            mt_ht = abs(line.price_subtotal)
            mt_ttc = abs(line.price_subtotal_incl)
            # PUMP = coût unitaire estimé depuis total_cost
            abs_qty = abs(qty)
            if abs_qty and line.total_cost:
                pump = abs(line.total_cost) / abs_qty
            else:
                pump = line.product_id.with_company(self.company_id).standard_price or 0.0

            prod_info = self._product_info(line.product_id)
            part_info = self._partner_info(order.partner_id)

            rows.append({
                **prod_info,
                **part_info,
                'date': order.date_order.date() if order.date_order else False,
                'annee': order.date_order.year if order.date_order else '',
                'mois': order.date_order.month if order.date_order else '',
                'prevision_vet_mois': 0.0,
                'prix_net': line.price_unit or 0.0,
                'bac': '',
                'piece': order.name,
                'type_facture': 'AVOIR' if sign < 0 else 'FACTURE',
                'pump': pump,
                'qte': abs_qty,
                'qte_signe': qty,
                'mt_ht': mt_ht,
                'mt_ttc': mt_ttc,
                'mtht_signe': mt_ht * sign,
                'marge': line.margin,
                'site': order.company_id.name or '',
            })
        return rows

    # ------------------------------------------------------------------
    # Requête principale
    # ------------------------------------------------------------------

    def _get_report_data(self):
        self.ensure_one()
        rows = []
        if self.source in ('sale', 'both'):
            rows += self._get_sale_lines()
        if self.source in ('pos', 'both'):
            rows += self._get_pos_lines()
        rows.sort(key=lambda r: (r['date'] or '', r['piece'], r['code_art']))
        return rows

    def _get_grouped_lines(self):
        """Lignes détaillées groupées par CATÉGORIE DE GESTION (pour usage interne)."""
        groups = {}
        for row in self._get_report_data():
            gkey = row['categorie_gestion'] or row['cat_art'] or 'Non défini'
            if gkey not in groups:
                groups[gkey] = {
                    'name': gkey,
                    'cat_art': row['cat_art'],
                    'lines': [],
                    'total_qte': 0.0,
                    'total_mt_ht': 0.0,
                    'total_mt_ttc': 0.0,
                    'total_marge': 0.0,
                }
            sign = 1 if row['type_facture'] == 'FACTURE' else -1
            groups[gkey]['lines'].append(row)
            groups[gkey]['total_qte'] += row['qte_signe']
            groups[gkey]['total_mt_ht'] += row['mtht_signe']
            groups[gkey]['total_mt_ttc'] += row['mt_ttc'] * sign
            groups[gkey]['total_marge'] += row['marge']
        return list(groups.values())

    def _get_summary_data(self):
        """Résumé agrégé par CATÉGORIE DE GESTION → CLIENT pour le PDF."""
        groups = {}
        for row in self._get_report_data():
            gkey = row['categorie_gestion'] or row['cat_art'] or 'Non défini'
            if gkey not in groups:
                groups[gkey] = {
                    'name': gkey,
                    'cat_art': row['cat_art'],
                    'clients': {},
                    'total_qte': 0.0,
                    'total_mt_ht': 0.0,
                    'total_mt_ttc': 0.0,
                    'total_marge': 0.0,
                }
            ckey = row['code_client'] or row['libelle_client'] or 'Client Comptoir'
            if ckey not in groups[gkey]['clients']:
                groups[gkey]['clients'][ckey] = {
                    'code_client': row['code_client'],
                    'libelle_client': row['libelle_client'],
                    'catg_client': row['catg_client'],
                    'code_catg_client': row['code_catg_client'],
                    'qte': 0.0,
                    'mt_ht': 0.0,
                    'mt_ttc': 0.0,
                    'marge': 0.0,
                }
            sign = 1 if row['type_facture'] == 'FACTURE' else -1
            groups[gkey]['clients'][ckey]['qte']    += row['qte_signe']
            groups[gkey]['clients'][ckey]['mt_ht']  += row['mtht_signe']
            groups[gkey]['clients'][ckey]['mt_ttc'] += row['mt_ttc'] * sign
            groups[gkey]['clients'][ckey]['marge']  += row['marge']
            groups[gkey]['total_qte']    += row['qte_signe']
            groups[gkey]['total_mt_ht']  += row['mtht_signe']
            groups[gkey]['total_mt_ttc'] += row['mt_ttc'] * sign
            groups[gkey]['total_marge']  += row['marge']

        result = []
        for g in groups.values():
            result.append({
                'name': g['name'],
                'cat_art': g['cat_art'],
                'clients': sorted(g['clients'].values(), key=lambda c: c['libelle_client']),
                'total_qte': g['total_qte'],
                'total_mt_ht': g['total_mt_ht'],
                'total_mt_ttc': g['total_mt_ttc'],
                'total_marge': g['total_marge'],
            })
        return result

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def action_print_report(self):
        self.ensure_one()
        return self.env.ref('custom_reports.action_report_sale_vte_x3').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise.")

        try:
            return self._build_excel()
        except UserError:
            raise
        except Exception as e:
            raise UserError(f"Erreur lors de la génération Excel : {e}") from e

    def _build_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = self._get_report_data()
        if not rows:
            raise UserError("Aucune donnée trouvée pour les critères sélectionnés.")

        wb = Workbook()
        ws = wb.active
        ws.title = "VTE X3"

        BLUE  = "1A5276"
        WHITE = "FFFFFF"
        GREY  = "F2F2F2"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def font(bold=False, color=None, size=9):
            kw = dict(name="Arial", bold=bold, size=size)
            if color:
                kw['color'] = color
            return Font(**kw)

        def aln(h="left"):
            return Alignment(horizontal=h)

        headers = [
            'DATE', 'ANNEE', 'MOIS', 'CAT_ART', 'CAREGORIE DE GESTION',
            'CODE_CATG_CLIENT', 'CATG_CLIENT', 'CODE_CLIENT', 'LIBELLE_CLIENT',
            'SOUSFAM_ART', 'SOUSFAM_ART_LIBELLE', 'CODE_ART', 'DESIGNATION_ART',
            'PRIX_CATALOGUE', 'PREVISION_VET_MOIS', 'PRIX_NET', 'BAC', 'PIECE',
            'TYPE_FACTURE', 'PUMP', 'QTE', 'QTE_SIGNE', 'MT_HT', 'MT_TTC',
            'MTHT_SIGNE', 'MARGE', 'SITE', 'RAYON', 'SRAYON',
        ]
        num_cols   = len(headers)
        last_col   = get_column_letter(num_cols)

        # ── Ligne 1 : titre société ────────────────────────────────
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = (
            f"{self.company_id.name}"
            f"  —  RAPPORT DE VENTES"
            f"  du {self.date_from.strftime('%d/%m/%Y')}"
            f"  au {self.date_to.strftime('%d/%m/%Y')}"
        )
        ws["A1"].font      = font(bold=True, color=WHITE, size=11)
        ws["A1"].fill      = fill(BLUE)
        ws["A1"].alignment = aln("center")
        ws.row_dimensions[1].height = 20

        # ── Ligne 2 : en-têtes colonnes ───────────────────────────
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, num_cols + 1):
            c = ws.cell(row=hrow, column=col)
            c.font      = font(bold=True, color=WHITE, size=9)
            c.fill      = fill(BLUE)
            c.alignment = aln("center")
            c.border    = brd
        ws.row_dimensions[hrow].height = 18

        # ── Colonnes numériques ────────────────────────────────────
        num_cols_right = {14, 15, 16, 20, 21, 22, 23, 24, 25, 26}
        money_cols     = {14, 15, 16, 20, 23, 24, 25, 26}

        # ── Données ───────────────────────────────────────────────
        data_start = hrow + 1
        for row_offset, row in enumerate(rows):
            r = data_start + row_offset
            ws.append([
                row['date'],
                row['annee'],
                row['mois'],
                row['cat_art'],
                row['categorie_gestion'],
                row['code_catg_client'],
                row['catg_client'],
                row['code_client'],
                row['libelle_client'],
                row['sousfam_art'],
                row['sousfam_art_libelle'],
                row['code_art'],
                row['designation_art'],
                row['prix_catalogue'],
                row['prevision_vet_mois'],
                row['prix_net'],
                row['bac'],
                row['piece'],
                row['type_facture'],
                row['pump'],
                row['qte'],
                row['qte_signe'],
                row['mt_ht'],
                row['mt_ttc'],
                row['mtht_signe'],
                row['marge'],
                row['site'],
                row['rayon'],
                row['srayon'],
            ])
            for col in range(1, num_cols + 1):
                c = ws.cell(row=r, column=col)
                c.font      = font(size=9)
                c.border    = brd
                c.alignment = aln("right" if col in num_cols_right else "left")
            for col in money_cols:
                ws.cell(row=r, column=col).number_format = '#,##0'
            ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'
            ws.row_dimensions[r].height = 14

        # ── Largeurs ──────────────────────────────────────────────
        col_widths = [12, 7, 6, 9, 20, 15, 20, 15, 25, 12, 22, 10, 28,
                      14, 16, 12, 6, 15, 10, 12, 10, 10, 14, 14, 14, 14, 12, 8, 8]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A3"

        # ── Export ────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()
        filename = (
            f"VTE_X3_{self.date_from.strftime('%d%m%Y')}"
            f"_{self.date_to.strftime('%d%m%Y')}.xlsx"
        )
        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'type':      'binary',
            'datas':     xlsx_data,
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }


class ReportSaleVteX3(models.AbstractModel):
    _name = 'report.custom_reports.report_sale_vte_x3_document'
    _description = 'Rapport de Ventes X3'

    @api.model
    def _get_report_values(self, docids, data=None):
        wizard = self.env['sale.vte.x3.report.wizard'].browse(docids)
        groups = wizard._get_summary_data()
        return {
            'doc': wizard,
            'company': wizard.company_id,
            'groups': groups,
            'date_from': wizard.date_from,
            'date_to': wizard.date_to,
        }
