import io
import base64

from odoo import models, fields, api, _
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from odoo.exceptions import ValidationError, UserError


class ReliquatReport(models.Model):
    _name = 'reliquat.report'
    _inherit = ['mail.thread','mail.activity.mixin']
    _description = 'Rapport de non livrés'
    _order = 'date_from desc'
    _rec_name = 'name'

    name = fields.Char(
        string='Nom du rapport',
        required=True,
        default=lambda self: self._get_default_name()
    )

    active = fields.Boolean(
        string='Actif',
        default=True
    )

    date_from = fields.Date(
        string='Date de début',
        required=True
    )

    date_to = fields.Date(
        string='Date de fin',
        required=True
    )

    period_type = fields.Selection([
        ('daily', 'Quotidien'),
        ('weekly', 'Hebdomadaire'),
        ('biweekly', 'Quinzaine'),
        ('monthly', 'Mensuel'),
        ('quarterly', 'Trimestriel'),
        ('semiannual', 'Semestriel'),
        ('yearly', 'Annuel'),
        ('custom', 'Personnalisé')
    ], string='Type de période', required=True, default='monthly')

    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('confirmed', 'Confirmé'),
        ('printed', 'Imprimé')
    ], string='État', default='draft')

    # Lignes de non livrés
    line_ids = fields.One2many(
        comodel_name='reliquat.report.line',
        inverse_name='report_id',
        string='Lignes de non livrés'
    )

    total_orders = fields.Integer(
        string='Nombre decommandes',
        compute='_compute_statistics'
    )

    total_qty_ordered = fields.Float(
        string='Quantité totale commandée',
        compute='_compute_statistics'
    )

    total_qty_received = fields.Float(
        string='Quantité totale reçue',
         compute='_compute_statistics'
    )

    total_qty_pending = fields.Float(
        string='Quantité en attente',
        compute='_compute_statistics'
    )

    satisfaction_rate = fields.Float(
        string='Taux de satisfaction',
        compute='_compute_statistics'
    )

    created_by = fields.Many2one(
        comodel_name='res.users',
        string='Créé par',
        default=lambda self: self.env.user
    )

    creation_date = fields.Datetime(
        string='Date de création',
        default=fields.Datetime.now
    )

    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Société',
        default=lambda self: self.env.company
    )

    partner_ids = fields.Many2many(
        comodel_name='res.partner',
        string='Fournisseurs filtrés',
        help="Filtre appliqué à la génération du rapport : si vide, tous les "
             "fournisseurs sont inclus. Pris en compte aussi à l'impression "
             "puisque les lignes (line_ids) sont générées selon ce filtre."
    )

    @api.model
    def _get_default_name(self):
        """Construit le nom par défaut avec les dates en jj/mm/aaaa"""

        if not self.date_from:
            raise ValueError(_("Veuillez définir la date de début."))

        date_from_str = self.date_from.strftime('%d/%m/%Y')
        date_to_str = self.date_to.strftime('%d/%m/%Y') if self.date_to else None

        if self.date_to and self.date_from != self.date_to:
            report_name = f"Rapport de non livrés du {date_from_str} au {date_to_str}"
        else:
            report_name = f"Rapport de non livrés du {date_from_str}"

    @api.onchange('date_from', 'date_to')
    def _onchange_dates(self):
        """Met à jour automatiquement le nom si les deux dates sont définies"""
        if self.date_from and self.date_to:
            date_from_str = self.date_from.strftime('%d/%m/%Y')
            date_to_str = self.date_to.strftime('%d/%m/%Y')
            self.name = f"Rapport de non livrés du {date_from_str} au {date_to_str}"

    @api.depends('line_ids')
    def _compute_statistics(self):
        for report in self:
            if report.line_ids:
                report.total_orders = len(report.line_ids)
                report.total_qty_ordered = sum(line.qty_ordered for line in report.line_ids)
                report.total_qty_received = sum(line.qty_received for line in report.line_ids)
                report.total_qty_pending = report.total_qty_ordered - report.total_qty_received

                if report.total_qty_ordered > 0:
                    report.satisfaction_rate = report.total_qty_received / report.total_qty_ordered
                else:
                    report.satisfaction_rate = 0.0
            else:
                report.total_orders = 0
                report.total_qty_ordered = 0.0
                report.total_qty_received = 0.0
                report.total_qty_pending = 0.0
                report.satisfaction_rate = 0.0

    def generate_report_data(self):
        """Génère les données du rapport basé sur les commandes d'achat"""
        self.ensure_one()

        # Supprimer les anciennes lignes
        self.line_ids.unlink()

        # Rechercher les commandes d'achat dans la période (+ filtre fournisseurs)
        domain = [
            ('date_order', '>=', self.date_from),
            ('date_order', '<=', self.date_to),
            ('state', 'in', ['purchase', 'done'])
        ]
        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))

        purchase_orders = self.env['purchase.order'].search(domain)

        lines_data = []
        for order in purchase_orders:
            for line in order.order_line:
                qty_received = sum(move.quantity for move in line.move_ids
                                   if move.state == 'done')

                qty_pending = line.product_qty - qty_received

                if qty_pending > 0:  # Seulement les lignes avec reliquats
                    satisfaction_rate = qty_received / line.product_qty if line.product_qty > 0 else 0

                    lines_data.append({
                        'report_id': self.id,
                        'purchase_order_id': order.id,
                        'partner_id': order.partner_id.id,
                        'product_id': line.product_id.id,
                        'qty_ordered': line.product_qty,
                        'qty_received': qty_received,
                        'qty_pending': qty_pending,
                        'satisfaction_rate': satisfaction_rate,
                        'order_date': order.date_order,
                    })

        # Créer les lignes
        self.env['reliquat.report.line'].create(lines_data)

        return True


    def action_print(self):
        self.state = 'printed'
        return self.env.ref('custom_reliquat_report.action_report_reliquat').report_action(self)

    def action_export_excel(self):
        """Exporte le rapport de non livrés en Excel — même modèle que le
        rapport « Mouvements par produit » (custom_stock) : openpyxl,
        ir.attachment, téléchargement via ir.actions.act_url."""
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl est requise."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Non livrés"

        BLUE  = "1A5276"
        WHITE = "FFFFFF"
        ROW_A = "D6EAF8"
        ROW_B = "EBF5FB"
        thin  = Side(style='thin', color="AAAAAA")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def font(bold=False, color="000000", size=9):
            return Font(name="Arial", bold=bold, color=color, size=size)

        def fill(c):
            return PatternFill("solid", fgColor=c)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        NCOLS = 8

        ws.merge_cells("A1:H1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = font(bold=True, size=11)

        ws.merge_cells("A2:H2")
        ws["A2"] = "%s — Du %s au %s" % (
            self.name,
            self.date_from.strftime('%d/%m/%Y'),
            self.date_to.strftime('%d/%m/%Y'),
        )
        ws["A2"].font = font(bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18

        if self.partner_ids:
            ws.merge_cells("A3:H3")
            ws["A3"] = "Fournisseur(s) filtré(s) : %s" % ', '.join(self.partner_ids.mapped('name'))
            ws["A3"].font = font(size=8)
        ws.append([])

        # Récap statistiques globales
        ws.append([
            "Total commandes", self.total_orders,
            "Qté commandée", "%.2f" % self.total_qty_ordered,
            "Qté reçue", "%.2f" % self.total_qty_received,
            "Qté en attente", "%.2f" % self.total_qty_pending,
        ])
        r = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            odd = col % 2 == 1
            c.font = font(bold=True, color=WHITE if odd else BLUE, size=9)
            c.fill = fill(BLUE if odd else ROW_A)
            c.border = brd
            c.alignment = aln("left" if odd else "right")
        ws.row_dimensions[r].height = 16

        ws.append(["Taux de satisfaction global : %.2f %%" % (self.satisfaction_rate * 100)])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1)
        c.font = font(bold=True, size=9)
        c.alignment = aln("center")
        ws.append([])

        hdrs = ["Date", "Commande", "Fournisseur", "Produit",
                "Qté Cdée", "Qté Reçue", "Reliquat", "Taux Sat. (%)"]
        ws.append(hdrs)
        r = ws.max_row
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.font = font(bold=True, size=9)
            c.fill = fill(ROW_A)
            c.border = brd
            c.alignment = aln("center")

        for i, line in enumerate(self.line_ids):
            ws.append([
                line.order_date.strftime('%d/%m/%Y') if line.order_date else '',
                line.purchase_order_id.name or '',
                line.partner_name or '',
                line.product_name or '',
                line.qty_ordered,
                line.qty_received,
                line.qty_pending,
                round(line.satisfaction_rate * 100, 2),
            ])
            r = ws.max_row
            bg = ROW_A if i % 2 == 0 else ROW_B
            for col in range(1, NCOLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = font(size=8)
                c.fill = fill(bg)
                c.border = brd
                c.alignment = aln("right" if col in (5, 6, 7, 8) else "left")
            for col in (5, 6, 7):
                ws.cell(row=r, column=col).number_format = '#,##0.##'
            ws.cell(row=r, column=8).number_format = '#,##0.00'

        for col, w in enumerate([12, 18, 22, 28, 12, 12, 12, 14], 1):
            ws.column_dimensions[chr(64 + col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read()).decode()

        fname = "Rapport_non_livres_%s.xlsx" % fields.Date.context_today(self).strftime('%d%m%Y')
        att = self.env['ir.attachment'].create({
            'name': fname,
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'new',
        }



