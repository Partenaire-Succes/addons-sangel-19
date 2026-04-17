from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook
import logging

_logger = logging.getLogger(__name__)


class StockExcelImportWizard(models.TransientModel):
    _name = "stock.excel.import.wizard"
    _description = "Stock Excel Import Wizard"

    file = fields.Binary(string="Excel File")
    file_name = fields.Char()

    company_id = fields.Many2one(
        "res.company",
        required=True,
        default=lambda self: self.env.company
    )

    location_id = fields.Many2one(
        "stock.location",
        required=True,
        domain="[('usage','=','internal')]"
    )

    warehouse_id = fields.Many2one(
        "stock.warehouse",
        string="Entrepôt",
        required=True,
        default=lambda self: self.env['stock.warehouse'].search(
            [('company_id', '=', self.env.company.id)], limit=1
        )
    )

    line_ids = fields.One2many(
        "stock.excel.import.line",
        "wizard_id",
        string="Lines"
    )

    state = fields.Selection([
        ("draft", "Draft"),
        ("loaded", "Loaded"),
        ("done", "Done"),
    ], default="draft")

    import_mode = fields.Selection([
        ("product", "Produits"),
        ("partner", "Contacts"),
    ], default="product")

    # -------------------------
    # LOAD FILE
    # -------------------------
    @staticmethod
    def _clean_code(value):
        """Convertit une valeur Excel en code article propre (sans .0 parasite)."""
        if value is None:
            return False
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()


    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload a file."))

        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        file_data = io.BytesIO(decoded_file)
        workbook = load_workbook(file_data)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("The Excel file is empty."))

        headers = [str(h).strip() for h in rows[0]]

        required_columns = ["product_code", "product_state", "quantity"]
        for col in required_columns:
            if col not in headers:
                raise UserError(_("Missing column: %s") % col)

        product_index = headers.index("product_code")
        status_index = headers.index("product_state")
        qty_index = headers.index("quantity")

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id]
        ))

        lines_vals = []

        for row in rows[1:]:
            if not row:
                continue

            product_code = self._clean_code(row[product_index])
            product_state = self._clean_code(row[status_index])
            quantity = float(row[qty_index] or 0.0)

            if not product_code:
                continue

            product = env["product.product"].search(
                [("default_code", "=", product_code)], limit=1
            )

            lines_vals.append((0, 0, {
                "product_code": product_code,
                "product_id": product.id if product else False,
                "p_state": product_state,
                "quantity": quantity,
                "found": bool(product),
            }))

        self.write({
            "line_ids": lines_vals,
            "state": "loaded",
            "import_mode": "product",   # ✅ on marque le mode
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }


    def action_load_file_partner(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload a file."))

        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        file_data = io.BytesIO(decoded_file)
        workbook = load_workbook(file_data)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("The Excel file is empty."))

        headers = [str(h).strip() for h in rows[0]]

        required_columns = ["customer_id", "customer_account"]
        for col in required_columns:
            if col not in headers:
                raise UserError(_("Missing column: %s") % col)

        customer_id_index = headers.index("customer_id")
        customer_account_index = headers.index("customer_account")

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id]
        ))

        lines_vals = []

        for row in rows[1:]:
            if not row:
                continue

            customer_id_val = self._clean_code(row[customer_id_index])
            customer_account_val = self._clean_code(row[customer_account_index])

            if not customer_id_val:
                continue

            partner = env["res.partner"].search(
                [("customer_id", "=", customer_id_val)], limit=1
            )

            lines_vals.append((0, 0, {
                "customer_id": customer_id_val,
                "customer_account": customer_account_val,
                "partner_id": partner.id if partner else False,
                "partner_found": bool(partner),
            }))

        self.write({
            "line_ids": lines_vals,
            "state": "loaded",
            "import_mode": "partner",   # ✅ on marque le mode
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }


    # -------------------------
    # CONFIRM — dispatche selon le mode
    # -------------------------
    def action_confirm(self):
        self.ensure_one()
        if self.import_mode == "product":
            return self._confirm_products()
        elif self.import_mode == "partner":
            return self._confirm_partners()


    def _confirm_products(self):
        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id]
        ))
        count_lines = len(self.line_ids.filtered(lambda l: l.found))

        for line in self.line_ids.filtered(lambda l: l.found):
            product = line.product_id.with_company(self.company_id)

            status = self.env["product.status"].search([
                ("code", "=", line.p_state),
                ("active", "=", True)
            ], limit=1)

            if status:
                tmpl = product.product_tmpl_id.with_context(
                    allowed_company_ids=[self.company_id.id]
                ).with_company(self.company_id)
                tmpl.current_company_status_id = status.id

            orderpoint = env["stock.warehouse.orderpoint"].search([
                ("product_id", "=", product.id),
                ("warehouse_id", "=", self.warehouse_id.id),
            ], limit=1)

            values = {
                "product_min_qty": line.quantity,
                "product_max_qty": line.quantity,
            }

            if orderpoint:
                orderpoint.write(values)
            else:
                values.update({
                    "product_id": product.id,
                    "location_id": self.location_id.id,
                    "company_id": self.company_id.id,
                    "warehouse_id": self.warehouse_id.id,
                })
                env["stock.warehouse.orderpoint"].create(values)

        self.state = "done"
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Produits mis à jour — %s ligne(s)' % count_lines,
                'message': "Import produits terminé avec succès.",
                'type': 'success',
            }
        }


    def _confirm_partners(self):
        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id]
        ))
        count_lines = len(self.line_ids.filtered(lambda l: l.partner_found))

        for line in self.line_ids.filtered(lambda l: l.partner_found):
            if line.customer_account:
                line.partner_id.with_company(self.company_id).customer_account = line.customer_account

        self.state = "done"
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Contacts mis à jour — %s ligne(s)' % count_lines,
                'message': "Import contacts terminé avec succès.",
                'type': 'success',
            }
        }

    # -------------------------
    # CONFIRM INVENTORY UPDATE
    # -------------------------
    # def action_confirm(self):
    #     self.ensure_one()

    #     env = self.env(context=dict(
    #         self.env.context,
    #         allowed_company_ids=[self.company_id.id]
    #     ))
    #     quants_to_apply = env["stock.quant"]
    #     count_lines = len(self.line_ids.filtered(lambda l: l.found))

    #     for line in self.line_ids.filtered(lambda l: l.found):

    #         product = line.product_id.with_company(self.company_id)

    #         status = self.env["product.status"].search([
    #             ("code", "=", line.p_state),
    #             ("active", "=", True)
    #         ], limit=1)

    #         if status:
    #             # ✅ Écrire sur le template avec le bon contexte société
    #             tmpl = product.product_tmpl_id.with_context(
    #                 allowed_company_ids=[self.company_id.id]
    #             ).with_company(self.company_id)
    #             tmpl.current_company_status_id = status.id

    #         if product:
    #             product.standard_price = line.cost

    #         quant = env["stock.quant"].search([
    #             ("product_id", "=", product.id),
    #             ("location_id", "=", self.location_id.id),
    #         ], limit=1)

    #         if quant:
    #             quant.inventory_quantity = line.quantity
    #         else:
    #             quant = env["stock.quant"].create({
    #                 "product_id": product.id,
    #                 "location_id": self.location_id.id,
    #                 "company_id": self.company_id.id,
    #                 "inventory_quantity": line.quantity,
    #             })

    #         quants_to_apply |= quant

    #     if quants_to_apply:
    #         quants_to_apply._apply_inventory()

    #     self.state = "done"

    #     # return {"type": "ir.actions.act_window_close"}
    #     return {
    #             'type': 'ir.actions.client',
    #             'tag': 'display_notification',
    #             'params': {
    #                 'title': 'Mise à jour du stock ligne total : %s' % count_lines,
    #                 'message': "Import terminé avec succès.",
    #                 'type': 'success',
    #             }
    #         }


    # def action_confirm(self):
    #     self.ensure_one()

    #     env = self.env(context=dict(
    #         self.env.context,
    #         allowed_company_ids=[self.company_id.id]
    #     ))

    #     count_lines = len(self.line_ids.filtered(lambda l: l.found))

    #     for line in self.line_ids.filtered(lambda l: l.found):

    #         product = line.product_id.with_company(self.company_id)

    #         # Mise à jour du statut produit
    #         status = self.env["product.status"].search([
    #             ("code", "=", line.p_state),
    #             ("active", "=", True)
    #         ], limit=1)

    #         if status:
    #             tmpl = product.product_tmpl_id.with_context(
    #                 allowed_company_ids=[self.company_id.id]
    #             ).with_company(self.company_id)
    #             tmpl.current_company_status_id = status.id

    #         # ✅ Recherche par warehouse_id.id (Many2one → int)
    #         orderpoint = env["stock.warehouse.orderpoint"].search([
    #             ("product_id", "=", product.id),
    #             ("warehouse_id", "=", self.warehouse_id.id),  # ← .id obligatoire
    #         ], limit=1)

    #         if orderpoint:
    #             orderpoint.product_max_qty = line.quantity
    #             orderpoint.product_min_qty = line.quantity
    #         else:
    #             env["stock.warehouse.orderpoint"].create({
    #                 "product_id": product.id,
    #                 "product_max_qty": line.quantity,
    #                 "product_min_qty": line.quantity,
    #                 "location_id": self.location_id.id,   # ← .id aussi
    #                 "company_id": self.company_id.id,
    #                 "warehouse_id": self.warehouse_id.id,  # ← .id aussi
    #             })

    #     self.state = "done"

    #     return {
    #         'type': 'ir.actions.client',
    #         'tag': 'display_notification',
    #         'params': {
    #             'title': 'Mise à jour du stock — %s ligne(s)' % count_lines,
    #             'message': "Import terminé avec succès.",
    #             'type': 'success',
    #         }
    #     }

class StockExcelImportLine(models.TransientModel):
    _name = "stock.excel.import.line"
    _description = "Stock Excel Import Line"

    wizard_id = fields.Many2one("stock.excel.import.wizard")

    product_id = fields.Many2one("product.product", readonly=True)
    product_name = fields.Char(related="product_id.name", string="Article", readonly=True)
    product_code = fields.Char("Code article")
    quantity = fields.Float("Quantité")
    cost = fields.Float("Coût")
    p_state = fields.Char("Statut Article", readonly=True)

    partner_id = fields.Many2one("res.partner", string="Contact", readonly=True)
    customer_id = fields.Char("ID Client")
    customer_account = fields.Char("Compte Client")
    partner_found = fields.Boolean("Contact trouvé")

    found = fields.Boolean("Produit trouvé")