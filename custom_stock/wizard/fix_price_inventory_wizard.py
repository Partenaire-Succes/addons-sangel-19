# -*- coding: utf-8 -*-
from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook
import logging

_logger = logging.getLogger(__name__)


class FixPriceInventoryWizard(models.TransientModel):
    _name = "fix.price.inventory.wizard"
    _description = "Correction Prix PMP — Inventaire"

    file = fields.Binary(string="Fichier Excel")
    file_name = fields.Char()

    company_id = fields.Many2one(
        "res.company",
        string="Société",
        required=True,
        default=lambda self: self.env.company,
    )

    line_ids = fields.One2many(
        "fix.price.inventory.line",
        "wizard_id",
        string="Lignes",
    )

    state = fields.Selection([
        ("draft",  "Brouillon"),
        ("loaded", "Chargé"),
        ("done",   "Terminé"),
    ], default="draft")

    # ── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _clean_code(value):
        """Nettoie le code article lu depuis Excel (supprime les .0 parasites)."""
        if value is None:
            return False
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @staticmethod
    def _clean_cost(value):
        """Parse un coût depuis Excel (gère la virgule décimale française)."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(',', '.').strip())
        except (ValueError, TypeError):
            return 0.0

    # ── Template de téléchargement ────────────────────────────────────────────

    def action_download_template(self):
        """Génère et propose au téléchargement un fichier Excel modèle."""
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl est requise."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        BLUE  = "1A5276"
        WHITE = "FFFFFF"

        headers = ["product_code", "cost"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
            cell.fill      = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 20

        # Exemples de lignes pour guider l'utilisateur
        examples = [
            ("4219", 801.31),
            ("1234", 1500.00),
        ]
        for row_data in examples:
            ws.append(row_data)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read()).decode()

        att = self.env["ir.attachment"].create({
            "name":      "Template_Correction_Prix_PMP.xlsx",
            "type":      "binary",
            "datas":     xlsx_data,
            "mimetype":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id":    self.id,
        })
        return {
            "type":   "ir.actions.act_url",
            "url":    f"/web/content/{att.id}?download=true",
            "target": "new",
        }

    # ── Chargement du fichier ─────────────────────────────────────────────────

    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez télécharger un fichier Excel."))

        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        workbook = load_workbook(io.BytesIO(decoded_file))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        headers = [str(h).strip() for h in rows[0]]

        if "product_code" not in headers:
            raise UserError(_("Colonne manquante : product_code"))

        # Accepte "cost", "Coût", "cout"…
        cost_col = next(
            (c for c in ("cost", "Coût", "cout", "COUT", "COÛT") if c in headers),
            None
        )
        if cost_col is None:
            raise UserError(_("Colonne manquante : 'cost' ou 'Coût'"))

        code_idx = headers.index("product_code")
        cost_idx = headers.index(cost_col)

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id],
        ))

        lines_vals = []

        for row in rows[1:]:
            if not row:
                continue

            product_code = self._clean_code(row[code_idx])
            cost_excel   = self._clean_cost(row[cost_idx])

            if not product_code:
                continue

            # ── Résolution du produit ──────────────────────────────────────
            product = env["product.product"].search(
                [("default_code", "=", product_code)], limit=1
            )
            tmpl = product.product_tmpl_id if product else env["product.template"].browse()

            if not tmpl:
                lines_vals.append((0, 0, {
                    "product_code":      product_code,
                    "cost_excel":        cost_excel,
                    "product_tmpl_id":   False,
                    "current_price":     0.0,
                    "current_std_price": 0.0,
                    "inv_lines_count":   0,
                    "action":            "not_found",
                    "found":             False,
                }))
                continue

            current_std_price = tmpl.standard_price

            # ── Nombre de lignes d'inventaire avec price == 0 ─────────────
            inv_lines_count = env["physical.inventory.line"].search_count([
                ("product_tmpl_id", "=", tmpl.id),
                ("price",           "=", 0.0),
            ])

            # ── Décision : corriger UNIQUEMENT si des lignes d'inventaire sont à 0 ──
            # Le standard_price n'est PAS modifié ici — correction séparée via fiche article
            action = "fix" if inv_lines_count > 0 else "skip"

            lines_vals.append((0, 0, {
                "product_code":      product_code,
                "cost_excel":        cost_excel,
                "product_tmpl_id":   tmpl.id,
                "current_price":     current_std_price,
                "current_std_price": current_std_price,
                "inv_lines_count":   inv_lines_count,
                "action":            action,
                "found":             True,
            }))

        if not lines_vals:
            raise UserError(_("Aucune ligne valide trouvée dans le fichier."))

        self.write({
            "line_ids": lines_vals,
            "state":    "loaded",
        })

        return {
            "type":      "ir.actions.act_window",
            "res_model": self._name,
            "res_id":    self.id,
            "view_mode": "form",
            "target":    "new",
        }

    # ── Confirmation ──────────────────────────────────────────────────────────

    def action_confirm(self):
        """
        Corrige UNIQUEMENT le champ `price` sur les lignes d'inventaire passées
        où ce champ est à 0. Le standard_price du produit n'est PAS modifié ici —
        sa correction se fait séparément via Fiche article → Mettre à jour le coût.
        """
        self.ensure_one()
        count_inv_lines = 0

        for line in self.line_ids.filtered(lambda l: l.action == "fix"):
            # Corriger price sur toutes les lignes d'inventaire à 0 pour ce produit
            inv_lines = self.env["physical.inventory.line"].search([
                ("product_tmpl_id", "=", line.product_tmpl_id.id),
                ("price",           "=", 0.0),
            ])
            if inv_lines:
                inv_lines.write({"price": line.cost_excel})
                count_inv_lines += len(inv_lines)

        self.state = "done"
        return {
            "type": "ir.actions.client",
            "tag":  "display_notification",
            "params": {
                "title":   _("Correction terminée — %d ligne(s) d'inventaire mises à jour") % count_inv_lines,
                "message": _(
                    "Le champ Prix unitaire (PMP) des lignes d'inventaire à 0 a été corrigé. "
                    "Le standard_price du produit reste inchangé."
                ),
                "type":   "success",
                "sticky": True,
            },
        }


class FixPriceInventoryLine(models.TransientModel):
    _name = "fix.price.inventory.line"
    _description = "Ligne Correction Prix PMP"

    wizard_id = fields.Many2one("fix.price.inventory.wizard")

    product_code        = fields.Char("Code Article",            readonly=True)
    product_tmpl_id     = fields.Many2one("product.template", "Produit trouvé", readonly=True)
    product_name        = fields.Char(related="product_tmpl_id.name", string="Désignation", readonly=True)
    current_price       = fields.Float("standard_price actuel",  readonly=True)
    current_std_price   = fields.Float("Prix PMP actuel",        readonly=True)
    inv_lines_count     = fields.Integer("Lignes inv. à 0",      readonly=True)
    cost_excel          = fields.Float("Prix à appliquer")

    action = fields.Selection([
        ("fix",       "À corriger"),
        ("skip",      "Ignoré (prix OK)"),
        ("not_found", "Produit non trouvé"),
    ], string="Statut", readonly=True)

    found = fields.Boolean("Trouvé", readonly=True)
