# -*- coding: utf-8 -*-
import io
import base64

from odoo import models, fields, api, _
from odoo.exceptions import UserError


class StockMoveProductReportWizard(models.TransientModel):
    _name = 'stock.move.product.report.wizard'
    _description = 'Assistant Rapport des mouvements par produit'

    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )
    product_ids = fields.Many2many(
        comodel_name='product.product',
        string='Produits',
        required=True,
        help="Sélectionnez un ou plusieurs articles : le rapport sortira "
             "tous les mouvements de stock validés pour chacun d'eux.",
    )
    date_debut = fields.Date(string='Date de début')
    date_fin = fields.Date(string='Date de fin')

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for rec in self:
            if rec.date_debut and rec.date_fin and rec.date_debut > rec.date_fin:
                raise UserError(_("La date de début doit être antérieure à la date de fin."))

    # ── Construction des données (partagée PDF / Excel) ──────────────────────

    def _operation_label(self, move):
        """Type d'opération lisible : type de transfert, sinon repli sur les emplacements."""
        if move.picking_type_id:
            return move.picking_type_id.name
        if move.location_dest_id.usage == 'internal' and move.location_id.usage != 'internal':
            return _('Réception')
        if move.location_dest_id.usage != 'internal' and move.location_id.usage == 'internal':
            return _('Livraison')
        if move.location_dest_id.usage == 'internal' and move.location_id.usage == 'internal':
            return _('Transfert interne')
        return _('Mouvement')

    def _move_direction(self, move):
        """Sens du mouvement du point de vue du stock : Entrée, Sortie ou Interne.
        Sert à séparer clairement ce qui entre de ce qui sort, pour une lecture
        professionnelle du rapport (totaux distincts entrées / sorties)."""
        dest_internal = move.location_dest_id.usage == 'internal'
        src_internal = move.location_id.usage == 'internal'
        if dest_internal and not src_internal:
            return 'in', _('Entrée')
        if not dest_internal and src_internal:
            return 'out', _('Sortie')
        return 'internal', _('Interne')

    def _get_report_data(self):
        """Retourne, pour chaque produit sélectionné : ses mouvements détaillés
        (date, type d'opération, sens entrée/sortie, référence, quantité, prix,
        valeur) ainsi que les sommes globales et les sommes séparées par sens
        (entrées / sorties / internes), pour une vue professionnelle claire."""
        self.ensure_one()

        domain = [
            ('product_id', 'in', self.product_ids.ids),
            ('state', '=', 'done'),
            ('company_id', '=', self.company_id.id),
        ]
        if self.date_debut:
            domain.append(('date', '>=', str(self.date_debut) + ' 00:00:00'))
        if self.date_fin:
            domain.append(('date', '<=', str(self.date_fin) + ' 23:59:59'))

        moves = self.env['stock.move'].search(domain, order='date asc, id asc')
        if not moves:
            raise UserError(_("Aucun mouvement trouvé pour les articles et la période sélectionnés."))

        result = []
        for product in self.product_ids:
            product_moves = moves.filtered(lambda m: m.product_id.id == product.id)
            if not product_moves:
                continue

            lines = []
            qty_total = 0.0
            value_total = 0.0
            sums = {
                'in': {'qty': 0.0, 'value': 0.0, 'count': 0},
                'out': {'qty': 0.0, 'value': 0.0, 'count': 0},
                'internal': {'qty': 0.0, 'value': 0.0, 'count': 0},
            }
            for move in product_moves:
                qty = sum(move.move_line_ids.mapped('quantity')) or move.product_uom_qty
                price = move.price_unit
                value = qty * price
                qty_total += qty
                value_total += value

                direction, direction_label = self._move_direction(move)
                sums[direction]['qty'] += qty
                sums[direction]['value'] += value
                sums[direction]['count'] += 1

                reference = move.reference or (move.picking_id.name if move.picking_id else move.name)
                lines.append({
                    'date': move.date,
                    'operation': self._operation_label(move),
                    'direction': direction,
                    'direction_label': direction_label,
                    'reference': reference,
                    'qty': qty,
                    'uom': move.product_uom.name,
                    'price_unit': price,
                    'value': value,
                })

            result.append({
                'product': product,
                'lines': lines,
                'nb_mouvements': len(lines),
                'qty_total': qty_total,
                'value_total': value_total,
                'in_qty': sums['in']['qty'],
                'in_value': sums['in']['value'],
                'in_count': sums['in']['count'],
                'out_qty': sums['out']['qty'],
                'out_value': sums['out']['value'],
                'out_count': sums['out']['count'],
                'internal_qty': sums['internal']['qty'],
                'internal_value': sums['internal']['value'],
                'internal_count': sums['internal']['count'],
                'solde_qty': sums['in']['qty'] - sums['out']['qty'],
                'solde_value': sums['in']['value'] - sums['out']['value'],
            })

        if not result:
            raise UserError(_("Aucun mouvement trouvé pour les articles et la période sélectionnés."))

        return result

    # ── Actions ───────────────────────────────────────────────────────────────

    def action_print_report(self):
        self.ensure_one()
        self._get_report_data()
        return self.env.ref('custom_stock.action_report_stock_move_product').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl est requise."))

        data = self._get_report_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Mouvements par produit"

        BLUE     = "1A5276"
        WHITE    = "FFFFFF"
        ROW_A    = "D6EAF8"
        IN_BG    = "D5F5E3"   # entrées : vert clair
        OUT_BG   = "FADBD8"   # sorties : rouge clair
        INT_BG   = "F2F3F4"   # internes : gris clair
        IN_HDR   = "1E8449"
        OUT_HDR  = "B03A2E"
        thin   = Side(style='thin', color="AAAAAA")
        brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
        DIR_BG = {'in': IN_BG, 'out': OUT_BG, 'internal': INT_BG}

        def font(bold=False, color="000000", size=9):
            return Font(name="Arial", bold=bold, color=color, size=size)

        def fill(c):
            return PatternFill("solid", fgColor=c)

        def aln(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        def fmt_date(d):
            return d.strftime('%d/%m/%Y %H:%M') if d else ''

        NCOLS = 7

        ws.merge_cells("A1:G1")
        ws["A1"] = self.company_id.name
        ws["A1"].font = font(bold=True, size=11)

        ws.merge_cells("A2:G2")
        period = ''
        if self.date_debut or self.date_fin:
            period = " — Du %s au %s" % (
                self.date_debut.strftime('%d/%m/%Y') if self.date_debut else '…',
                self.date_fin.strftime('%d/%m/%Y') if self.date_fin else '…',
            )
        ws["A2"] = "MOUVEMENTS PAR PRODUIT" + period
        ws["A2"].font = font(bold=True, color=WHITE, size=11)
        ws["A2"].fill = fill(BLUE)
        ws["A2"].alignment = aln("center")
        ws.row_dimensions[2].height = 18
        ws.append([])

        hdrs = ["Date", "Type d'opération", "Sens", "Référence", "Quantité", "Prix unitaire", "Valeur"]

        for prod in data:
            product = prod['product']

            # Bandeau produit : nom + nombre de mouvements
            ws.append([
                "%s — %s" % (product.default_code or '', product.display_name),
                "", "", "%d mouvement(s)" % prod['nb_mouvements'], "", "", "",
            ])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
            for col in range(1, NCOLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = font(bold=True, color=WHITE, size=10)
                c.fill = fill(BLUE)
                c.border = brd
                c.alignment = aln("left" if col == 1 else "right")
            ws.row_dimensions[r].height = 18

            # Bandeau récap : Entrées / Sorties / Solde — pour distinguer
            # clairement ce qui entre de ce qui sort (vue professionnelle)
            ws.append([
                "Entrées : %.2f (%s) — Valeur : %.2f" % (prod['in_qty'], prod['in_count'], prod['in_value']),
                "", "",
                "Sorties : %.2f (%s) — Valeur : %.2f" % (prod['out_qty'], prod['out_count'], prod['out_value']),
                "", "",
                "Solde : %.2f / %.2f" % (prod['solde_qty'], prod['solde_value']),
            ])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
            for col, bg, color in ((1, IN_BG, IN_HDR), (4, OUT_BG, OUT_HDR), (7, ROW_A, BLUE)):
                c = ws.cell(row=r, column=col)
                c.font = font(bold=True, color=color, size=9)
                c.fill = fill(bg)
                c.border = brd
                c.alignment = aln("left" if col != 7 else "right")
            ws.cell(row=r, column=7).border = brd
            ws.row_dimensions[r].height = 16

            ws.append(hdrs)
            r = ws.max_row
            for col in range(1, NCOLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = font(bold=True, size=9)
                c.fill = fill(ROW_A)
                c.border = brd
                c.alignment = aln("center")

            for line in prod['lines']:
                ws.append([
                    fmt_date(line['date']),
                    line['operation'],
                    line['direction_label'],
                    line['reference'],
                    line['qty'],
                    line['price_unit'],
                    line['value'],
                ])
                r = ws.max_row
                bg = DIR_BG[line['direction']]
                for col in range(1, NCOLS + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = font(size=8)
                    c.fill = fill(bg)
                    c.border = brd
                    c.alignment = aln("right" if col in (5, 6, 7) else ("center" if col == 3 else "left"))
                ws.cell(row=r, column=5).number_format = '#,##0.##'
                ws.cell(row=r, column=6).number_format = '#,##0.00'
                ws.cell(row=r, column=7).number_format = '#,##0.00'

            ws.append([])

        for col, w in enumerate([18, 20, 12, 20, 14, 16, 16], 1):
            ws.column_dimensions[chr(64 + col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_data = base64.b64encode(buf.read()).decode()

        fname = "Mouvements_par_produit_%s.xlsx" % fields.Date.context_today(self).strftime('%d%m%Y')
        att = self.env['ir.attachment'].create({
            'name': fname,
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'new',
        }
