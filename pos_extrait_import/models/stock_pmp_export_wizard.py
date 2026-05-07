# -*- coding: utf-8 -*-
import base64
import io

from odoo import models, fields, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class StockPmpExportWizard(models.TransientModel):
    _name = 'stock.pmp.export.wizard'
    _description = 'Export PMP Stock par Article'

    company_id = fields.Many2one(
        'res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
    )

    def action_export_excel(self):
        self.ensure_one()
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl non installé. pip install openpyxl"))

        wb = openpyxl.Workbook()

        self._build_sheet_pmp(wb)
        self._build_sheet_zero(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        att = self.env['ir.attachment'].create({
            'name':      'export_pmp_articles.xlsx',
            'type':      'binary',
            'datas':     base64.b64encode(output.read()),
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'self',
        }

    # ── Helpers styles ────────────────────────────────────────────────────────

    def _styles(self):
        thin   = Side(style='thin', color='CCCCCC')
        return {
            'border':     Border(left=thin, right=thin, top=thin, bottom=thin),
            'hdr_font':   Font(color="FFFFFF", bold=True, size=11),
            'data_font':  Font(size=10),
            'total_font': Font(bold=True, size=10),
            'center':     Alignment(horizontal='center', vertical='center'),
            'left':       Alignment(horizontal='left',   vertical='center'),
            'right':      Alignment(horizontal='right',  vertical='center'),
            'fill_hdr':   PatternFill("solid", fgColor="1F4E79"),
            'fill_alt':   PatternFill("solid", fgColor="EBF3FB"),
            'fill_white': PatternFill("solid", fgColor="FFFFFF"),
            'fill_total': PatternFill("solid", fgColor="D6E4F0"),
            'fill_zero':  PatternFill("solid", fgColor="C00000"),
        }

    def _write_header(self, ws, headers, st):
        ws.row_dimensions[1].height = 28
        for col, (label, width) in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=label)
            cell.font      = st['hdr_font']
            cell.fill      = st['fill_hdr']
            cell.alignment = st['center']
            cell.border    = st['border']
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def _write_cell(self, ws, row, col, val, align, fill, font, border, num_fmt=None):
        cell           = ws.cell(row=row, column=col, value=val)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align
        cell.border    = border
        if num_fmt:
            cell.number_format = num_fmt

    # ── Onglet 1 : PMP par article (regroupé) ────────────────────────────────

    def _build_sheet_pmp(self, wb):
        move_groups = self.env['stock.move'].read_group(
            domain=[
                ('company_id', '=', self.company_id.id),
                ('state', '=', 'done'),
                '|', ('is_in', '=', True), ('value', '>', 0),
            ],
            fields=['value:sum', 'quantity:sum'],
            groupby=['product_id'],
        )
        if not move_groups:
            raise UserError(_("Aucun mouvement trouvé pour la société sélectionnée."))

        ws     = wb.active
        ws.title = "PMP par Article"
        st     = self._styles()

        HEADERS = [
            ("Code Article",  18),
            ("Nom Article",   42),
            ("Quantité",      14),
            ("Valeur Totale", 18),
            ("PMP",           14),
        ]
        self._write_header(ws, HEADERS, st)

        for row_idx, group in enumerate(move_groups, 2):
            product = self.env['product.product'].browse(group['product_id'][0])
            qty     = group['quantity'] or 0.0
            value   = group['value']       or 0.0
            pmp     = round(value / qty, 2) if qty else 0.0
            fill    = st['fill_alt'] if row_idx % 2 == 0 else st['fill_white']

            ws.row_dimensions[row_idx].height = 20
            for col, (val, align, fmt) in enumerate([
                (product.default_code or '', st['center'], None),
                (product.display_name,       st['left'],   None),
                (round(qty, 3),              st['right'],  '#,##0.00'),
                (round(value, 2),            st['right'],  '#,##0.00'),
                (pmp,                        st['right'],  '#,##0.00'),
            ], 1):
                self._write_cell(ws, row_idx, col, val, align, fill, st['data_font'], st['border'], fmt)

        # Totaux
        total_row   = len(move_groups) + 2
        total_qty   = sum(g['quantity'] or 0 for g in move_groups)
        total_value = sum(g['value']       or 0 for g in move_groups)
        total_pmp   = round(total_value / total_qty, 2) if total_qty else 0.0

        ws.row_dimensions[total_row].height = 22
        for col, (val, align, fmt) in enumerate([
            ("TOTAL",                st['center'], None),
            ("",                     st['left'],   None),
            (round(total_qty,   3),  st['right'],  '#,##0.00'),
            (round(total_value, 2),  st['right'],  '#,##0.00'),
            (total_pmp,              st['right'],  '#,##0.00'),
        ], 1):
            self._write_cell(ws, total_row, col, val, align, st['fill_total'], st['total_font'], st['border'], fmt)

    # ── Onglet 2 : Réceptions à valeur zéro (détail ligne par ligne) ─────────

    def _build_sheet_zero(self, wb):
        moves = self.env['stock.move'].search([
            ('company_id', '=', self.company_id.id),
            ('state', '=', 'done'),
            ('is_in', '=', True),
            ('value', '=', 0),
        ])

        ws       = wb.create_sheet(title="Réceptions à Zéro")
        st       = self._styles()

        HEADERS = [
            ("Code Article", 18),
            ("Nom du mvt",   24),
            ("Date",         14),
            ("Quantité",     14),
        ]
        self._write_header(ws, HEADERS, st)

        # Tri par code article croissant
        moves_sorted = moves.sorted(key=lambda m: m.product_id.default_code or '')

        for row_idx, move in enumerate(moves_sorted, 2):
            fill = st['fill_alt'] if row_idx % 2 == 0 else st['fill_white']
            date_val = move.date.date() if move.date else ''
            ws.row_dimensions[row_idx].height = 20
            for col, (val, align, fmt) in enumerate([
                (move.product_id.default_code or '',              st['center'], None),
                (move.picking_id.name or move.reference or '',    st['left'],   None),
                (date_val,                                        st['center'], 'DD/MM/YYYY'),
                (round(move.quantity, 3),                      st['right'],  '#,##0.00'),
            ], 1):
                self._write_cell(ws, row_idx, col, val, align, fill, st['data_font'], st['border'], fmt)
