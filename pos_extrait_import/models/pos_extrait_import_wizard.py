# -*- coding: utf-8 -*-
import base64
import io
import logging
from datetime import datetime, date
from collections import defaultdict, OrderedDict

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl non installé. pip install openpyxl")


# Correspondance entre les clés internes et les en-têtes Excel acceptés (normalisés en minuscules)
COL_ALIASES = {
    'date_order':   ['date', 'date_order', 'date commande'],
    'session_key':  ['session', 'session_key', 'nom_session'],
    'order_ref':    ['commande', 'order_ref', 'ref_commande', 'num_commande', 'order'],
    'product_ref':  ['code', 'code article', 'code_article', 'product_ref', 'default_code',
                     'ref_produit', 'code_produit', 'code art'],
    'product_name': ['produit', 'designation', 'libelle', 'article', 'product_name', 'nom_produit'],
    'qty':          ['qty', 'quantite', 'qte', 'quantity', 'qté'],
    'price_ht':     ['prix unit ht', 'prix_unit_ht', 'prix ht', 'price_ht', 'ht', 'montant_ht'],
    'price_unit':   ['prix unit ttc', 'prix_unit_ttc', 'prix ttc', 'price_unit', 'ttc',
                     'prix_unit', 'price'],
    'margin':       ['marge', 'margin', 'marge_brute', 'profit'],
}
REQUIRED_COLS = ['date_order', 'session_key', 'order_ref', 'qty', 'price_unit']


class PosExtraitImportWizard(models.TransientModel):
    _name = 'pos.extrait.import.wizard'
    _description = 'Assistant Import Extrait POS'

    # ── Configuration ──────────────────────────────────────────────────────
    pos_config_id = fields.Many2one(
        comodel_name='pos.config',
        string='Point de Vente',
        required=True,
    )
    payment_method_id = fields.Many2one(
        comodel_name='pos.payment.method',
        string='Mode de paiement',
        required=True,
        help="Mode de paiement appliqué à toutes les commandes importées.",
    )

    # ── Fichier ────────────────────────────────────────────────────────────
    import_file     = fields.Binary(string='Fichier Excel (.xlsx)', attachment=False)
    import_filename = fields.Char(string='Nom du fichier')

    # ── État ───────────────────────────────────────────────────────────────
    state = fields.Selection(
        selection=[
            ('draft',   'Configuration'),
            ('preview', 'Prévisualisation'),
            ('done',    'Terminé'),
            ('error',   'Terminé avec erreurs'),
        ],
        default='draft',
        readonly=True,
    )

    # ── Lignes de prévisualisation ─────────────────────────────────────────
    preview_line_ids = fields.One2many(
        comodel_name='pos.extrait.preview.line',
        inverse_name='wizard_id',
        string='Lignes à importer',
        readonly=True,
    )

    # ── Compteurs prévisualisation ─────────────────────────────────────────
    preview_total    = fields.Integer(string='Total lignes',   compute='_compute_preview_stats')
    preview_ok       = fields.Integer(string='OK',             compute='_compute_preview_stats')
    preview_errors   = fields.Integer(string='Erreurs',        compute='_compute_preview_stats')
    preview_orders   = fields.Integer(string='Commandes',      compute='_compute_preview_stats')
    preview_sessions = fields.Integer(string='Sessions',       compute='_compute_preview_stats')
    can_import       = fields.Boolean(string='Importable',     compute='_compute_preview_stats')

    # ── Résultats d'import ─────────────────────────────────────────────────
    import_log       = fields.Html(string="Journal d'import", readonly=True)
    sessions_created = fields.Integer(string='Sessions créées',  readonly=True)
    orders_created   = fields.Integer(string='Commandes créées', readonly=True)
    lines_created    = fields.Integer(string='Lignes créées',    readonly=True)
    errors_count     = fields.Integer(string='Erreurs',          readonly=True)

    # ──────────────────────────────────────────────────────────────────────
    @api.depends('preview_line_ids', 'preview_line_ids.line_state')
    def _compute_preview_stats(self):
        for rec in self:
            lines = rec.preview_line_ids
            rec.preview_total    = len(lines)
            rec.preview_ok       = len(lines.filtered(lambda l: l.line_state == 'ok'))
            rec.preview_errors   = len(lines.filtered(lambda l: l.line_state == 'error'))
            rec.preview_orders   = len(set(lines.mapped('order_ref')))
            rec.preview_sessions = len(set(lines.mapped('session_key')))
            rec.can_import       = rec.preview_total > 0 and rec.preview_errors == 0

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 1 : Télécharger le template Excel
    # ══════════════════════════════════════════════════════════════════════
    def action_download_template(self):
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl non installé. Exécutez : pip install openpyxl"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import POS Extrait"

        COLUMNS = [
            ('date_order',   'DATE\n(JJ/MM/AAAA)',          True,  16),
            ('session_key',  'SESSION\n(nom session)',       True,  18),
            ('order_ref',    'COMMANDE\n(réf. unique)',      True,  22),
            ('product_ref',  'CODE\n(code article)',         True,  14),
            ('product_name', 'PRODUIT\n(nom, optionnel)',    False, 28),
            ('qty',          'QTY\n(négatif = retour)',      True,  12),
            ('price_ht',     'PRIX UNIT HT\n(optionnel, calculé\nsi absent)',  False, 18),
            ('price_unit',   'PRIX UNIT TTC\n(TTC unité)',                    True,  16),
            ('marge',        'MARGE\n(optionnel)',                            False, 14),
        ]

        fill_req = PatternFill("solid", fgColor="1F4E79")
        fill_opt = PatternFill("solid", fgColor="2E75B6")
        fill_ex  = PatternFill("solid", fgColor="EBF3FB")
        font_hdr = Font(color="FFFFFF", bold=True, size=10)
        font_dat = Font(size=10)
        align_c  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin     = Side(style='thin', color='CCCCCC')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[1].height = 48
        for c, (key, label, req, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font      = font_hdr
            if key in ('marge', 'price_ht'):
                cell.fill = fill_opt
            elif req:
                cell.fill = fill_req
            else:
                cell.fill = fill_opt
            cell.alignment = align_c
            cell.border    = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width

        examples = [
            ['01/01/2026', 'IMPORT-JANV', 'BSM – 000001', 4928, 'SAUCISSE POULET', 119,  415.25, 490,  4621],
            ['01/01/2026', 'IMPORT-JANV', 'BSM – 000001', 4942, 'PONDEUSE',        204,  381.36, 450,  6457],
            ['01/01/2026', 'IMPORT-JANV', 'BSM – 000002', 4929, 'LANGUE DE BOEUF', 99,   381.36, 450,  4209],
            ['01/01/2026', 'IMPORT-JANV', 'BSM – 000002', 4808, 'POULET EFFILE',   107, 2330.51, 2750, 55704],
            ['01/02/2026', 'IMPORT-FEVR', 'BSM – 000050', 4805, 'VOLAILLE FERMIER', 3, 11855.93, 13990, 1200],
        ]
        for r, row in enumerate(examples, 2):
            ws.row_dimensions[r].height = 20
            for c, v in enumerate(row, 1):
                cell           = ws.cell(row=r, column=c, value=v)
                cell.font      = font_dat
                cell.fill      = fill_ex
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        att = self.env['ir.attachment'].create({
            'name':      'template_import_pos_extrait.xlsx',
            'type':      'binary',
            'datas':     base64.b64encode(output.read()),
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{att.id}?download=true', 'target': 'self'}

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 2 : Prévisualiser
    # ══════════════════════════════════════════════════════════════════════
    def action_preview(self):
        self.ensure_one()
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl non installé. pip install openpyxl"))
        if not self.import_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        if not self.pos_config_id:
            raise UserError(_("Veuillez sélectionner un Point de Vente."))
        if not self.payment_method_id:
            raise UserError(_("Veuillez sélectionner un mode de paiement."))

        self.preview_line_ids.unlink()

        rows = self._read_excel_file()
        parse_errors = []
        orders_data  = self._parse_rows_to_orders(rows, parse_errors)

        if not orders_data and parse_errors:
            raise UserError(_("Impossible de parser le fichier :\n") + "\n".join(parse_errors[:10]))

        sessions_map = self._group_by_session(orders_data)
        preview_vals = []
        seq          = 0

        for session_key, s_orders in sessions_map.items():
            for order_data in s_orders:
                for line_data in order_data['lines']:
                    seq += 10
                    state, msg, prod_id, computed_price_ht = self._validate_product_line(line_data)

                    preview_vals.append({
                        'wizard_id':           self.id,
                        'sequence':            seq,
                        'session_key':         session_key,
                        'date_order':          order_data['date_order'].strftime('%d/%m/%Y'),
                        'order_ref':           order_data['name'],
                        'product_ref':         line_data.get('product_ref', ''),
                        'product_name':        line_data.get('product_name', ''),
                        'qty':                 line_data['qty'],
                        'price_ht':            computed_price_ht,
                        'price_unit':          line_data['price_unit'],
                        'margin':              line_data.get('margin', 0.0),
                        'line_state':          state,
                        'message':             msg,
                        'resolved_product_id': prod_id,
                    })

        if preview_vals:
            self.env['pos.extrait.preview.line'].create(preview_vals)

        self.write({'state': 'preview'})
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def _validate_product_line(self, line_data):
        """Valide une ligne produit. Retourne (state, message, prod_id, price_ht)."""
        price_ttc = line_data.get('price_unit', 0.0)
        price_ht  = line_data.get('price_ht', 0.0)

        product = self._find_product(line_data.get('product_ref'), line_data.get('product_name'))
        if not product:
            ref  = line_data.get('product_ref', '')
            name = line_data.get('product_name', '')
            return 'error', f"Produit introuvable (code: '{ref}' / nom: '{name}')", 0, price_ht or price_ttc

        if not price_ht:
            price_ht = self._compute_price_ht(product, price_ttc)
            msg = f"✅ Prêt (HT calculé : {price_ht:.2f})"
        else:
            msg = "✅ Prêt"

        return 'ok', msg, product.id, price_ht

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 3 : Retour au draft
    # ══════════════════════════════════════════════════════════════════════
    def action_back_to_draft(self):
        self.ensure_one()
        self.preview_line_ids.unlink()
        self.write({'state': 'draft'})
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 4 : Valider l'import
    # ══════════════════════════════════════════════════════════════════════
    def action_import(self):
        self.ensure_one()
        if not self.preview_line_ids:
            raise UserError(_("Aucune ligne en prévisualisation. Lancez d'abord la prévisualisation."))

        err_lines = self.preview_line_ids.filtered(lambda l: l.line_state == 'error')
        if err_lines:
            raise UserError(_(
                f"{len(err_lines)} ligne(s) en erreur. "
                "Corrigez votre fichier et relancez la prévisualisation."
            ))

        logs = []
        errors = []
        sessions_created = orders_created = lines_created = 0

        try:
            orders_map, sessions_map = self._build_data_from_preview()
            logs.append(
                f"<p>✅ <b>{len(orders_map)}</b> commande(s) / "
                f"<b>{len(sessions_map)}</b> session(s) à créer.</p>"
            )

            for session_key, s_order_refs in sessions_map.items():
                try:
                    s_orders = [orders_map[r] for r in s_order_refs if r in orders_map]
                    session  = self._create_session(session_key, s_orders)
                    sessions_created += 1
                    logs.append(f"<p>📅 Session <b>{session.name}</b> — {len(s_orders)} commande(s).</p>")

                    for order_data in s_orders:
                        try:
                            order, n = self._create_order(session, order_data)
                            orders_created += 1
                            lines_created  += n
                        except Exception as exc:
                            errors.append(f"Commande <b>{order_data.get('name')}</b> : {exc}")
                            _logger.exception("Erreur commande %s", order_data.get('name'))

                except Exception as exc:
                    errors.append(f"Session <b>{session_key}</b> : {exc}")
                    _logger.exception("Erreur session %s", session_key)

        except UserError:
            raise
        except Exception as exc:
            raise UserError(_(f"Erreur inattendue : {exc}"))

        self.write({
            'state':            'error' if errors else 'done',
            'import_log':       self._build_log_html(logs, errors, sessions_created, orders_created, lines_created),
            'sessions_created': sessions_created,
            'orders_created':   orders_created,
            'lines_created':    lines_created,
            'errors_count':     len(errors),
        })
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def _build_data_from_preview(self):
        orders_map   = OrderedDict()
        sessions_map = defaultdict(list)

        for line in self.preview_line_ids.sorted('sequence'):
            ref = line.order_ref
            sk  = line.session_key

            if ref not in orders_map:
                orders_map[ref] = {
                    'name':       ref,
                    'date_order': datetime.strptime(line.date_order, '%d/%m/%Y'),
                    'lines':      [],
                }
                sessions_map[sk].append(ref)

            orders_map[ref]['lines'].append({
                'resolved_product_id': line.resolved_product_id,
                'qty':        line.qty,
                'price_ht':   line.price_ht,
                'price_unit': line.price_unit,
                'margin':     line.margin,
            })

        return orders_map, sessions_map

    # ══════════════════════════════════════════════════════════════════════
    # LECTURE EXCEL
    # ══════════════════════════════════════════════════════════════════════
    def _read_excel_file(self):
        file_data = base64.b64decode(self.import_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        ws = None
        for name in wb.sheetnames:
            nl = name.lower()
            if 'import' in nl or 'pos' in nl or 'vente' in nl or 'passage' in nl or 'extrait' in nl:
                ws = wb[name]
                break
        ws = ws or wb.active

        raw_headers = [
            str(c.value).strip().lower().split('\n')[0].strip() if c.value else ''
            for c in ws[1]
        ]

        col_map = {}
        for idx, raw in enumerate(raw_headers):
            for key, aliases in COL_ALIASES.items():
                if key not in col_map:
                    for alias in aliases:
                        if alias == raw or alias in raw:
                            col_map[key] = idx
                            break

        missing = [k for k in REQUIRED_COLS if k not in col_map]
        if missing:
            raise UserError(_(
                f"Colonnes obligatoires manquantes : {', '.join(missing)}.\n"
                f"En-têtes détectés : {[h for h in raw_headers if h]}\n"
                "Colonnes requises : DATE, SESSION, COMMANDE, QTY, PRIX UNIT TTC."
            ))

        if 'product_ref' not in col_map and 'product_name' not in col_map:
            raise UserError(_(
                "Colonne produit manquante : 'CODE' (ou code article) "
                "ou 'PRODUIT' (ou designation) requis."
            ))

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(v for v in row if v not in (None, '', ' ')):
                continue
            d = {'_row': row_idx}
            for key, col in col_map.items():
                d[key] = row[col] if col < len(row) else None
            rows.append(d)

        if not rows:
            raise UserError(_("Aucune donnée trouvée. Données attendues à partir de la ligne 2."))
        return rows

    # ══════════════════════════════════════════════════════════════════════
    # PARSING LIGNES → COMMANDES
    # ══════════════════════════════════════════════════════════════════════
    def _parse_rows_to_orders(self, rows, errors):
        orders  = OrderedDict()
        counter = 1

        for row in rows:
            rn = row['_row']
            dt = self._parse_date(row.get('date_order'))
            if not dt:
                errors.append(f"Ligne {rn} : date invalide '{row.get('date_order')}'")
                continue

            sk = self._clean_str(row.get('session_key'))
            if not sk:
                errors.append(f"Ligne {rn} : colonne SESSION vide.")
                continue

            pref  = self._to_product_code(row.get('product_ref'))
            pname = self._clean_str(row.get('product_name'))
            if not pref and not pname:
                errors.append(f"Ligne {rn} : code article et nom produit vides.")
                continue

            try:
                qty       = float(row.get('qty') or 0)
                price_ttc = float(row.get('price_unit') or 0)
                price_ht  = float(row.get('price_ht') or 0)   # 0 = absent → sera calculé via TVA
                margin    = float(row.get('margin') or 0)
            except (ValueError, TypeError):
                errors.append(f"Ligne {rn} : valeurs numériques invalides.")
                continue

            ref = self._clean_str(row.get('order_ref'))
            if not ref:
                ref = f"IMPORT-{dt.strftime('%Y%m%d')}-{counter:05d}"
                counter += 1

            # Clé composite : session + ref_commande pour éviter les collisions entre sessions
            key = f"{sk}|{ref}"

            if key not in orders:
                orders[key] = {
                    'name':        ref,
                    'session_key': sk,
                    'date_order':  dt,
                    'lines':       [],
                }

            orders[key]['lines'].append({
                'product_ref':  pref,
                'product_name': pname,
                'qty':          qty,
                'price_ht':     price_ht,
                'price_unit':   price_ttc,
                'margin':       margin,
                '_row':         rn,
            })

        return list(orders.values())

    def _group_by_session(self, orders_data):
        """Groupe les commandes par nom de session (colonne SESSION du fichier)."""
        sessions = defaultdict(list)
        for order in orders_data:
            sessions[order['session_key']].append(order)
        # Trier les sessions par date de première commande
        return dict(sorted(sessions.items(), key=lambda kv: min(o['date_order'] for o in kv[1])))

    # ══════════════════════════════════════════════════════════════════════
    # CRÉATION ENREGISTREMENTS
    # ══════════════════════════════════════════════════════════════════════
    def _create_session(self, session_key, orders):
        PosSession   = self.env['pos.session']
        session_name = f"{session_key}"

        existing = PosSession.search([
            ('name', '=', session_name),
            ('config_id', '=', self.pos_config_id.id),
        ], limit=1)
        if existing:
            return existing

        dates    = [o['date_order'] for o in orders]
        start_dt = min(dates).replace(hour=0,  minute=0,  second=0,  microsecond=0)
        stop_dt  = max(dates).replace(hour=23, minute=59, second=59, microsecond=0)
        ctx      = {'tracking_disable': True, 'mail_notrack': True, 'no_recompute': True}

        session = PosSession.with_context(**ctx).sudo().create({
            'name': session_name, 'config_id': self.pos_config_id.id, 'start_at': start_dt,
        })
        session.with_context(**ctx).sudo().write({'state': 'closed', 'stop_at': stop_dt})
        return session

    def _create_order(self, session, order_data):
        ctx = {'tracking_disable': True, 'mail_notrack': True}

        lines_vals   = []
        amount_total = amount_tax = 0.0

        for line_data in order_data['lines']:
            pid     = line_data.get('resolved_product_id', 0)
            product = self.env['product.product'].browse(pid) if pid else \
                      self._find_product(line_data.get('product_ref', ''), line_data.get('product_name', ''))
            if not product:
                raise UserError(_("Produit non trouvé."))

            qty        = line_data['qty']
            price_ttc  = line_data['price_unit']
            # price_ht est déjà calculé via TVA si absent dans le fichier (cf. _validate_product_line)
            price_ht   = line_data.get('price_ht') or self._compute_price_ht(product, price_ttc)
            margin     = line_data.get('margin', 0.0)

            sub  = price_ht  * qty   # price_subtotal (HT)
            subi = price_ttc * qty   # price_subtotal_incl (TTC)
            # total_cost = price_subtotal - margin  →  margin = price_subtotal - total_cost
            total_cost = sub - margin

            amount_total += subi
            amount_tax   += (subi - sub)

            taxes = product.taxes_id.filtered(lambda t: t.company_id == self.env.company)
            lines_vals.append((0, 0, {
                'product_id':             product.id,
                'qty':                    qty,
                'price_unit':             price_ttc,
                'discount':               0.0,
                'tax_ids':                [(6, 0, taxes.ids)] if taxes else [(5,)],
                'price_subtotal':         sub,
                'price_subtotal_incl':    subi,
                'total_cost':             total_cost,
                'is_total_cost_computed': True,
            }))

        amount_paid = amount_total

        order = self.env['pos.order'].with_context(**ctx).sudo().create({
            'name':          order_data['name'],
            'date_order':    order_data['date_order'],
            'session_id':    session.id,
            'config_id':     self.pos_config_id.id,
            'partner_id':    False,
            'state':         'draft',
            'lines':         lines_vals,
            'amount_total':  amount_total,
            'amount_tax':    amount_tax,
            'amount_paid':   amount_paid,
            'amount_return': 0.0,
        })

        self.env['pos.payment'].with_context(**ctx).sudo().create({
            'pos_order_id':      order.id,
            'payment_method_id': self.payment_method_id.id,
            'amount':            amount_paid,
            'session_id':        session.id,
        })

        order.with_context(**ctx).sudo().write({'state': 'done'})
        return order, len(order_data['lines'])

    # ══════════════════════════════════════════════════════════════════════
    # HELPERS RECHERCHE
    # ══════════════════════════════════════════════════════════════════════
    def _compute_price_ht(self, product, price_ttc):
        """Calcule le prix HT depuis le prix TTC en utilisant les taxes du produit.
        Formule : price_ht = price_ttc / (1 + somme_taux_tva)
        Ne tient compte que des taxes en % non price_include (TVA classique ajoutée sur HT).
        """
        taxes = product.taxes_id.filtered(
            lambda t: t.company_id == self.env.company
                      and t.amount_type == 'percent'
                      and not t.price_include
        )
        if not taxes:
            return price_ttc
        total_rate = sum(t.amount for t in taxes)
        if total_rate == 0:
            return price_ttc
        return round(price_ttc / (1.0 + total_rate / 100.0), 6)

    def _find_product(self, ref, name):
        P = self.env['product.product']
        if ref:
            p = P.search([('default_code', '=', self._to_product_code(ref))], limit=1)
            if p:
                return p
        if name:
            p = P.search([('name', 'ilike', name)], limit=1)
            if p:
                return p
        return None

    # ══════════════════════════════════════════════════════════════════════
    # HELPERS UTILITAIRES
    # ══════════════════════════════════════════════════════════════════════
    def _parse_date(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day, 9, 0, 0)
        s = str(value).strip()
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                    '%d-%m-%Y %H:%M', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    @staticmethod
    def _clean_str(v):
        if v is None:
            return ''
        s = str(v).strip()
        return '' if s.lower() in ('none', 'false', 'nan') else s

    @staticmethod
    def _to_product_code(v):
        """Convertit un code produit en string propre : 4928.0 → '4928', 4928 → '4928'."""
        if v is None:
            return ''
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()

    def _build_log_html(self, logs, errors, sessions, orders, lines):
        err_color = "#C00000"
        ok_color  = "#1F4E79"
        alt       = "#EBF3FB"

        html  = "<div style='font-family:Arial,sans-serif;font-size:13px;'>" + "".join(logs)
        if errors:
            html += f"<br/><p style='color:{err_color};'><b>⚠️ {len(errors)} erreur(s) :</b></p><ul>"
            for e in errors[:100]:
                html += f"<li style='color:{err_color};margin-bottom:4px;'>{e}</li>"
            if len(errors) > 100:
                html += f"<li>… et {len(errors)-100} autre(s). Voir logs serveur.</li>"
            html += "</ul>"

        rows_data = [
            ("Sessions créées",  sessions,    ok_color),
            ("Commandes créées", orders,      ok_color),
            ("Lignes créées",    lines,       ok_color),
            ("Erreurs",          len(errors), err_color if errors else ok_color),
        ]
        html += "<br/><table style='border-collapse:collapse;min-width:320px;margin-top:12px;'>"
        html += (f"<tr style='background:{ok_color};color:#fff;'>"
                 f"<th style='padding:10px 16px;text-align:left;'>Indicateur</th>"
                 f"<th style='padding:10px 16px;text-align:center;'>Valeur</th></tr>")
        for i, (label, value, color) in enumerate(rows_data):
            bg = alt if i % 2 == 0 else "#FFFFFF"
            html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:8px 16px;border-bottom:1px solid #DDD;'>{label}</td>"
                f"<td style='padding:8px 16px;text-align:center;border-bottom:1px solid #DDD;'>"
                f"<b style='color:{color};'>{value}</b></td></tr>"
            )
        html += "</table></div>"
        return html
