# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class ImportBarcodesWizard(models.TransientModel):
    _name = 'import.barcodes.wizard'
    _description = 'Import Excel - Codes-Barres Produits'

    file = fields.Binary(string="Fichier Excel", required=False)
    file_name = fields.Char()
    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('loaded', 'Chargé'),
        ('done', 'Terminé'),
    ], default='draft')
    line_ids = fields.One2many(
        'import.barcodes.wizard.line',
        'wizard_id',
        string='Lignes',
    )
    count_ok = fields.Integer(string='Lignes valides', compute='_compute_counts')
    count_errors = fields.Integer(string='Lignes en erreur', compute='_compute_counts')

    @api.depends('line_ids.status')
    def _compute_counts(self):
        for wizard in self:
            wizard.count_ok = len(wizard.line_ids.filtered(lambda l: l.status == 'ok'))
            wizard.count_errors = len(wizard.line_ids.filtered(lambda l: l.status != 'ok'))

    # -------------------------------------------------------------------------
    # TÉLÉCHARGER LE TEMPLATE EXCEL
    # -------------------------------------------------------------------------
    def action_download_template(self):
        """Génère et télécharge un fichier Excel modèle pour l'import des codes-barres."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl est requise pour générer le template."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Import Codes-Barres"

        headers = ['code_article', 'barcode']
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Ligne exemple
        ws.append(['ART001', '1234567890123'])

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'template_import_codes_barres.xlsx',
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'self',
        }

    # -------------------------------------------------------------------------
    # CHARGER LE FICHIER
    # -------------------------------------------------------------------------
    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez charger un fichier Excel."))

        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        file_data = io.BytesIO(decoded_file)

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_data)
        except Exception:
            raise UserError(_(
                "Impossible de lire le fichier. "
                "Vérifiez qu'il s'agit bien d'un fichier Excel (.xlsx)."
            ))

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

        required_columns = ['code_article', 'barcode']
        for col in required_columns:
            if col not in headers:
                raise UserError(
                    _("Colonne manquante : '%s'\nColonnes trouvées : %s") % (
                        col, ', '.join(headers)
                    )
                )

        idx_code = headers.index('code_article')
        idx_barcode = headers.index('barcode')

        # Récupérer tous les codes-barres déjà enregistrés en base (contrôle unicité rapide)
        existing_barcodes = set(
            self.env['product.multiple.barcodes'].search([]).mapped('product_multi_barcode')
        )

        lines_vals = []

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or row[idx_code] is None:
                continue

            # Normalisation du code article (Excel peut lire les entiers comme float)
            raw_code = row[idx_code]
            if isinstance(raw_code, float) and raw_code == int(raw_code):
                product_code = str(int(raw_code))
            else:
                product_code = str(raw_code).strip()
            if not product_code:
                continue

            # Normalisation du code-barres (Excel peut lire les entiers comme float)
            raw_barcode = row[idx_barcode]
            if raw_barcode is None:
                continue
            if isinstance(raw_barcode, float) and raw_barcode == int(raw_barcode):
                barcode = str(int(raw_barcode))
            else:
                barcode = str(raw_barcode).strip()
            if not barcode:
                continue

            # Recherche du produit par code article (exact, unique)
            product = self.env['product.template'].search(
                [('default_code', '=', product_code)],
                limit=1,
            )

            if not product:
                lines_vals.append((0, 0, {
                    'product_code': product_code,
                    'barcode': barcode,
                    'product_id': False,
                    'status': 'product_not_found',
                }))
                continue

            # Vérifier si le code-barres existe déjà
            if barcode in existing_barcodes:
                lines_vals.append((0, 0, {
                    'product_code': product_code,
                    'barcode': barcode,
                    'product_id': product.id,
                    'status': 'barcode_exists',
                }))
                continue

            # Tout est OK : prêt à importer
            # Marquer le barcode comme "vu" pour éviter les doublons intra-fichier
            existing_barcodes.add(barcode)
            lines_vals.append((0, 0, {
                'product_code': product_code,
                'barcode': barcode,
                'product_id': product.id,
                'status': 'ok',
            }))

        if not lines_vals:
            raise UserError(_(
                "Aucune ligne de données trouvée dans le fichier (hors en-têtes)."
            ))

        self.write({
            'line_ids': lines_vals,
            'state': 'loaded',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # -------------------------------------------------------------------------
    # CONFIRMER L'IMPORT
    # -------------------------------------------------------------------------
    def action_confirm(self):
        self.ensure_one()

        valid_lines = self.line_ids.filtered(lambda l: l.status == 'ok')

        if not valid_lines:
            raise UserError(_(
                "Aucune ligne valide à importer. "
                "Vérifiez les erreurs dans le tableau."
            ))

        created = 0
        for line in valid_lines:
            # Vérifier s'il existe déjà un code-barres actif sur ce produit
            has_active = self.env['product.multiple.barcodes'].search_count([
                ('product_template_id', '=', line.product_id.id),
                ('is_active_barcode', '=', True),
            ])

            self.env['product.multiple.barcodes'].create({
                'product_multi_barcode': line.barcode,
                'product_template_id': line.product_id.id,
                'product_id': line.product_id.product_variant_id.id,
                # Activer automatiquement si aucun code-barres actif n'existe encore
                'is_active_barcode': not has_active,
            })
            created += 1

        self.write({'state': 'done'})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Import terminé"),
                'message': _(
                    "%d code(s)-barres importé(s) avec succès."
                ) % created,
                'type': 'success',
                'sticky': False,
            },
        }


class ImportBarcodesWizardLine(models.TransientModel):
    _name = 'import.barcodes.wizard.line'
    _description = 'Ligne import codes-barres produits'

    wizard_id = fields.Many2one('import.barcodes.wizard', ondelete='cascade')
    product_code = fields.Char(string='Code Article (Excel)', readonly=True)
    barcode = fields.Char(string='Code-Barres (Excel)', readonly=True)
    product_id = fields.Many2one('product.template', string='Produit trouvé', readonly=True)
    status = fields.Selection([
        ('ok', 'Prêt'),
        ('product_not_found', 'Produit non trouvé'),
        ('barcode_exists', 'Code-barres déjà existant'),
    ], string='Statut', readonly=True)
