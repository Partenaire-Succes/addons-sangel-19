# -*- coding: utf-8 -*-
import base64
import io
import logging
from datetime import datetime, date
from collections import defaultdict, OrderedDict

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl non installé. pip install openpyxl")


# Correspondance entre les clés internes et les en-têtes Excel acceptés (normalisés en minuscules)
COL_ALIASES = {
    'date_order':    ['date_order', 'date'],
    'order_ref':     ['order_ref', 'commande', 'order', 'ref_commande', 'num_commande'],
    'customer_ref':  ['customer_ref', 'id client', 'id_client', 'customer_id',
                      'client_id', 'id_clt', 'ref_client'],
    'customer_name': ['customer_name', 'nom', 'name', 'client', 'nom_client'],
    'product_ref':   ['product_ref', 'code article', 'code_article', 'code art',
                      'default_code', 'ref_produit', 'code_produit'],
    'product_name':  ['product_name', 'produit', 'designation', 'libelle',
                      'article', 'nom_produit'],
    'qty':           ['qty', 'quantite', 'qte', 'quantity', 'qté'],
    'price_ht':      ['prix_ht', 'prix ht', 'price_ht', 'ht', 'montant_ht'],
    'price_unit':    ['price_unit', 'prix_ttc', 'prix ttc', 'ttc', 'prix_unit',
                      'price', 'montant_ttc'],
    'discount':      ['discount', 'remise', 'remise_pct'],
    'note':          ['note', 'notes', 'commentaire', 'remarque'],
}
REQUIRED_COLS = ['date_order', 'qty', 'price_unit']


class PosHistoryImportWizard(models.TransientModel):
    _name = 'pos.history.import.wizard'
    _description = 'Assistant Import Historique POS'

    # ── Configuration ──────────────────────────────────────────────────────
    pos_config_id = fields.Many2one(
        comodel_name='pos.config',
        string='Point de Vente',
        required=True,
    )
    payment_method_id = fields.Many2one(
        comodel_name='pos.payment.method',
        string='Mode de paiement',
        required=True,
        help="Mode de paiement appliqué à toutes les commandes importées.",
    )

    # ── Fichier ────────────────────────────────────────────────────────────
    import_file     = fields.Binary(string='Fichier Excel (.xlsx)', attachment=False)
    import_filename = fields.Char(string='Nom du fichier')

    # ── État ───────────────────────────────────────────────────────────────
    state = fields.Selection(
        selection=[
            ('draft',   'Configuration'),
            ('preview', 'Prévisualisation'),
            ('done',    'Terminé'),
            ('error',   'Terminé avec erreurs'),
        ],
        default='draft',
        readonly=True,
    )

    # ── Lignes de prévisualisation ─────────────────────────────────────────
    preview_line_ids = fields.One2many(
        comodel_name='pos.import.preview.line',
        inverse_name='wizard_id',
        string='Lignes à importer',
        readonly=True,
    )

    # ── Compteurs prévisualisation ─────────────────────────────────────────
    preview_total    = fields.Integer(string='Total lignes',   compute='_compute_preview_stats')
    preview_ok       = fields.Integer(string='OK',             compute='_compute_preview_stats')
    preview_warnings = fields.Integer(string='Avertissements', compute='_compute_preview_stats')
    preview_errors   = fields.Integer(string='Erreurs',        compute='_compute_preview_stats')
    preview_orders   = fields.Integer(string='Commandes',      compute='_compute_preview_stats')
    preview_sessions = fields.Integer(string='Sessions',       compute='_compute_preview_stats')
    can_import       = fields.Boolean(string='Importable',     compute='_compute_preview_stats')

    # ── Résultats d'import ─────────────────────────────────────────────────
    import_log       = fields.Html(string="Journal d'import", readonly=True)
    sessions_created = fields.Integer(string='Sessions créées',  readonly=True)
    orders_created   = fields.Integer(string='Commandes créées', readonly=True)
    lines_created    = fields.Integer(string='Lignes créées',    readonly=True)
    errors_count     = fields.Integer(string='Erreurs',          readonly=True)

    # ──────────────────────────────────────────────────────────────────────
    @api.depends('preview_line_ids', 'preview_line_ids.line_state')
    def _compute_preview_stats(self):
        for rec in self:
            lines = rec.preview_line_ids
            rec.preview_total    = len(lines)
            rec.preview_ok       = len(lines.filtered(lambda l: l.line_state == 'ok'))
            rec.preview_warnings = len(lines.filtered(lambda l: l.line_state == 'warning'))
            rec.preview_errors   = len(lines.filtered(lambda l: l.line_state == 'error'))
            rec.preview_orders   = len(set(lines.mapped('order_ref')))
            rec.preview_sessions = len(set(lines.mapped('session_key')))
            rec.can_import       = rec.preview_total > 0 and rec.preview_errors == 0

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 1 : Télécharger le template Excel
    # ══════════════════════════════════════════════════════════════════════
    def action_download_template(self):
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl non installé. Exécutez : pip install openpyxl"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import POS"

        COLUMNS = [
            ('date_order',    'Date\n(JJ/MM/AAAA)',            True,  16),
            ('order_ref',     'Commande\n(réf. unique)',        True,  18),
            ('customer_ref',  'Id client\n(réf. client)',       False, 16),
            ('customer_name', 'Nom\n(nom client)',              False, 22),
            ('product_ref',   'Code article *\n(réf. interne)', True,  16),
            ('product_name',  'Produit\n(nom fallback)',        False, 28),
            ('qty',           'Qty *\n(négatif = retour)',      True,  12),
            ('price_ht',      'Prix_ht *\n(HT unité)',          True,  14),
            ('price_unit',    'Prix_ttc *\n(TTC unité)',        True,  14),
            ('note',          'Note\n(optionnel)',               False, 20),
        ]

        fill_req = PatternFill("solid", fgColor="1F4E79")
        fill_opt = PatternFill("solid", fgColor="2E75B6")
        fill_ex  = PatternFill("solid", fgColor="EBF3FB")
        font_hdr = Font(color="FFFFFF", bold=True, size=10)
        font_dat = Font(size=10)
        align_c  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin     = Side(style='thin', color='CCCCCC')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[1].height = 48
        for c, (key, label, req, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font      = font_hdr
            cell.fill      = fill_req if req else fill_opt
            cell.alignment = align_c
            cell.border    = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = width

        examples = [
            # Commande 2 lignes
            ['01/01/2026', '01-1-53026', None,    None,          '3253', 'SAUCISSE POULET 10X34G', 1,    975,  1150, ''],
            ['01/01/2026', '01-1-53026', None,    None,          '4657', 'PONDEUSE',               3.82, 10514, 11460, ''],
            # Commande 1 ligne avec client
            ['01/01/2026', '01-1-53027', 'C-001', 'Jean Dupont', '4657', 'PONDEUSE',               4.02, 11065, 12060, ''],
            # Retour (qty négative)
            ['01/01/2026', '01-1-53031', None,    None,          '4044', 'LANGUE DE BOEUF',        -1.115, -2604, -2604, 'Retour'],
            # Jour 2
            ['02/01/2026', '01-2-53100', None,    None,          '2676', 'POULET EFFILE',           1.116, 3069, 3069, ''],
        ]
        for r, row in enumerate(examples, 2):
            ws.row_dimensions[r].height = 20
            for c, v in enumerate(row, 1):
                cell           = ws.cell(row=r, column=c, value=v)
                cell.font      = font_dat
                cell.fill      = fill_ex
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        att = self.env['ir.attachment'].create({
            'name':      'template_import_pos_historique.xlsx',
            'type':      'binary',
            'datas':     base64.b64encode(output.read()),
            'mimetype':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id':    self.id,
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{att.id}?download=true', 'target': 'self'}

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 2 : Prévisualiser
    # ══════════════════════════════════════════════════════════════════════
    def action_preview(self):
        """Parse le fichier Excel, valide et crée les lignes de prévisualisation."""
        self.ensure_one()
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl non installé. pip install openpyxl"))
        if not self.import_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        if not self.pos_config_id:
            raise UserError(_("Veuillez sélectionner un Point de Vente."))
        if not self.payment_method_id:
            raise UserError(_("Veuillez sélectionner un mode de paiement."))

        self.preview_line_ids.unlink()

        rows = self._read_excel_file()
        parse_errors = []
        orders_data  = self._parse_rows_to_orders(rows, parse_errors)

        if not orders_data and parse_errors:
            raise UserError(_("Impossible de parser le fichier :\n") + "\n".join(parse_errors[:10]))

        sessions_map = self._group_by_session(orders_data)
        preview_vals = []
        seq          = 0

        for session_key, s_orders in sessions_map.items():
            for order_data in s_orders:
                is_first = True
                for line_data in order_data['lines']:
                    seq += 10
                    state, msg, prod_id, partner_id = self._validate_product_line(order_data, line_data)

                    preview_vals.append({
                        'wizard_id':    self.id,
                        'sequence':     seq,
                        'session_key':  session_key,
                        'date_order':   order_data['date_order'].strftime('%d/%m/%Y'),
                        'order_ref':    order_data['name'],
                        'customer_info': self._fmt_customer(
                            order_data.get('customer_ref'),
                            order_data.get('customer_name'),
                        ) if is_first else '',
                        'product_ref':  line_data.get('product_ref', ''),
                        'product_name': line_data.get('product_name', ''),
                        'qty':          line_data['qty'],
                        'price_ht':     line_data.get('price_ht', 0.0),
                        'price_unit':   line_data['price_unit'],
                        'note':         line_data.get('note', ''),
                        'line_state':   state,
                        'message':      msg,
                        'resolved_product_id': prod_id,
                        'resolved_partner_id': partner_id if is_first else 0,
                    })
                    is_first = False

        if preview_vals:
            self.env['pos.import.preview.line'].create(preview_vals)

        self.write({'state': 'preview'})
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def _validate_product_line(self, order_data, line_data):
        """Valide une ligne produit. Retourne (state, message, prod_id, partner_id)."""
        msgs  = []
        state = 'ok'
        prod_id = partner_id = 0

        # Produit
        product = self._find_product(line_data.get('product_ref'), line_data.get('product_name'))
        if product:
            prod_id = product.id
        else:
            state = 'error'
            msgs.append(
                f"Produit introuvable "
                f"(réf: '{line_data.get('product_ref')}' / nom: '{line_data.get('product_name')}')"
            )

        # Client (optionnel)
        cref  = order_data.get('customer_ref', '')
        cname = order_data.get('customer_name', '')
        if cref or cname:
            partner = self._find_partner_no_create(cref, cname)
            if partner:
                partner_id = partner.id
            else:
                if state != 'error':
                    state = 'warning'
                msgs.append(f"Client '{cname or cref}' non trouvé → sera créé à l'import.")

        msg = ' | '.join(msgs) if msgs else ('✅ Prêt' if state == 'ok' else '')
        return state, msg, prod_id, partner_id

    def _find_partner_no_create(self, ref, name):
        P = self.env['res.partner']
        if ref:
            p = P.search([('ref', '=', ref)], limit=1)
            if p:
                return p
        if name:
            p = P.search([('name', 'ilike', name)], limit=1)
            if p:
                return p
        return None

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 3 : Retour au draft
    # ══════════════════════════════════════════════════════════════════════
    def action_back_to_draft(self):
        self.ensure_one()
        self.preview_line_ids.unlink()
        self.write({'state': 'draft'})
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    # ══════════════════════════════════════════════════════════════════════
    # ACTION 4 : Valider l'import
    # ══════════════════════════════════════════════════════════════════════
    def action_import(self):
        self.ensure_one()
        if not self.preview_line_ids:
            raise UserError(_("Aucune ligne en prévisualisation. Lancez d'abord la prévisualisation."))

        err_lines = self.preview_line_ids.filtered(lambda l: l.line_state == 'error')
        if err_lines:
            raise UserError(_(
                f"{len(err_lines)} ligne(s) en erreur. "
                "Corrigez votre fichier et relancez la prévisualisation."
            ))

        logs = []
        errors = []
        sessions_created = orders_created = lines_created = 0

        try:
            orders_map, sessions_map = self._build_data_from_preview()
            logs.append(
                f"<p>✅ <b>{len(orders_map)}</b> commande(s) / "
                f"<b>{len(sessions_map)}</b> session(s) à créer.</p>"
            )

            for session_key, s_order_refs in sessions_map.items():
                try:
                    s_orders = [orders_map[r] for r in s_order_refs if r in orders_map]
                    session  = self._create_session(session_key, s_orders)
                    sessions_created += 1
                    logs.append(f"<p>📅 Session <b>{session.name}</b> — {len(s_orders)} commande(s).</p>")

                    for order_data in s_orders:
                        try:
                            order, n = self._create_order(session, order_data)
                            orders_created += 1
                            lines_created  += n
                        except Exception as exc:
                            errors.append(f"Commande <b>{order_data.get('name')}</b> : {exc}")
                            _logger.exception("Erreur commande %s", order_data.get('name'))

                except Exception as exc:
                    errors.append(f"Session <b>{session_key}</b> : {exc}")
                    _logger.exception("Erreur session %s", session_key)

        except UserError:
            raise
        except Exception as exc:
            raise UserError(_(f"Erreur inattendue : {exc}"))

        self.write({
            'state':            'error' if errors else 'done',
            'import_log':       self._build_log_html(logs, errors, sessions_created, orders_created, lines_created),
            'sessions_created': sessions_created,
            'orders_created':   orders_created,
            'lines_created':    lines_created,
            'errors_count':     len(errors),
        })
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
        }

    def _build_data_from_preview(self):
        orders_map   = OrderedDict()
        sessions_map = defaultdict(list)

        for line in self.preview_line_ids.sorted('sequence'):
            ref = line.order_ref
            sk  = line.session_key

            if ref not in orders_map:
                # Reconstituer ref/nom bruts depuis customer_info ("ref / nom")
                info = line.customer_info or ''
                if ' / ' in info:
                    raw_ref, raw_name = info.split(' / ', 1)
                else:
                    raw_ref, raw_name = '', info

                orders_map[ref] = {
                    'name':               ref,
                    'date_order':         datetime.strptime(line.date_order, '%d/%m/%Y'),
                    'customer_ref':       raw_ref,
                    'customer_name':      raw_name,
                    'resolved_partner_id': line.resolved_partner_id,
                    'lines':              [],
                }
                sessions_map[sk].append(ref)
            else:
                # Enrichir les infos client si absentes sur la 1ère ligne
                if line.resolved_partner_id and not orders_map[ref]['resolved_partner_id']:
                    orders_map[ref]['resolved_partner_id'] = line.resolved_partner_id

            orders_map[ref]['lines'].append({
                'resolved_product_id': line.resolved_product_id,
                'qty':       line.qty,
                'price_ht':  line.price_ht,
                'price_unit': line.price_unit,
                'note':      line.note,
            })

        return orders_map, sessions_map

    # ══════════════════════════════════════════════════════════════════════
    # LECTURE EXCEL
    # ══════════════════════════════════════════════════════════════════════
    def _read_excel_file(self):
        file_data = base64.b64decode(self.import_file)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        ws = None
        for name in wb.sheetnames:
            nl = name.lower()
            if 'import' in nl or 'pos' in nl or 'vente' in nl or 'passage' in nl:
                ws = wb[name]
                break
        ws = ws or wb.active

        # Normaliser les en-têtes
        raw_headers = [
            str(c.value).strip().lower().split('\n')[0].strip() if c.value else ''
            for c in ws[1]
        ]

        col_map = {}
        for idx, raw in enumerate(raw_headers):
            for key, aliases in COL_ALIASES.items():
                if key not in col_map:
                    for alias in aliases:
                        if alias == raw or alias in raw:
                            col_map[key] = idx
                            break

        missing = [k for k in REQUIRED_COLS if k not in col_map]
        if missing:
            raise UserError(_(
                f"Colonnes obligatoires manquantes : {', '.join(missing)}.\n"
                f"En-têtes détectés : {[h for h in raw_headers if h]}\n"
                "Colonnes requises : date_order (ou Date), qty (ou Qty), "
                "price_unit (ou Prix_ttc)."
            ))

        # Vérifier qu'on a au moins product_ref ou product_name
        if 'product_ref' not in col_map and 'product_name' not in col_map:
            raise UserError(_(
                "Colonne produit manquante : 'Code article' (ou product_ref) "
                "ou 'Produit' (ou product_name) requis."
            ))

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(v for v in row if v not in (None, '', ' ')):
                continue
            d = {'_row': row_idx}
            for key, col in col_map.items():
                d[key] = row[col] if col < len(row) else None
            rows.append(d)

        if not rows:
            raise UserError(_("Aucune donnée trouvée. Données attendues à partir de la ligne 2."))
        return rows

    # ══════════════════════════════════════════════════════════════════════
    # PARSING LIGNES → COMMANDES
    # ══════════════════════════════════════════════════════════════════════
    def _parse_rows_to_orders(self, rows, errors):
        orders  = OrderedDict()
        counter = 1

        for row in rows:
            rn = row['_row']
            dt = self._parse_date(row.get('date_order'))
            if not dt:
                errors.append(f"Ligne {rn} : date invalide '{row.get('date_order')}'")
                continue

            pref  = self._clean_str(row.get('product_ref'))
            pname = self._clean_str(row.get('product_name'))
            if not pref and not pname:
                errors.append(f"Ligne {rn} : code article et nom produit vides.")
                continue

            try:
                qty       = float(row.get('qty') or 0)
                price_ttc = float(row.get('price_unit') or 0)
                price_ht  = float(row.get('price_ht') or price_ttc)
                disc      = float(row.get('discount') or 0)
            except (ValueError, TypeError):
                errors.append(f"Ligne {rn} : valeurs numériques invalides.")
                continue

            ref = self._clean_str(row.get('order_ref'))
            if not ref:
                ref = f"IMPORT-{dt.strftime('%Y%m%d')}-{counter:05d}"
                counter += 1

            if ref not in orders:
                orders[ref] = {
                    'name':          ref,
                    'date_order':    dt,
                    'customer_ref':  self._clean_str(row.get('customer_ref')),
                    'customer_name': self._clean_str(row.get('customer_name')),
                    'lines':         [],
                }
            else:
                # Enrichir le client si absent sur les premières lignes
                if not orders[ref]['customer_ref'] and row.get('customer_ref'):
                    orders[ref]['customer_ref'] = self._clean_str(row['customer_ref'])
                if not orders[ref]['customer_name'] and row.get('customer_name'):
                    orders[ref]['customer_name'] = self._clean_str(row['customer_name'])

            orders[ref]['lines'].append({
                'product_ref':  pref,
                'product_name': pname,
                'qty':          qty,
                'price_ht':     price_ht,
                'price_unit':   price_ttc,
                'discount':     disc,
                'note':         self._clean_str(row.get('note')),
                '_row':         rn,
            })

        return list(orders.values())

    def _group_by_session(self, orders_data):
        """1 session par jour."""
        sessions = defaultdict(list)
        for order in orders_data:
            key = order['date_order'].strftime('%Y-%m-%d')
            sessions[key].append(order)
        return dict(sorted(sessions.items()))

    # ══════════════════════════════════════════════════════════════════════
    # CRÉATION ENREGISTREMENTS
    # ══════════════════════════════════════════════════════════════════════
    def _create_session(self, session_key, orders):
        PosSession   = self.env['pos.session']
        session_name = f"[IMPORT] {session_key}"

        existing = PosSession.search([
            ('name', '=', session_name),
            ('config_id', '=', self.pos_config_id.id),
        ], limit=1)
        if existing:
            return existing

        dates    = [o['date_order'] for o in orders]
        start_dt = min(dates).replace(hour=0,  minute=0,  second=0,  microsecond=0)
        stop_dt  = max(dates).replace(hour=23, minute=59, second=59, microsecond=0)
        ctx      = {'tracking_disable': True, 'mail_notrack': True, 'no_recompute': True}

        session = PosSession.with_context(**ctx).sudo().create({
            'name': session_name, 'config_id': self.pos_config_id.id, 'start_at': start_dt,
        })
        session.with_context(**ctx).sudo().write({'state': 'closed', 'stop_at': stop_dt})
        return session

    def _create_order(self, session, order_data):
        ctx = {'tracking_disable': True, 'mail_notrack': True}

        partner_id = order_data.get('resolved_partner_id', 0)
        if not partner_id:
            partner    = self._find_or_create_partner(
                order_data.get('customer_ref', ''),
                order_data.get('customer_name', ''),
            )
            partner_id = partner.id if partner else False

        lines_vals   = []
        amount_total = amount_tax = 0.0

        for line_data in order_data['lines']:
            pid     = line_data.get('resolved_product_id', 0)
            product = self.env['product.product'].browse(pid) if pid else \
                      self._find_product(line_data.get('product_ref', ''), line_data.get('product_name', ''))
            if not product:
                raise UserError(_("Produit non trouvé."))

            qty       = line_data['qty']
            price_ttc = line_data['price_unit']
            price_ht  = line_data.get('price_ht', price_ttc)

            # Utiliser les montants exacts du fichier source
            sub  = price_ht  * qty  # HT
            subi = price_ttc * qty  # TTC

            amount_total += subi
            amount_tax   += (subi - sub)

            taxes = product.taxes_id.filtered(lambda t: t.company_id == self.env.company)
            lv = {
                'product_id':          product.id,
                'qty':                 qty,
                'price_unit':          price_ttc,
                'discount':            line_data.get('discount', 0.0),
                'tax_ids':             [(6, 0, taxes.ids)] if taxes else [(5,)],
                'price_subtotal':      sub,
                'price_subtotal_incl': subi,
            }
            if line_data.get('note'):
                lv['note'] = line_data['note']
            lines_vals.append((0, 0, lv))

        amount_paid   = amount_total
        amount_return = 0.0

        order = self.env['pos.order'].with_context(**ctx).sudo().create({
            'name':          order_data['name'],
            'date_order':    order_data['date_order'],
            'session_id':    session.id,
            'config_id':     self.pos_config_id.id,
            'partner_id':    partner_id or False,
            'state':         'done',
            'lines':         lines_vals,
            'amount_total':  amount_total,
            'amount_tax':    amount_tax,
            'amount_paid':   amount_paid,
            'amount_return': amount_return,
        })

        self.env['pos.payment'].with_context(**ctx).sudo().create({
            'pos_order_id':      order.id,
            'payment_method_id': self.payment_method_id.id,
            'amount':            amount_paid,
            'session_id':        session.id,
        })

        return order, len(order_data['lines'])

    # ══════════════════════════════════════════════════════════════════════
    # HELPERS RECHERCHE
    # ══════════════════════════════════════════════════════════════════════
    def _find_or_create_partner(self, ref, name):
        P = self.env['res.partner']
        if ref:
            p = P.search([('ref', '=', ref)], limit=1)
            if p:
                return p
        if name:
            p = P.search([('name', 'ilike', name)], limit=1)
            if p:
                return p
            return P.sudo().create({'name': name, 'customer_rank': 1})
        return P

    def _find_product(self, ref, name):
        P = self.env['product.product']
        if ref:
            p = P.search([('default_code', '=', ref)], limit=1)
            if p:
                return p
        if name:
            p = P.search([('name', 'ilike', name)], limit=1)
            if p:
                return p
        return None

    # ══════════════════════════════════════════════════════════════════════
    # HELPERS UTILITAIRES
    # ══════════════════════════════════════════════════════════════════════
    def _parse_date(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day, 9, 0, 0)
        s = str(value).strip()
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                    '%d-%m-%Y %H:%M', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    @staticmethod
    def _clean_str(v):
        if v is None:
            return ''
        s = str(v).strip()
        return '' if s.lower() in ('none', 'false', 'nan') else s

    @staticmethod
    def _to_float(v):
        try:
            return float(v or 0)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _fmt_customer(ref, name):
        parts = [p for p in [ref, name] if p]
        return ' / '.join(parts) if parts else ''

    def _build_log_html(self, logs, errors, sessions, orders, lines):
        err_color = "#C00000"
        ok_color  = "#1F4E79"
        alt       = "#EBF3FB"

        html  = "<div style='font-family:Arial,sans-serif;font-size:13px;'>" + "".join(logs)
        if errors:
            html += f"<br/><p style='color:{err_color};'><b>⚠️ {len(errors)} erreur(s) :</b></p><ul>"
            for e in errors[:100]:
                html += f"<li style='color:{err_color};margin-bottom:4px;'>{e}</li>"
            if len(errors) > 100:
                html += f"<li>… et {len(errors)-100} autre(s). Voir logs serveur.</li>"
            html += "</ul>"

        rows_data = [
            ("Sessions créées",  sessions,    ok_color),
            ("Commandes créées", orders,      ok_color),
            ("Lignes créées",    lines,       ok_color),
            ("Erreurs",          len(errors), err_color if errors else ok_color),
        ]
        html += "<br/><table style='border-collapse:collapse;min-width:320px;margin-top:12px;'>"
        html += (f"<tr style='background:{ok_color};color:#fff;'>"
                 f"<th style='padding:10px 16px;text-align:left;'>Indicateur</th>"
                 f"<th style='padding:10px 16px;text-align:center;'>Valeur</th></tr>")
        for i, (label, value, color) in enumerate(rows_data):
            bg = alt if i % 2 == 0 else "#FFFFFF"
            html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:8px 16px;border-bottom:1px solid #DDD;'>{label}</td>"
                f"<td style='padding:8px 16px;text-align:center;border-bottom:1px solid #DDD;'>"
                f"<b style='color:{color};'>{value}</b></td></tr>"
            )
        html += "</table></div>"
        return html
