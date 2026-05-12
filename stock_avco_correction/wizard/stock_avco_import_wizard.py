import base64
import io
import logging

import openpyxl
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class StockAvcoImportWizard(models.TransientModel):
    _name = 'stock.avco.import.wizard'
    _description = "Assistant de correction AVCO par import Excel"

    state = fields.Selection([
        ('import',  'Import fichier'),
        ('preview', 'Vérification'),
        ('done',    'Terminé'),
    ], default='import', string="Etape")

    company_id = fields.Many2one(
        'res.company', string="Société", required=True,
        default=lambda self: self.env.company,
    )

    excel_file     = fields.Binary(string="Fichier Excel", attachment=False)
    excel_filename = fields.Char(string="Nom du fichier")

    line_ids = fields.One2many(
        'stock.avco.import.wizard.line', 'wizard_id', string="Lignes"
    )

    summary_html         = fields.Html(string="Résumé",                   readonly=True)
    nb_moves_corrected   = fields.Integer(string="Mouvements corrigés",   readonly=True)
    nb_po_updated        = fields.Integer(string="Commandes mises à jour", readonly=True)
    nb_invoice_updated   = fields.Integer(string="Factures mises à jour",  readonly=True)
    nb_products_updated  = fields.Integer(string="Produits mis à jour",    readonly=True)
    total_value_injected = fields.Float(string="Valeur totale (FCFA)",     readonly=True)
    nb_not_found         = fields.Integer(string="Codes non trouvés",      readonly=True)

    # ------------------------------------------------------------------
    # ÉTAPE 1 — Analyser
    # ------------------------------------------------------------------
    def action_load_file(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))

        rows = self._parse_excel(self.excel_file, self.excel_filename)
        if not rows:
            raise UserError(_(
                "Fichier vide ou format incorrect.\n"
                "Colonnes attendues : 'code article' et 'pmp'"
            ))

        self.line_ids.unlink()
        lines_vals = []
        not_found  = []

        for code_raw, pmp_excel in rows:
            if not code_raw:
                continue

            code_padded = str(code_raw).strip().zfill(4)
            pmp_excel   = float(pmp_excel) if pmp_excel else 0.0

            product = self.env['product.product'].search([
                ('default_code', '=', code_padded),
                ('active', '=', True),
            ], limit=1)

            if not product:
                not_found.append(code_padded)
                lines_vals.append({
                    'wizard_id':    self.id,
                    'code_article': code_padded,
                    'product_id':   False,
                    'pmp_excel':    pmp_excel,
                    'nb_moves':     0,
                    'total_qty':    0.0,
                    'old_value':    0.0,
                    'new_value':    0.0,
                    'state':        'not_found',
                })
                continue

            moves = self.env['stock.move'].search([
                ('product_id', '=', product.id),
                ('is_in', '=', True),
                ('picking_id', '!=', False),
                ('state', '=', 'done'),
                ('company_id', '=', self.company_id.id),
            ])

            if not moves:
                lines_vals.append({
                    'wizard_id':    self.id,
                    'code_article': code_padded,
                    'product_id':   product.id,
                    'pmp_excel':    pmp_excel,
                    'nb_moves':     0,
                    'total_qty':    0.0,
                    'old_value':    0.0,
                    'new_value':    0.0,
                    'state':        'no_move',
                })
                continue

            total_qty  = sum(m.quantity for m in moves)
            old_value  = sum(m.value for m in moves)
            new_value  = total_qty * pmp_excel

            lines_vals.append({
                'wizard_id':    self.id,
                'code_article': code_padded,
                'product_id':   product.id,
                'pmp_excel':    pmp_excel,
                'nb_moves':     len(moves),
                'total_qty':    total_qty,
                'old_value':    old_value,
                'new_value':    new_value,
                'state':        'ready',
            })

        self.env['stock.avco.import.wizard.line'].create(lines_vals)

        nb_ready = sum(1 for v in lines_vals if v['state'] == 'ready')
        nb_no_move = sum(1 for v in lines_vals if v['state'] == 'no_move')

        self.summary_html = self._build_load_summary_html(
            len(rows), nb_ready, nb_no_move, len(not_found),
            not_found, self.company_id.name,
        )
        self.state = 'preview'
        return self._reload()

    # ------------------------------------------------------------------
    # ÉTAPE 2 — Appliquer
    # ------------------------------------------------------------------
    def action_apply(self):
        self.ensure_one()

        company = self.company_id
        lines_to_apply = self.line_ids.filtered(
            lambda l: l.state in ('ready', 'no_move') and l.product_id
        )
        if not lines_to_apply:
            raise UserError(_("Aucune ligne à corriger."))

        nb_moves_fixed = 0
        nb_po_fixed    = 0
        nb_inv_fixed   = 0
        nb_products    = 0
        total_value    = 0.0

        for line in lines_to_apply:
            moves = self.env['stock.move'].search([
                ('product_id', '=', line.product_id.id),
                ('is_in', '=', True),
                ('picking_id', '!=', False),
                ('state', '=', 'done'),
                ('company_id', '=', company.id),
            ])
            if not moves:
                continue

            po_lines_fixed  = set()
            invoices_fixed  = set()
            product_touched = False

            for move in moves:
                correct_price = line.pmp_excel
                correct_value = move.quantity * correct_price

                update_vals = {}
                if abs(move.price_unit - correct_price) > 0.01:
                    update_vals['price_unit'] = correct_price
                if abs(move.value - correct_value) > 0.01:
                    update_vals['value'] = correct_value

                if update_vals:
                    if not move.avco_corrected:
                        move.write({
                            'avco_original_price_unit': move.price_unit,
                            'avco_original_value':      move.value,
                            'avco_correction_date':     fields.Datetime.now(),
                            'avco_correction_user_id':  self.env.user.id,
                            'avco_corrected':           True,
                        })

                    if 'price_unit' in update_vals:
                        move.write({'price_unit': correct_price})

                    if 'value' in update_vals:
                        self.env['product.value'].create({
                            'move_id':    move.id,
                            'value':      correct_value,
                            'company_id': company.id,
                        })

                    nb_moves_fixed  += 1
                    product_touched  = True

                total_value += correct_value

                pol = move.purchase_line_id

                if pol and pol.company_id == company and \
                        pol.id not in po_lines_fixed and \
                        abs(pol.price_unit - correct_price) > 0.01:
                    pol.write({'price_unit': correct_price})
                    po_lines_fixed.add(pol.id)
                    nb_po_fixed    += 1
                    product_touched = True

                if pol and pol.company_id == company:
                    for invoice in pol.order_id.invoice_ids.filtered(
                        lambda inv: inv.move_type == 'in_invoice'
                                    and inv.state != 'cancel'
                                    and inv.company_id == company
                                    and inv.id not in invoices_fixed
                    ):
                        inv_lines = invoice.invoice_line_ids.filtered(
                            lambda il: il.product_id == move.product_id
                                       and abs(il.price_unit - correct_price) > 0.01
                        )
                        if inv_lines:
                            inv_lines.write({'price_unit': correct_price})
                            invoices_fixed.add(invoice.id)
                            nb_inv_fixed += 1

            # Mise à jour standard_price dans tous les cas (avec ou sans mouvements)
            variant = line.product_id.with_company(company).sudo()
            if abs(variant.standard_price - line.pmp_excel) > 0.01:
                old_std = variant.standard_price
                variant.write({'standard_price': line.pmp_excel})
                _logger.info(
                    "standard_price [%s] %s : %.2f → %.2f",
                    company.name, line.product_id.display_name,
                    old_std, line.pmp_excel,
                )

            if product_touched:
                nb_products += 1

            line.write({'state': 'done'})

        self.write({
            'state':                'done',
            'nb_moves_corrected':   nb_moves_fixed,
            'nb_po_updated':        nb_po_fixed,
            'nb_invoice_updated':   nb_inv_fixed,
            'nb_products_updated':  nb_products,
            'total_value_injected': total_value,
            'nb_not_found':         len(self.line_ids.filtered(lambda l: l.state == 'not_found')),
        })
        self.summary_html = self._build_apply_summary_html(
            nb_products, nb_moves_fixed, nb_po_fixed,
            nb_inv_fixed, total_value, company.name,
        )
        _logger.info(
            "Correction AVCO [%s] : %d produits | %d moves | %d PO | %d factures | %.2f FCFA",
            company.name, nb_products, nb_moves_fixed, nb_po_fixed, nb_inv_fixed, total_value,
        )
        return self._reload()

    def action_reset(self):
        self.line_ids.unlink()
        self.write({
            'state': 'import', 'excel_file': False, 'excel_filename': False,
            'summary_html': False, 'nb_moves_corrected': 0, 'nb_po_updated': 0,
            'nb_invoice_updated': 0, 'nb_products_updated': 0,
            'total_value_injected': 0.0, 'nb_not_found': 0,
        })
        return self._reload()

    # ------------------------------------------------------------------
    # PARSING EXCEL
    # ------------------------------------------------------------------
    def _parse_excel(self, file_b64, filename):
        file_bytes = base64.b64decode(file_b64)
        rows = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            headers = {}
            for col in ws.iter_cols(1, ws.max_column, 1, 1):
                cell = col[0]
                if cell.value:
                    h = str(cell.value).strip().lower()
                    if 'code' in h:
                        headers.setdefault('code', cell.column - 1)
                    elif any(k in h for k in ('pmp', 'prix', 'cout', 'coût', 'price')):
                        headers.setdefault('pmp', cell.column - 1)

            if 'code' not in headers or 'pmp' not in headers:
                found = [str(ws.cell(1, i + 1).value) for i in range(ws.max_column)
                         if ws.cell(1, i + 1).value]
                raise UserError(_(
                    "Colonnes 'code article' et 'pmp' non trouvées.\nDétectées : %s"
                ) % ', '.join(found))

            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[headers['code']]
                pmp  = row[headers['pmp']]
                if code is not None:
                    try:
                        rows.append((str(code).strip(), float(pmp) if pmp else 0.0))
                    except (ValueError, TypeError):
                        continue

        except UserError:
            raise
        except Exception as e:
            if HAS_XLRD:
                try:
                    wb  = xlrd.open_workbook(file_contents=file_bytes)
                    ws  = wb.sheet_by_index(0)
                    hdr = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]
                    ci  = next((i for i, h in enumerate(hdr) if 'code' in h), None)
                    pi  = next((i for i, h in enumerate(hdr)
                                if any(k in h for k in ('pmp', 'prix', 'cout'))), None)
                    if ci is None or pi is None:
                        raise UserError(_("Colonnes non trouvées."))
                    for r in range(1, ws.nrows):
                        code = str(ws.cell_value(r, ci)).strip()
                        pmp  = ws.cell_value(r, pi)
                        if code:
                            try:
                                rows.append((code, float(pmp) if pmp else 0.0))
                            except (ValueError, TypeError):
                                continue
                except UserError:
                    raise
                except Exception as e2:
                    raise UserError(_("Impossible de lire le fichier : %s") % str(e2))
            else:
                raise UserError(_("Impossible de lire le fichier : %s") % str(e))
        return rows

    # ------------------------------------------------------------------
    # HTML
    # ------------------------------------------------------------------
    def _build_load_summary_html(self, total, nb_ready, nb_no_move,
                                  nb_not_found, not_found_codes, company_name):
        nf_list = ''
        if not_found_codes:
            items   = ''.join(f'<li>{c}</li>' for c in not_found_codes[:20])
            more    = (f'<li>... et {len(not_found_codes) - 20} autres</li>'
                       if len(not_found_codes) > 20 else '')
            nf_list = f'<ul style="color:#dc3545">{items}{more}</ul>'

        return f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <div style="background:#e8f4fd;border:1px solid #bee5eb;border-radius:6px;
                      padding:8px 12px;margin-bottom:10px;font-size:13px;">
            <b>Société :</b> {company_name}
          </div>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Total produits dans le fichier</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{total}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#28a745;">Produits à corriger (mouvements trouvés)</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#28a745;">{nb_ready}</td>
            </tr>
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;color:#e67e22;">Sans mouvement de réception</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#e67e22;">{nb_no_move}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#dc3545;">Codes non trouvés dans Odoo</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#dc3545;">{nb_not_found}</td>
            </tr>
          </table>
          {f'<h4 style="color:#dc3545;margin-top:12px;">Codes non trouvés :</h4>{nf_list}' if nf_list else ''}
        </div>
        """

    def _build_apply_summary_html(self, nb_products, nb_moves, nb_po,
                                   nb_invoices, total_value, company_name):
        return f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <div style="background:#e8f4fd;border:1px solid #bee5eb;border-radius:6px;
                      padding:8px 12px;margin-bottom:10px;">
            <b>Société :</b> {company_name}
          </div>
          <h3 style="border-bottom:2px solid #28a745;padding-bottom:8px;color:#28a745;">
            Corrections appliquées avec succès !
          </h3>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Produits traités</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_products}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;">Mouvements de stock corrigés</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_moves}</td>
            </tr>
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Commandes d'achat mises à jour</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_po}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;">Factures fournisseurs mises à jour</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_invoices}</td>
            </tr>
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Valeur totale recalculée</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">
                {total_value:,.2f} FCFA
              </td>
            </tr>
          </table>
        </div>
        """

    def _reload(self):
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }


class StockAvcoImportWizardLine(models.TransientModel):
    _name = 'stock.avco.import.wizard.line'
    _description = "Ligne de correction AVCO"
    _order = 'state, code_article'

    wizard_id    = fields.Many2one('stock.avco.import.wizard', ondelete='cascade')
    code_article = fields.Char(string="Code Article", readonly=True)
    product_id   = fields.Many2one('product.product', string="Produit",          readonly=True)
    pmp_excel    = fields.Float(string="PMP Excel (FCFA)", digits=(16, 4),       readonly=True)

    nb_moves  = fields.Integer(string="Nb mouvements", readonly=True)
    total_qty = fields.Float(string="Qté totale",       digits=(16, 3),          readonly=True)
    old_value = fields.Float(string="Valeur actuelle",  digits=(16, 2),          readonly=True)
    new_value = fields.Float(string="Nouvelle valeur",  digits=(16, 2),          readonly=True)

    state = fields.Selection([
        ('ready',     'À corriger'),
        ('not_found', 'Code non trouvé'),
        ('no_move',   'Aucune réception'),
        ('done',      'Corrigé'),
    ], string="Statut", readonly=True)
