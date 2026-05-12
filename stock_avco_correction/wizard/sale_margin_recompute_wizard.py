# -*- coding: utf-8 -*-
import base64
import io
import logging

import openpyxl

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class PosMarginCostWizard(models.TransientModel):
    _name = 'pos.margin.cost.wizard'
    _description = "Correction des coûts et marges POS — Import Excel"

    state = fields.Selection([
        ('import',  'Import fichier'),
        ('preview', 'Aperçu'),
        ('done',    'Terminé'),
    ], default='import', readonly=True)

    company_id = fields.Many2one(
        'res.company', string="Société", required=True,
        default=lambda self: self.env.company,
    )
    date_from = fields.Date(string="Date de début")
    date_to   = fields.Date(string="Date de fin")

    excel_file     = fields.Binary(string="Fichier Excel", attachment=False)
    excel_filename = fields.Char(string="Nom du fichier")

    line_ids = fields.One2many(
        'pos.margin.cost.wizard.line', 'wizard_id', string="Lignes"
    )

    count_ok        = fields.Integer("Articles à corriger",  compute='_compute_stats')
    count_not_found = fields.Integer("Codes non trouvés",    compute='_compute_stats')
    count_no_lines  = fields.Integer("Sans lignes POS",      compute='_compute_stats')
    total_lines_pos = fields.Integer("Lignes POS impactées", compute='_compute_stats')
    summary_html    = fields.Html("Résumé", readonly=True)

    @api.depends('line_ids.state', 'line_ids.nb_pos_lines')
    def _compute_stats(self):
        for rec in self:
            rec.count_ok        = len(rec.line_ids.filtered(lambda l: l.state == 'ok'))
            rec.count_not_found = len(rec.line_ids.filtered(lambda l: l.state == 'not_found'))
            rec.count_no_lines  = len(rec.line_ids.filtered(lambda l: l.state == 'no_lines'))
            rec.total_lines_pos = sum(rec.line_ids.filtered(lambda l: l.state == 'ok').mapped('nb_pos_lines'))

    # ── Étape 1 : Analyser ───────────────────────────────────────────────────

    def action_preview(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))

        self.line_ids.unlink()
        rows = self._parse_excel(self.excel_file)

        lines_vals = []
        for code, correct_cost in rows:
            vals = self._analyse_product(code, correct_cost)
            vals['wizard_id'] = self.id
            lines_vals.append(vals)

        if lines_vals:
            self.env['pos.margin.cost.wizard.line'].create(lines_vals)

        self.state = 'preview'
        return self._reload()

    def _pos_line_domain(self, product_id):
        domain = [
            ('product_id', '=', product_id),
            ('order_id.company_id', '=', self.company_id.id),
            ('order_id.state', 'not in', ['cancel']),
        ]
        if self.date_from:
            domain.append(('order_id.date_order', '>=', self.date_from))
        if self.date_to:
            domain.append(('order_id.date_order', '<=', self.date_to))
        return domain

    def _analyse_product(self, code, correct_cost):
        base = {'code_article': code, 'correct_unit_cost': correct_cost}

        product = self.env['product.product'].search(
            [('default_code', '=', code), ('active', '=', True)], limit=1
        )
        if not product:
            return {**base, 'state': 'not_found',
                    'message': f"Code '{code}' introuvable dans Odoo."}

        base['product_id'] = product.id

        if correct_cost <= 0:
            return {**base, 'state': 'not_found',
                    'message': "Le coût doit être supérieur à 0."}

        pos_lines = self.env['pos.order.line'].search(self._pos_line_domain(product.id))
        if not pos_lines:
            return {**base, 'state': 'no_lines',
                    'message': "Aucune ligne POS trouvée pour cet article."}

        total_qty     = sum(abs(l.qty) for l in pos_lines)
        old_total_cost = sum(l.total_cost for l in pos_lines)
        new_total_cost = sum(l.qty * correct_cost for l in pos_lines)

        return {
            **base,
            'nb_pos_lines':   len(pos_lines),
            'total_qty':      total_qty,
            'old_total_cost': old_total_cost,
            'new_total_cost': new_total_cost,
            'state':          'ok',
            'message':        '',
        }

    # ── Étape 2 : Appliquer ──────────────────────────────────────────────────

    def action_apply(self):
        self.ensure_one()
        ok_lines = self.line_ids.filtered(lambda l: l.state == 'ok')
        if not ok_lines:
            raise UserError(_("Aucun article valide à corriger."))

        nb_products = 0
        nb_pos_lines = 0
        errors = []

        for wl in ok_lines:
            try:
                pos_lines = self.env['pos.order.line'].search(
                    self._pos_line_domain(wl.product_id.id)
                )
                for line in pos_lines:
                    line.sudo().write({
                        'total_cost':            line.qty * wl.correct_unit_cost,
                        'is_total_cost_computed': True,
                    })
                    nb_pos_lines += 1

                nb_products += 1
                wl.state = 'done'
                _logger.info(
                    "[POS MARGIN] %s (%s) : coût unitaire %.4f → %d lignes POS corrigées",
                    wl.code_article, wl.product_id.display_name,
                    wl.correct_unit_cost, len(pos_lines),
                )
            except Exception as e:
                errors.append(f"{wl.code_article} : {e}")
                _logger.exception("Erreur correction coût POS %s", wl.code_article)

        self.summary_html = self._build_result_html(nb_products, nb_pos_lines, errors)
        self.state = 'done'
        return self._reload()

    def action_reset(self):
        self.line_ids.unlink()
        self.write({
            'state': 'import',
            'excel_file': False,
            'excel_filename': False,
            'summary_html': False,
        })
        return self._reload()

    # ── Parsing Excel ────────────────────────────────────────────────────────

    def _parse_excel(self, file_b64):
        """
        Lit le fichier Excel.
        Colonnes attendues : 'code article' (ou 'code') | 'cout' (ou 'pmp', 'prix', 'cost')
        Données à partir de la ligne 2.
        """
        data = base64.b64decode(file_b64)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("Impossible de lire le fichier Excel : %s") % str(e))

        ws = wb.active
        headers = {}
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            cell = col[0]
            if cell.value:
                h = str(cell.value).strip().lower()
                if any(k in h for k in ('code',)):
                    headers.setdefault('code', cell.column - 1)
                elif any(k in h for k in ('cout', 'coût', 'pmp', 'prix', 'cost', 'price')):
                    headers.setdefault('cost', cell.column - 1)

        if 'code' not in headers or 'cost' not in headers:
            found = [str(ws.cell(1, i + 1).value) for i in range(ws.max_column)
                     if ws.cell(1, i + 1).value]
            raise UserError(_(
                "Colonnes non trouvées dans le fichier.\n"
                "Attendu : une colonne 'code article' et une colonne 'cout'.\n"
                "Colonnes détectées : %s"
            ) % ', '.join(found))

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[headers['code']]
            cost = row[headers['cost']]
            if code is None:
                continue
            code = str(code).strip()
            if code.endswith('.0'):
                code = code[:-2]
            code = code.zfill(4)
            try:
                cost = float(cost) if cost is not None else 0.0
            except (ValueError, TypeError):
                cost = 0.0
            if code:
                rows.append((code, cost))

        if not rows:
            raise UserError(_(
                "Aucune donnée trouvée dans le fichier.\n"
                "Format attendu :\n"
                "  Colonne A : code article\n"
                "  Colonne B : cout unitaire\n"
                "  Données à partir de la ligne 2."
            ))
        return rows

    # ── HTML ─────────────────────────────────────────────────────────────────

    def _build_result_html(self, nb_products, nb_pos_lines, errors):
        html = f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <h3 style="color:#28a745;border-bottom:2px solid #28a745;padding-bottom:8px;">
            Correction des coûts POS terminée
          </h3>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Articles traités</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_products}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;">Lignes POS corrigées</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_pos_lines}</td>
            </tr>
          </table>
          <p style="margin-top:12px;color:#6c757d;font-size:12px;">
            Le champ <b>Coût total</b> (<code>total_cost</code>) a été mis à jour sur chaque
            ligne POS. Les marges sont recalculées automatiquement à l'affichage.
          </p>
        """
        if errors:
            items = ''.join(f'<li>{e}</li>' for e in errors)
            html += f"""
          <p style="color:#dc3545;font-weight:bold;margin-top:10px;">
            ❌ {len(errors)} erreur(s) :
          </p>
          <ul style="color:#dc3545;">{items}</ul>
            """
        html += "</div>"
        return html

    def _reload(self):
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }


class PosMarginCostWizardLine(models.TransientModel):
    _name = 'pos.margin.cost.wizard.line'
    _description = "Ligne de correction coût POS"
    _order = 'state, code_article'

    wizard_id         = fields.Many2one('pos.margin.cost.wizard', ondelete='cascade')
    code_article      = fields.Char(string="Code Article",    readonly=True)
    product_id        = fields.Many2one('product.product',    string="Article",     readonly=True)
    correct_unit_cost = fields.Float(string="Coût unitaire Excel", digits=(16, 4), readonly=True)

    nb_pos_lines   = fields.Integer(string="Nb lignes POS",      readonly=True)
    total_qty      = fields.Float(string="Qté totale",           digits=(16, 3), readonly=True)
    old_total_cost = fields.Float(string="Coût total actuel",    digits=(16, 2), readonly=True)
    new_total_cost = fields.Float(string="Nouveau coût total",   digits=(16, 2), readonly=True)

    state = fields.Selection([
        ('ok',        'À corriger'),
        ('not_found', 'Code non trouvé'),
        ('no_lines',  'Aucune ligne POS'),
        ('done',      'Corrigé'),
    ], string="Statut", readonly=True)
    message = fields.Char(string="Message", readonly=True)
