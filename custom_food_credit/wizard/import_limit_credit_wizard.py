# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook
import logging

_logger = logging.getLogger(__name__)


class ImportLimitCreditWizard(models.TransientModel):
    _name = 'import.limit.credit.wizard'
    _description = 'Import Excel - Mise à jour Limite Crédit Consommée'

    file = fields.Binary(string="Fichier Excel", required=True)
    file_name = fields.Char()
    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('loaded', 'Chargé'),
        ('done', 'Terminé'),
    ], default='draft')
    line_ids = fields.One2many(
        'import.limit.credit.line',
        'wizard_id',
        string='Lignes'
    )
    count_ok = fields.Integer(string='Lignes valides', compute='_compute_counts')
    count_errors = fields.Integer(string='Lignes en erreur', compute='_compute_counts')

    @api.depends('line_ids.status')
    def _compute_counts(self):
        for wizard in self:
            wizard.count_ok = len(wizard.line_ids.filtered(lambda l: l.status == 'ok'))
            wizard.count_errors = len(wizard.line_ids.filtered(lambda l: l.status != 'ok'))

    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez charger un fichier Excel."))

        # Supprimer les anciennes lignes
        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        file_data = io.BytesIO(decoded_file)

        try:
            workbook = load_workbook(file_data)
        except Exception:
            raise UserError(_("Impossible de lire le fichier. Vérifiez qu'il s'agit bien d'un fichier Excel (.xlsx)."))

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        # Lecture des en-têtes (première ligne)
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]

        required_columns = ['partner_id', 'Limite Credit Consommée']
        for col in required_columns:
            if col not in headers:
                raise UserError(
                    _("Colonne manquante : '%s'\nColonnes trouvées : %s") % (col, ', '.join(headers))
                )

        idx_partner = headers.index('partner_id')
        idx_consumed = headers.index('Limite Credit Consommée')

        lines_vals = []

        for row_num, row in enumerate(rows[1:], start=2):
            # Ignorer les lignes vides
            if not row or (row[idx_partner] is None and row[idx_consumed] is None):
                continue

            partner_name = str(row[idx_partner]).strip() if row[idx_partner] else ''
            if not partner_name:
                continue

            try:
                amount_new = float(row[idx_consumed] or 0.0)
            except (ValueError, TypeError):
                amount_new = 0.0

            # Recherche du partenaire (nom exact, insensible à la casse)
            partner = self.env['res.partner'].search(
                [('customer_id', '=ilike', partner_name)],
                limit=1
            )

            if not partner:
                lines_vals.append((0, 0, {
                    'partner_name': partner_name,
                    'partner_id': False,
                    'limit_credit_id': False,
                    'amount_limit_consumed_new': amount_new,
                    'amount_limit_consumed_old': 0.0,
                    'status': 'partner_not_found',
                }))
                continue

            # Recherche de la limite de crédit pour ce partenaire
            limit_credit = self.env['limit.credit'].search(
                [('partner_id', '=', partner.id)],
                limit=1
            )

            if not limit_credit:
                lines_vals.append((0, 0, {
                    'partner_name': partner_name,
                    'partner_id': partner.id,
                    'limit_credit_id': False,
                    'amount_limit_consumed_new': amount_new,
                    'amount_limit_consumed_old': 0.0,
                    'status': 'no_limit_credit',
                }))
                continue

            # Tout est OK
            lines_vals.append((0, 0, {
                'partner_name': partner_name,
                'partner_id': partner.id,
                'limit_credit_id': limit_credit.id,
                'amount_limit_consumed_new': amount_new,
                'amount_limit_consumed_old': limit_credit.amount_limit_consumed,
                'status': 'ok',
            }))

        if not lines_vals:
            raise UserError(_("Aucune ligne de données trouvée dans le fichier (hors en-têtes)."))

        self.write({
            'line_ids': lines_vals,
            'state': 'loaded',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_confirm(self):
        self.ensure_one()

        valid_lines = self.line_ids.filtered(lambda l: l.status == 'ok')

        if not valid_lines:
            raise UserError(_("Aucune ligne valide à traiter. Vérifiez les erreurs dans le tableau."))

        for line in valid_lines:
            line.limit_credit_id.write({
                'amount_limit_consumed': line.amount_limit_consumed_new,
            })
            # Traçabilité via limit.credit.operation
            self.env['limit.credit.operation'].create({
                'limit_id': line.limit_credit_id.id,
                'name': _("Import Excel - Mise à jour consommation"),
                'amount_operation': line.amount_limit_consumed_new - line.amount_limit_consumed_old,
                'operation_date': fields.Datetime.now(),
            })

        self.write({'state': 'done'})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Import terminé"),
                'message': _("%d limite(s) de crédit mise(s) à jour avec succès.") % len(valid_lines),
                'type': 'success',
                'sticky': False,
            },
        }


class ImportLimitCreditLine(models.TransientModel):
    _name = 'import.limit.credit.line'
    _description = 'Ligne import limite crédit'

    wizard_id = fields.Many2one('import.limit.credit.wizard')
    partner_name = fields.Char(string='Nom Client (Excel)', readonly=True)
    partner_id = fields.Many2one('res.partner', string='Client trouvé', readonly=True)
    limit_credit_id = fields.Many2one('limit.credit', string='Limite Crédit', readonly=True)
    amount_limit_consumed_old = fields.Float(string='Consommé actuel', readonly=True)
    amount_limit_consumed_new = fields.Float(string='Nouveau montant', readonly=True)
    status = fields.Selection([
        ('ok', 'Prêt'),
        ('partner_not_found', 'Client non trouvé'),
        ('no_limit_credit', 'Pas de limite crédit'),
    ], string='Statut', readonly=True)
