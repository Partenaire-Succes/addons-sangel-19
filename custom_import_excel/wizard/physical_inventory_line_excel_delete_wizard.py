# -*- coding: utf-8 -*-
import base64
import io
import logging

import openpyxl

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class PhysicalInventoryLineExcelDeleteWizard(models.TransientModel):
    _name = 'physical.inventory.line.excel.delete.wizard'
    _description = "Suppression de lignes d'inventaire par import Excel"

    state = fields.Selection([
        ('import',  'Import fichier'),
        ('preview', 'Aperçu'),
        ('done',    'Terminé'),
    ], default='import', readonly=True)

    company_id = fields.Many2one(
        'res.company', string="Société", required=True,
        default=lambda self: self.env.company,
    )

    excel_file     = fields.Binary(string="Fichier Excel", attachment=False)
    excel_filename = fields.Char(string="Nom du fichier")

    line_ids = fields.One2many(
        'physical.inventory.line.excel.delete.wizard.line', 'wizard_id', string="Lignes"
    )

    count_ok          = fields.Integer("Lignes à supprimer",       compute='_compute_stats')
    count_not_found    = fields.Integer("Produit/inventaire introuvable", compute='_compute_stats')
    count_no_lines     = fields.Integer("Aucune ligne correspondante",    compute='_compute_stats')
    total_inv_lines    = fields.Integer("Total lignes d'inventaire",      compute='_compute_stats')
    summary_html       = fields.Html("Résumé", readonly=True)

    @api.depends('line_ids.state', 'line_ids.nb_lines')
    def _compute_stats(self):
        for rec in self:
            rec.count_ok       = len(rec.line_ids.filtered(lambda l: l.state == 'ok'))
            rec.count_not_found = len(rec.line_ids.filtered(lambda l: l.state == 'not_found'))
            rec.count_no_lines  = len(rec.line_ids.filtered(lambda l: l.state == 'no_lines'))
            rec.total_inv_lines = sum(rec.line_ids.filtered(lambda l: l.state == 'ok').mapped('nb_lines'))

    # ------------------------------------------------------------------
    # Étape 1 — Analyser
    # ------------------------------------------------------------------

    def action_preview(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))

        self.line_ids.unlink()
        rows = self._parse_excel(self.excel_file)

        lines_vals = []
        for code, inventaire_name in rows:
            vals = self._analyse_row(code, inventaire_name)
            vals['wizard_id'] = self.id
            lines_vals.append(vals)

        if lines_vals:
            self.env['physical.inventory.line.excel.delete.wizard.line'].create(lines_vals)

        self.state = 'preview'
        return self._reload()

    def _analyse_row(self, code, inventaire_name):
        base = {'code_excel': code, 'inventaire_excel': inventaire_name}

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id],
        ))

        # physical.inventory.line se rattache par product_tmpl_id (pas par
        # variante) : on cherche directement sur product.template, dont le
        # default_code est le champ code_article utilisé partout ailleurs
        # dans ce module pour l'inventaire (custom_stock/product_template.py
        # : code_article = related='default_code').
        product_tmpl = env['product.template'].with_context(active_test=False).search(
            [('default_code', '=', code)], limit=1
        )
        if not product_tmpl:
            return {**base, 'state': 'not_found',
                    'message': f"Code article '{code}' introuvable dans Odoo."}
        base['product_tmpl_id'] = product_tmpl.id

        inventory = env['physical.inventory'].search(
            [('name', '=', inventaire_name), ('company_id', '=', self.company_id.id)], limit=1
        )
        if not inventory:
            return {**base, 'state': 'not_found',
                    'message': f"Inventaire '{inventaire_name}' introuvable pour la société "
                               f"'{self.company_id.name}'."}
        base['inventory_physical_id'] = inventory.id

        inv_lines = self.env['physical.inventory.line'].with_context(active_test=False).search([
            ('product_tmpl_id', '=', product_tmpl.id),
            ('inventory_physical_id', '=', inventory.id),
        ])
        if not inv_lines:
            return {**base, 'state': 'no_lines',
                    'message': "Aucune ligne d'inventaire trouvée pour cet article dans cet inventaire."}

        return {
            **base,
            'inv_line_ids': [(6, 0, inv_lines.ids)],
            'nb_lines':     len(inv_lines),
            'state':        'ok',
            'message':      '',
        }

    # ------------------------------------------------------------------
    # Étape 2 — Supprimer
    # ------------------------------------------------------------------

    def action_delete(self):
        self.ensure_one()
        ok_lines = self.line_ids.filtered(lambda l: l.state == 'ok')
        if not ok_lines:
            raise UserError(_("Aucune ligne valide à supprimer."))

        nb_articles = 0
        nb_inv_lines = 0
        errors = []

        for wl in ok_lines:
            try:
                # physical.inventory.line.unlink() refuse la suppression dès
                # que l'inventaire parent est en mode 'normal' (le mode par
                # défaut), sauf sous ce contexte — même contournement que
                # PhysicalInventory.create_line_physical(). Ici la
                # suppression est déjà validée explicitement par l'utilisateur
                # (aperçu + confirmation), donc le contournement est légitime.
                inv_lines = wl.inv_line_ids.with_context(from_generate_lines=True)
                count = len(inv_lines)
                inv_lines.unlink()
                nb_inv_lines += count
                nb_articles += 1
                wl.state = 'done'
                _logger.info(
                    "[DELETE PHYSICAL INVENTORY LINE] %s (%s) — %d ligne(s) supprimée(s)",
                    wl.code_excel, wl.inventaire_excel, count,
                )
            except Exception as e:
                errors.append(f"{wl.code_excel} / {wl.inventaire_excel} : {e}")
                _logger.exception(
                    "Erreur suppression physical.inventory.line %s / %s",
                    wl.code_excel, wl.inventaire_excel,
                )

        self.summary_html = self._build_result_html(nb_articles, nb_inv_lines, errors)
        self.state = 'done'
        return self._reload()

    def action_reset(self):
        self.line_ids.unlink()
        self.write({
            'state': 'import',
            'excel_file': False,
            'excel_filename': False,
            'summary_html': False,
        })
        return self._reload()

    # ------------------------------------------------------------------
    # Parsing Excel
    # ------------------------------------------------------------------

    @staticmethod
    def _clean_code(value):
        """Nettoie le code article lu depuis Excel (int, float ou str avec
        zéros non significatifs) et le complète à 4 chiffres, comme les
        autres imports du module (cf. pos.margin.cost.wizard)."""
        if value is None:
            return False
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        code = str(value).strip()
        if code.endswith('.0'):
            code = code[:-2]
        return code.zfill(4) if code else False

    def _parse_excel(self, file_b64):
        """Colonnes attendues : 'CODE' (code article) | 'INVENTAIRE' (nom
        exact de l'inventaire physique parent). Données à partir de la
        ligne 2."""
        data = base64.b64decode(file_b64)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("Impossible de lire le fichier Excel : %s") % str(e))

        ws = wb.active
        headers = {}
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            cell = col[0]
            if cell.value:
                h = str(cell.value).strip().lower()
                if 'code' in h:
                    headers.setdefault('code', cell.column - 1)
                elif 'inventaire' in h or 'inventory' in h:
                    headers.setdefault('inventaire', cell.column - 1)

        if 'code' not in headers or 'inventaire' not in headers:
            found = [str(ws.cell(1, i + 1).value) for i in range(ws.max_column)
                     if ws.cell(1, i + 1).value]
            raise UserError(_(
                "Colonnes non trouvées dans le fichier.\n"
                "Attendu : une colonne 'CODE' et une colonne 'INVENTAIRE'.\n"
                "Colonnes détectées : %s"
            ) % ', '.join(found))

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = self._clean_code(row[headers['code']])
            inventaire = row[headers['inventaire']]
            if not code or inventaire is None:
                continue
            inventaire = str(inventaire).strip()
            if code and inventaire:
                rows.append((code, inventaire))

        if not rows:
            raise UserError(_(
                "Aucune donnée trouvée dans le fichier.\n"
                "Format attendu :\n"
                "  Colonne CODE : code article\n"
                "  Colonne INVENTAIRE : nom exact de l'inventaire physique\n"
                "  Données à partir de la ligne 2."
            ))
        return rows

    # ------------------------------------------------------------------
    # HTML
    # ------------------------------------------------------------------

    def _build_result_html(self, nb_articles, nb_inv_lines, errors):
        html = f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <h3 style="color:#28a745;border-bottom:2px solid #28a745;padding-bottom:8px;">
            Suppression terminée
          </h3>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Articles traités</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_articles}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;">Lignes d'inventaire supprimées</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{nb_inv_lines}</td>
            </tr>
          </table>
        """
        if errors:
            items = ''.join(f'<li>{e}</li>' for e in errors)
            html += f"""
          <p style="color:#dc3545;font-weight:bold;margin-top:10px;">
            ❌ {len(errors)} erreur(s) :
          </p>
          <ul style="color:#dc3545;">{items}</ul>
            """
        html += "</div>"
        return html

    def _reload(self):
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }


class PhysicalInventoryLineExcelDeleteWizardLine(models.TransientModel):
    _name = 'physical.inventory.line.excel.delete.wizard.line'
    _description = "Ligne de suppression de ligne d'inventaire (import Excel)"
    _order = 'state, code_excel'

    wizard_id = fields.Many2one(
        'physical.inventory.line.excel.delete.wizard', ondelete='cascade'
    )

    code_excel       = fields.Char(string="Code Article (Excel)", readonly=True)
    inventaire_excel = fields.Char(string="Inventaire (Excel)",   readonly=True)

    product_tmpl_id       = fields.Many2one('product.template', string="Article",   readonly=True)
    inventory_physical_id = fields.Many2one('physical.inventory', string="Inventaire", readonly=True)
    inv_line_ids = fields.Many2many(
        'physical.inventory.line',
        relation='pil_excel_delete_wizard_line_rel',
        column1='wizard_line_id', column2='inv_line_id',
        string="Lignes d'inventaire concernées", readonly=True,
    )
    nb_lines = fields.Integer(string="Nb lignes", readonly=True)

    state = fields.Selection([
        ('ok',        'À supprimer'),
        ('not_found', 'Introuvable'),
        ('no_lines',  'Aucune ligne'),
        ('done',      'Supprimée'),
    ], string="Statut", readonly=True)
    message = fields.Char(string="Message", readonly=True)
