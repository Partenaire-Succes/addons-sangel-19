import base64
import io
import logging

import openpyxl
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class ProductStatusImportWizard(models.TransientModel):
    _name = 'product.status.import.wizard'
    _description = "Import Excel des statuts magasin"

    state = fields.Selection([
        ('import',  'Import fichier'),
        ('preview', 'Vérification'),
        ('done',    'Terminé'),
    ], default='import', string="Étape")

    company_id = fields.Many2one(
        'res.company', string="Magasin", required=True,
        default=lambda self: self.env.company,
    )

    excel_file     = fields.Binary(string="Fichier Excel", attachment=False)
    excel_filename = fields.Char(string="Nom du fichier")

    line_ids = fields.One2many(
        'product.status.import.wizard.line', 'wizard_id', string="Lignes"
    )

    summary_html      = fields.Html(string="Résumé", readonly=True)
    nb_updated        = fields.Integer(string="Statuts mis à jour",   readonly=True)
    nb_created        = fields.Integer(string="Statuts créés",        readonly=True)
    nb_not_found      = fields.Integer(string="Articles non trouvés", readonly=True)
    nb_unknown_status = fields.Integer(string="Statuts inconnus",     readonly=True)

    # ------------------------------------------------------------------
    # ÉTAPE 1 — Charger et analyser
    # ------------------------------------------------------------------
    def action_load_file(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))

        rows = self._parse_excel(self.excel_file, self.excel_filename)
        if not rows:
            raise UserError(_(
                "Fichier vide ou format incorrect.\n"
                "Colonnes attendues : 'code article' et 'statut'."
            ))

        self.line_ids.unlink()

        all_codes   = [r['code']   for r in rows if r['code']]
        all_statuts = list({r['statut'] for r in rows if r['statut']})

        products = self.env['product.product'].search(
            [('default_code', 'in', all_codes), ('active', '=', True)]
        )
        product_map = {p.default_code: p for p in products}

        statuses = self.env['product.status'].search([('code', 'in', all_statuts)])
        status_map = {s.code: s.id for s in statuses}

        # Récupère tous les statuts actuels pour ce magasin en une seule requête
        tmpl_ids = products.mapped('product_tmpl_id').ids
        existing_statuts = self.env['product.company.status'].search([
            ('product_id', 'in', tmpl_ids),
            ('company_id', '=', self.company_id.id),
        ])
        current_status_by_tmpl = {e.product_id.id: e.status_id.id for e in existing_statuts}

        lines_vals = []
        for row in rows:
            code   = row['code']
            statut = row['statut']

            if not code:
                continue

            product = product_map.get(code)
            status_id = status_map.get(statut) if statut else False

            if not product:
                line_state = 'not_found'
            elif not status_id:
                line_state = 'unknown_status'
            else:
                line_state = 'ready'

            tmpl_id = product.product_tmpl_id.id if product else False
            current_status_id = current_status_by_tmpl.get(tmpl_id, False) if tmpl_id else False

            lines_vals.append({
                'wizard_id':         self.id,
                'code_article':      code,
                'code_statut':       statut or '',
                'product_id':        product.id if product else False,
                'status_id':         status_id or False,
                'current_status_id': current_status_id,
                'state':             line_state,
            })

        self.env['product.status.import.wizard.line'].create(lines_vals)

        nb_ready          = sum(1 for v in lines_vals if v['state'] == 'ready')
        nb_not_found      = sum(1 for v in lines_vals if v['state'] == 'not_found')
        nb_unknown_status = sum(1 for v in lines_vals if v['state'] == 'unknown_status')

        self.summary_html = self._build_load_summary_html(
            len(rows), nb_ready, nb_not_found, nb_unknown_status
        )
        self.state = 'preview'
        return self._reload()

    # ------------------------------------------------------------------
    # ÉTAPE 2 — Appliquer
    # ------------------------------------------------------------------
    def action_apply(self):
        self.ensure_one()

        lines_to_apply = self.line_ids.filtered(lambda l: l.state == 'ready')
        if not lines_to_apply:
            raise UserError(_("Aucune ligne en état 'Prêt' à traiter."))

        nb_updated = 0
        nb_created = 0

        tmpl_ids = lines_to_apply.mapped(lambda l: l.product_id.product_tmpl_id.id)
        existing_map = {
            e.product_id.id: e
            for e in self.env['product.company.status'].search([
                ('product_id', 'in', tmpl_ids),
                ('company_id', '=', self.company_id.id),
            ])
        }

        for line in lines_to_apply:
            tmpl_id  = line.product_id.product_tmpl_id.id
            existing = existing_map.get(tmpl_id)

            if existing:
                if existing.status_id.id != line.status_id.id:
                    existing.write({'status_id': line.status_id.id})
                    nb_updated += 1
                    _logger.info(
                        "Statut [%s] %s : %s → %s",
                        self.company_id.name,
                        line.product_id.display_name,
                        line.current_status_id.code,
                        line.status_id.code,
                    )
            else:
                self.env['product.company.status'].create({
                    'product_id': tmpl_id,
                    'company_id': self.company_id.id,
                    'status_id':  line.status_id.id,
                })
                nb_created += 1
                _logger.info(
                    "Statut créé [%s] %s : %s",
                    self.company_id.name,
                    line.product_id.display_name,
                    line.status_id.code,
                )

            line.write({'state': 'done'})

        nb_not_found      = len(self.line_ids.filtered(lambda l: l.state == 'not_found'))
        nb_unknown_status = len(self.line_ids.filtered(lambda l: l.state == 'unknown_status'))

        self.write({
            'state':            'done',
            'nb_updated':        nb_updated,
            'nb_created':        nb_created,
            'nb_not_found':      nb_not_found,
            'nb_unknown_status': nb_unknown_status,
        })
        self.summary_html = self._build_apply_summary_html(
            nb_updated, nb_created, nb_not_found, nb_unknown_status
        )
        _logger.info(
            "Import statuts [%s] : %d mis à jour | %d créés | %d non trouvés | %d statuts inconnus",
            self.company_id.name, nb_updated, nb_created, nb_not_found, nb_unknown_status,
        )
        return self._reload()

    def action_reset(self):
        self.line_ids.unlink()
        self.write({
            'state': 'import',
            'excel_file': False,
            'excel_filename': False,
            'summary_html': False,
            'nb_updated': 0,
            'nb_created': 0,
            'nb_not_found': 0,
            'nb_unknown_status': 0,
        })
        return self._reload()

    # ------------------------------------------------------------------
    # PARSING — colonnes : 'code article' + 'statut'
    # ------------------------------------------------------------------
    def _parse_excel(self, file_b64, filename):
        file_bytes = base64.b64decode(file_b64)
        rows = []

        CODE_KEYS   = ('code', 'article', 'reference', 'réf')
        STATUT_KEYS = ('statut', 'status', 'etat', 'état')

        def detect_cols(headers):
            ci, si = None, None
            for i, h in enumerate(headers):
                h = str(h).strip().lower() if h else ''
                if any(k in h for k in CODE_KEYS) and ci is None:
                    ci = i
                elif any(k in h for k in STATUT_KEYS) and si is None:
                    si = i
            return ci, si

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            ci, si = detect_cols(header_row)

            if ci is None or si is None:
                found = [str(h) for h in header_row if h]
                raise UserError(_(
                    "Colonnes 'code article' et 'statut' non trouvées.\n"
                    "Colonnes détectées : %s"
                ) % ', '.join(found))

            for row in ws.iter_rows(min_row=2, values_only=True):
                code   = str(row[ci]).strip().zfill(4) if row[ci] is not None else None
                statut = str(row[si]).strip().upper()  if row[si] is not None else None
                if code and code != '0000':
                    rows.append({'code': code, 'statut': statut})

        except UserError:
            raise
        except Exception as e:
            if HAS_XLRD:
                try:
                    wb  = xlrd.open_workbook(file_contents=file_bytes)
                    ws  = wb.sheet_by_index(0)
                    hdr = [ws.cell_value(0, c) for c in range(ws.ncols)]
                    ci, si = detect_cols(hdr)
                    if ci is None or si is None:
                        raise UserError(_("Colonnes 'code article' et 'statut' non trouvées."))
                    for r in range(1, ws.nrows):
                        code   = str(ws.cell_value(r, ci)).strip().zfill(4) if ws.cell_value(r, ci) else None
                        statut = str(ws.cell_value(r, si)).strip().upper()  if ws.cell_value(r, si) else None
                        if code and code != '0000':
                            rows.append({'code': code, 'statut': statut})
                except UserError:
                    raise
                except Exception as e2:
                    raise UserError(_("Impossible de lire le fichier : %s") % str(e2))
            else:
                raise UserError(_("Impossible de lire le fichier : %s") % str(e))

        return rows

    # ------------------------------------------------------------------
    # HTML
    # ------------------------------------------------------------------
    def _build_load_summary_html(self, total, nb_ready, nb_not_found, nb_unknown_status):
        return f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <div style="background:#e8f4fd;border:1px solid #bee5eb;border-radius:6px;
                      padding:8px 12px;margin-bottom:10px;font-size:13px;">
            <b>Magasin :</b> {self.company_id.name}
          </div>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;">Total lignes dans le fichier</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;">{total}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#28a745;"><b>Prêts à importer</b></td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#28a745;">{nb_ready}</td>
            </tr>
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;color:#dc3545;">Articles non trouvés</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#dc3545;">{nb_not_found}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#e67e22;">Code statut inconnu</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#e67e22;">{nb_unknown_status}</td>
            </tr>
          </table>
          <p style="margin-top:10px;color:#555;font-size:12px;">
            Vérifiez les lignes ci-dessous puis cliquez sur <b>Appliquer les statuts</b>.
          </p>
        </div>
        """

    def _build_apply_summary_html(self, nb_updated, nb_created, nb_not_found, nb_unknown_status):
        return f"""
        <div style="font-family:Arial,sans-serif;padding:10px;">
          <div style="background:#e8f4fd;border:1px solid #bee5eb;border-radius:6px;
                      padding:8px 12px;margin-bottom:10px;font-size:13px;">
            <b>Magasin :</b> {self.company_id.name}
          </div>
          <h3 style="border-bottom:2px solid #28a745;padding-bottom:8px;color:#28a745;">
            Import des statuts terminé !
          </h3>
          <table style="width:100%;border-collapse:collapse;">
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;color:#28a745;">Statuts mis à jour</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#28a745;">{nb_updated}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#17a2b8;">Nouveaux statuts créés</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#17a2b8;">{nb_created}</td>
            </tr>
            <tr style="background:#f8f9fa;">
              <td style="padding:8px;border:1px solid #dee2e6;color:#dc3545;">Articles non trouvés (ignorés)</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#dc3545;">{nb_not_found}</td>
            </tr>
            <tr>
              <td style="padding:8px;border:1px solid #dee2e6;color:#e67e22;">Statuts inconnus (ignorés)</td>
              <td style="padding:8px;border:1px solid #dee2e6;font-weight:bold;color:#e67e22;">{nb_unknown_status}</td>
            </tr>
          </table>
        </div>
        """

    def _reload(self):
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }


class ProductStatusImportWizardLine(models.TransientModel):
    _name = 'product.status.import.wizard.line'
    _description = "Ligne import statut magasin"
    _order = 'state, code_article'

    wizard_id         = fields.Many2one('product.status.import.wizard', ondelete='cascade')
    code_article      = fields.Char(string="Code Article", readonly=True)
    code_statut       = fields.Char(string="Code Statut",  readonly=True)

    product_id        = fields.Many2one('product.product', string="Article",        readonly=True)
    status_id         = fields.Many2one('product.status',  string="Nouveau statut", readonly=True)
    current_status_id = fields.Many2one('product.status',  string="Statut actuel",  readonly=True)

    state = fields.Selection([
        ('ready',          'Prêt'),
        ('not_found',      'Article non trouvé'),
        ('unknown_status', 'Statut inconnu'),
        ('done',           'Mis à jour'),
    ], string="État", readonly=True)
