# -*- coding: utf-8 -*-
from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook
import logging

_logger = logging.getLogger(__name__)


class SupplierPriceImportWizard(models.TransientModel):
    _name = "supplier.price.import.wizard"
    _description = "Import Liste Prix Fournisseurs"

    file = fields.Binary(string="Fichier Excel")
    file_name = fields.Char()

    company_id = fields.Many2one(
        "res.company",
        string="Société",
        required=True,
        default=lambda self: self.env.company,
    )

    line_ids = fields.One2many(
        "supplier.price.import.line",
        "wizard_id",
        string="Lignes",
    )

    state = fields.Selection([
        ("draft", "Brouillon"),
        ("loaded", "Chargé"),
        ("done", "Terminé"),
    ], default="draft")

    @staticmethod
    def _clean_val(value):
        """Nettoie une valeur Excel (supprime les .0 parasites, strip)."""
        if value is None:
            return False
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    def action_load_file(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez télécharger un fichier Excel."))

        self.line_ids.unlink()

        decoded_file = base64.b64decode(self.file)
        workbook = load_workbook(io.BytesIO(decoded_file))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        headers = [str(h).strip() for h in rows[0]]
        required_columns = ["partner_id", "product_tmpl_id", "company_id", "price"]
        for col in required_columns:
            if col not in headers:
                raise UserError(_("Colonne manquante dans le fichier : %s") % col)

        partner_idx = headers.index("partner_id")
        product_idx = headers.index("product_tmpl_id")
        company_idx = headers.index("company_id")
        price_idx = headers.index("price")

        env = self.env(context=dict(self.env.context, active_test=False))
        lines_vals = []

        for row in rows[1:]:
            if not row:
                continue

            partner_val = self._clean_val(row[partner_idx])
            product_val = self._clean_val(row[product_idx])
            company_val = self._clean_val(row[company_idx])
            price_raw = row[price_idx]

            if not partner_val and not product_val and not company_val:
                continue

            try:
                price = float(price_raw) if price_raw is not None else 0.0
            except (ValueError, TypeError):
                price = 0.0

            # --- Résolution fournisseur ---
            partner = env["res.partner"].search(
                [("name", "=", partner_val), ("supplier_rank", ">", 0)], limit=1
            )
            if not partner:
                partner = env["res.partner"].search([("name", "=", partner_val)], limit=1)

            # --- Résolution produit (par code interne) ---
            product_tmpl = env["product.template"].search(
                [("default_code", "=", product_val)], limit=1
            )

            # --- Résolution société ---
            company = env["res.company"].search([("name", "=", company_val)], limit=1)

            # --- Vérification existence ligne fournisseur ---
            supplierinfo = False
            action = "not_found"

            if partner and product_tmpl and company:
                supplierinfo = env["product.supplierinfo"].search([
                    ("partner_id", "=", partner.id),
                    ("product_tmpl_id", "=", product_tmpl.id),
                    ("company_id", "=", company.id),
                ], limit=1)
                action = "update" if supplierinfo else "create"

            lines_vals.append((0, 0, {
                "partner_val": partner_val,
                "product_val": product_val,
                "company_val": company_val,
                "price": price,
                "partner_id": partner.id if partner else False,
                "product_tmpl_id": product_tmpl.id if product_tmpl else False,
                "company_res_id": company.id if company else False,
                "supplierinfo_id": supplierinfo.id if supplierinfo else False,
                "action": action,
            }))

        self.write({
            "line_ids": lines_vals,
            "state": "loaded",
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_confirm(self):
        self.ensure_one()
        created = 0
        updated = 0

        for line in self.line_ids.filtered(lambda l: l.action in ("create", "update")):
            company = line.company_res_id or self.company_id
            env = self.env(context=dict(
                self.env.context,
                allowed_company_ids=[company.id],
            ))

            if line.action == "update" and line.supplierinfo_id:
                line.supplierinfo_id.with_company(company).price = line.price
                updated += 1

            elif line.action == "create":
                env["product.supplierinfo"].with_company(company).create({
                    "partner_id": line.partner_id.id,
                    "product_tmpl_id": line.product_tmpl_id.id,
                    "company_id": company.id,
                    "price": line.price,
                    "min_qty": 0.0,
                    "currency_id": company.currency_id.id,
                })
                created += 1

        self.state = "done"
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import terminé"),
                "message": _("%d ligne(s) créée(s)  •  %d ligne(s) mise(s) à jour") % (created, updated),
                "type": "success",
                "sticky": True,
            },
        }


class SupplierPriceImportLine(models.TransientModel):
    _name = "supplier.price.import.line"
    _description = "Ligne Import Prix Fournisseur"

    wizard_id = fields.Many2one("supplier.price.import.wizard")

    # Valeurs brutes lues depuis l'Excel
    partner_val = fields.Char("Fournisseur (Excel)", readonly=True)
    product_val = fields.Char("Code Produit (Excel)", readonly=True)
    company_val = fields.Char("Société (Excel)", readonly=True)
    price = fields.Float("Prix unitaire")

    # Enregistrements Odoo résolus
    partner_id = fields.Many2one("res.partner", string="Fournisseur trouvé", readonly=True)
    product_tmpl_id = fields.Many2one("product.template", string="Produit trouvé", readonly=True)
    company_res_id = fields.Many2one("res.company", string="Société trouvée", readonly=True)
    supplierinfo_id = fields.Many2one("product.supplierinfo", string="Prix existant", readonly=True)

    action = fields.Selection([
        ("create", "Créer"),
        ("update", "Mettre à jour"),
        ("not_found", "Non trouvé"),
    ], string="Action", readonly=True)
