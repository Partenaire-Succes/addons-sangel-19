# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from openpyxl import load_workbook


class OrderpointImportWizard(models.TransientModel):
    _name = 'orderpoint.import.wizard'
    _description = "Import des Règles de Réassort (Min/Max)"

    file = fields.Binary(string="Fichier Excel", required=True)
    file_name = fields.Char()

    company_id = fields.Many2one(
        'res.company',
        required=True,
        default=lambda self: self.env.company,
    )
    warehouse_id = fields.Many2one(
        'stock.warehouse',
        string='Entrepôt',
        required=True,
        default=lambda self: self.env['stock.warehouse'].search(
            [('company_id', '=', self.env.company.id)], limit=1
        ),
    )
    location_id = fields.Many2one(
        'stock.location',
        string='Emplacement (création uniquement)',
        domain="[('usage', '=', 'internal')]",
        help="Requis uniquement si des nouvelles règles doivent être créées. "
             "Laissez vide si toutes les règles existent déjà.",
    )

    @api.onchange('warehouse_id')
    def _onchange_warehouse_id(self):
        if self.warehouse_id and self.warehouse_id.lot_stock_id:
            self.location_id = self.warehouse_id.lot_stock_id

    state = fields.Selection([
        ('draft',  'Brouillon'),
        ('loaded', 'Chargé'),
        ('done',   'Terminé'),
    ], default='draft')

    line_ids = fields.One2many(
        'orderpoint.import.wizard.line',
        'wizard_id',
        string='Lignes',
    )

    # Noms de colonnes acceptés (insensible à la casse, accents inclus)
    _COL_ALIASES = {
        'product_code': [
            'product_code', 'code article', 'code_article',
            'code', 'reference', 'ref', 'code art', 'codearticle',
        ],
        'qty_min': [
            'qty_min', 'quantité min', 'quantite min', 'qte min',
            'qty min', 'stock min', 'min', 'quantite_min',
            'quantité minimum', 'quantite minimum',
        ],
        'qty_max': [
            'qty_max', 'quantité max', 'quantite max', 'qte max',
            'qty max', 'stock max', 'max', 'quantite_max',
            'quantité maximum', 'quantite maximum',
        ],
    }

    @staticmethod
    def _to_float(value):
        """Convertit une cellule Excel en float — gère None, formules non calculées, etc."""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _clean_code(value):
        if value is None:
            return False
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        return str(value).strip()

    @staticmethod
    def _normalize(text):
        """Minuscules + supprime accents pour comparaison souple."""
        import unicodedata
        return unicodedata.normalize('NFD', str(text).strip().lower()).encode(
            'ascii', 'ignore'
        ).decode('ascii')

    def _find_col(self, headers_norm, key):
        """Retourne l'index de la première variante reconnue, ou -1."""
        aliases_norm = [self._normalize(a) for a in self._COL_ALIASES[key]]
        for i, h in enumerate(headers_norm):
            if h in aliases_norm:
                return i
        return -1

    def action_load_file(self):
        """Lit le fichier Excel et pré-visualise les lignes avant confirmation."""
        self.ensure_one()

        if not self.file:
            raise UserError(_("Veuillez charger un fichier Excel."))

        self.line_ids.unlink()

        decoded = base64.b64decode(self.file)
        # data_only=True lit les valeurs calculées plutôt que les formules
        wb = load_workbook(io.BytesIO(decoded), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise UserError(_("Le fichier Excel est vide."))

        # Normalise les en-têtes (minuscules + sans accents)
        headers_norm = [self._normalize(h) if h else '' for h in rows[0]]

        idx_code = self._find_col(headers_norm, 'product_code')
        idx_min  = self._find_col(headers_norm, 'qty_min')
        idx_max  = self._find_col(headers_norm, 'qty_max')

        missing = []
        if idx_code == -1: missing.append('Code Article  (ex: "Code Article" ou "product_code")')
        if idx_min  == -1: missing.append('Quantité Min  (ex: "Quantité min" ou "qty_min")')
        if idx_max  == -1: missing.append('Quantité Max  (ex: "Quantité max" ou "qty_max")')
        if missing:
            found_cols = ', '.join(f'"{h}"' for h in rows[0] if h)
            raise UserError(
                _("Colonnes introuvables dans le fichier :\n- %s\n\n"
                  "Colonnes détectées dans votre fichier : %s")
                % ('\n- '.join(missing), found_cols)
            )

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id],
        ))

        lines_vals = []
        for row in rows[1:]:
            if not any(row):
                continue
            code = self._clean_code(row[idx_code])
            if not code:
                continue
            qty_min = self._to_float(row[idx_min])
            qty_max = self._to_float(row[idx_max])

            product = env['product.product'].search(
                [('default_code', '=', code)], limit=1
            )

            # Informations complémentaires pour l'aperçu
            current_min = 0.0
            current_max = 0.0
            qty_on_hand = 0.0
            rule_exists = False
            action_label = 'Produit introuvable'

            if product:
                # Stock actuel dans l'entrepôt sélectionné
                location = self.warehouse_id.lot_stock_id
                if location:
                    quant = env['stock.quant'].search([
                        ('product_id', '=', product.id),
                        ('location_id', 'child_of', location.id),
                    ])
                    qty_on_hand = sum(quant.mapped('quantity'))
                else:
                    qty_on_hand = product.qty_available

                # Règle de réassort existante
                orderpoint = env['stock.warehouse.orderpoint'].search([
                    ('product_id', '=', product.id),
                    ('warehouse_id', '=', self.warehouse_id.id),
                ], limit=1)

                if orderpoint:
                    rule_exists  = True
                    current_min  = orderpoint.product_min_qty
                    current_max  = orderpoint.product_max_qty
                    action_label = 'Mise à jour'
                else:
                    action_label = 'Création'

            lines_vals.append((0, 0, {
                'product_code': code,
                'product_id':   product.id if product else False,
                'qty_min':      qty_min,
                'qty_max':      qty_max,
                'found':        bool(product),
                'rule_exists':  rule_exists,
                'current_min':  current_min,
                'current_max':  current_max,
                'qty_on_hand':  qty_on_hand,
                'action_label': action_label,
            }))

        if not lines_vals:
            raise UserError(_("Aucune ligne de données trouvée dans le fichier."))

        self.write({'line_ids': lines_vals, 'state': 'loaded'})

        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    def action_confirm(self):
        """Crée ou met à jour les règles de réassort pour les produits trouvés."""
        self.ensure_one()

        env = self.env(context=dict(
            self.env.context,
            allowed_company_ids=[self.company_id.id],
        ))

        lines_ok    = self.line_ids.filtered(lambda l: l.found)
        lines_ko    = self.line_ids.filtered(lambda l: not l.found)
        count_ok    = len(lines_ok)
        count_ko    = len(lines_ko)
        count_cree  = 0
        count_maj   = 0

        for line in lines_ok:
            product = line.product_id.with_company(self.company_id)
            orderpoint = env['stock.warehouse.orderpoint'].search([
                ('product_id',  '=', product.id),
                ('warehouse_id', '=', self.warehouse_id.id),
            ], limit=1)

            values = {
                'product_min_qty': line.qty_min,
                'product_max_qty': line.qty_max,
            }

            if orderpoint:
                orderpoint.write(values)
                count_maj += 1
            else:
                # Détermine l'emplacement : celui du wizard ou le stock principal de l'entrepôt
                location = self.location_id or self.warehouse_id.lot_stock_id
                if not location:
                    raise UserError(
                        _("Impossible de créer la règle pour '%s' : "
                          "aucun emplacement trouvé. "
                          "Veuillez renseigner le champ Emplacement.")
                        % line.product_code
                    )
                values.update({
                    'product_id':   product.id,
                    'location_id':  location.id,
                    'company_id':   self.company_id.id,
                    'warehouse_id': self.warehouse_id.id,
                })
                env['stock.warehouse.orderpoint'].create(values)
                count_cree += 1

        self.state = 'done'

        msg = (
            f"Import terminé : {count_maj} règle(s) mise(s) à jour, "
            f"{count_cree} règle(s) créée(s)."
        )
        if count_ko:
            msg += f"\n⚠ {count_ko} code(s) article introuvable(s) — ignoré(s)."

        return {
            'type': 'ir.actions.client',
            'tag':  'display_notification',
            'params': {
                'title':   'Règles de réassort — %s ligne(s) traitée(s)' % count_ok,
                'message': msg,
                'type':    'success' if not count_ko else 'warning',
                'sticky':  bool(count_ko),
            },
        }


class OrderpointImportWizardLine(models.TransientModel):
    _name = 'orderpoint.import.wizard.line'
    _description = "Ligne Import Règles de Réassort"

    wizard_id    = fields.Many2one('orderpoint.import.wizard', ondelete='cascade')
    product_code = fields.Char(string='Code Article')
    product_id   = fields.Many2one('product.product', string='Produit', readonly=True)
    product_name = fields.Char(related='product_id.name', string='Désignation', readonly=True)

    # Valeurs à importer
    qty_min      = fields.Float(string='Nv Min', readonly=True)
    qty_max      = fields.Float(string='Nv Max', readonly=True)

    # Informations actuelles (avant import)
    rule_exists  = fields.Boolean(string='Règle existante', readonly=True)
    current_min  = fields.Float(string='Min actuel', readonly=True)
    current_max  = fields.Float(string='Max actuel', readonly=True)
    qty_on_hand  = fields.Float(string='Stock actuel', readonly=True)
    action_label = fields.Char(string='Action', readonly=True)

    found        = fields.Boolean(string='Trouvé', readonly=True)
